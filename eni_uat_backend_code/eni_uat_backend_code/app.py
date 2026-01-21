from flask import Flask, request, jsonify, send_file
import msal
import pyodbc
import os
import math
import io
import base64
from datetime import datetime
import pandas as pd
from azure.ai.formrecognizer import DocumentAnalysisClient
from azure.core.credentials import AzureKeyCredential
from PIL import Image, ImageDraw, ImageFont
import logging
import re 
from flask_cors import CORS
import azure.functions as func
import tempfile
import cv2
import numpy as np
from ultralytics import YOLO
import json
from typing import List, Tuple, Dict, Optional
from easyocr import Reader
from collections import defaultdict
from shapely.geometry import Polygon, LineString
from neo4j import GraphDatabase
from pathlib import Path
from flask_cors import cross_origin
import requests
import snowflake.connector
from dotenv import load_dotenv
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import serialization
from apscheduler.schedulers.blocking import BlockingScheduler
from apscheduler.schedulers.background import BackgroundScheduler
import pytz
from concurrent.futures import ThreadPoolExecutor, as_completed
from shapely.geometry import Point
from PIL import ImageEnhance
import fitz
from PIL import ImageOps
import sys
import time
import threading
import mimetypes
from io import BytesIO
from azure.storage.blob import BlobServiceClient
from opencensus.ext.azure.log_exporter import AzureLogHandler
from typing import Union
from openai import AzureOpenAI
from io import BytesIO
from collections import deque
import uuid


# Load environment variables
load_dotenv()

# Azure SQL Database
CLIENT_ID = os.getenv("CLIENT_ID")
TENANT_ID = os.getenv("TENANT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")
AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"

# Azure Form Recognizer
FORM_RECOGNIZER_ENDPOINT = os.getenv("FORM_RECOGNIZER_ENDPOINT")
FORM_RECOGNIZER_API_KEY = os.getenv("FORM_RECOGNIZER_API_KEY")

# Neo4j connection settings
NEO4J_URI = os.getenv("NEO4J_URI")
NEO4J_USER = os.getenv("NEO4J_USER")
NEO4J_PASSWORD = os.getenv("NEO4J_PASSWORD")

# Open AI Configuration
azure_openai_key = os.getenv("azure_openai_key")
api_base = os.getenv("api_base")
azure_openai_version = os.getenv("azure_openai_version")
deployment_name = os.getenv("deployment_name")

# DIRECT LINE TOKEN
DIRECT_LINE_SECRET = os.getenv("DIRECT_LINE_SECRET")
COPILOT_REGION = os.getenv("COPILOT_REGION")

# Blob Variables
connection_string = os.getenv("connection_string")
container_name = os.getenv("container_name")
success_container_name = os.getenv("success_container_name")
api_url_pid = os.getenv("api_url_pid")
api_url_sd = os.getenv("api_url_sd")


# ----------------------- CLEAN LOGGING SETUP -----------------------

app_insights_key = "7a54c6b0-0cf7-4ca5-a89e-1a29a25d73fd"

# Create root logger
logger = logging.getLogger()
logger.setLevel(logging.INFO)

# Remove all existing handlers (important in Azure App Service)
for handler in logger.handlers[:]:
    logger.removeHandler(handler)

# Standard console handler
console_handler = logging.StreamHandler(sys.stdout)
formatter = logging.Formatter(
    "%(asctime)s - %(levelname)s - %(filename)s:%(lineno)d - %(message)s"
)
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

# Application Insights handler
ai_handler = AzureLogHandler(
    connection_string=f"InstrumentationKey={app_insights_key}"
)
ai_handler.setFormatter(formatter)
logger.addHandler(ai_handler)

# ----------------- SILENCE AZURE & OPENAI INTERNAL LOGS -----------------

# OpenAI SDK internal logs (removes _client.py:1025 POST messages)
logging.getLogger("openai").setLevel(logging.ERROR)

# httpx logs (HTTP layer used by AzureOpenAI)
logging.getLogger("httpx").setLevel(logging.ERROR)

# Azure SDK general noise
logging.getLogger("azure").setLevel(logging.WARNING)

# Azure pipeline verbose logs
logging.getLogger("azure.core.pipeline.policies.http_logging_policy").setLevel(logging.ERROR)

# Snowflake connection noise
logging.getLogger("snowflake.connector.connection").setLevel(logging.WARNING)

# ----------------------- END LOGGING SETUP -----------------------

# Flask app configuration
UPLOAD_FOLDER = 'images'

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

CORS(app,origins=["http://localhost:3000"])
# CORS(app,origins=["https://enifrontendv2-f9bybsd2d8d3cvaw.centralus-01.azurewebsites.net","http://localhost:3000"])
# CORS(app,origins=["https://nl-tst-tag-register-euw-app-c7edeug7eeabf0e3.westeurope-01.azurewebsites.net","http://localhost:3000"])


class LineDetectionOptimizer:
    """Enhanced line detection with improved accuracy and deduplication for small segments."""
    
    def __init__(self, angle_tolerance=3, distance_threshold=22, min_overlap=0.5, gap_threshold=50, min_segment_length=3):
        self.angle_tolerance = angle_tolerance
        self.distance_threshold = distance_threshold
        self.min_overlap = min_overlap
        self.gap_threshold = gap_threshold
        self.min_segment_length = min_segment_length
    
    @staticmethod
    def normalize_line(x1, y1, x2, y2):
        """Normalize line so start point is always top-left."""
        if (y1, x1) > (y2, x2):
            return x2, y2, x1, y1
        return x1, y1, x2, y2
    
    @staticmethod
    def get_line_angle(x1, y1, x2, y2):
        """Get angle in degrees (0-180, always acute)."""
        import numpy as np
        angle = np.arctan2(abs(y2 - y1), abs(x2 - x1)) * 180 / np.pi
        return min(angle, 180 - angle)
    
    @staticmethod
    def point_to_line_distance(px, py, x1, y1, x2, y2):
        """Calculate perpendicular distance from point to line."""
        import numpy as np
        num = abs((y2 - y1) * px - (x2 - x1) * py + x2 * y1 - y2 * x1)
        den = np.sqrt((y2 - y1)**2 + (x2 - x1)**2)
        return num / (den + 1e-6)
    
    def lines_are_collinear(self, x1a, y1a, x2a, y2a, x1b, y1b, x2b, y2b):
        """Check if two lines are collinear (same line, possibly different segments)."""
        import math
        
        # Check segment length
        len_a = math.sqrt((x2a - x1a)**2 + (y2a - y1a)**2)
        len_b = math.sqrt((x2b - x1b)**2 + (y2b - y1b)**2)
        if len_a < self.min_segment_length or len_b < self.min_segment_length:
            return False
        
        angle_a = self.get_line_angle(x1a, y1a, x2a, y2a)
        angle_b = self.get_line_angle(x1b, y1b, x2b, y2b)
        
        angle_diff = abs(angle_a - angle_b)
        if angle_diff > self.angle_tolerance:
            return False
        
        d1 = self.point_to_line_distance(x1b, y1b, x1a, y1a, x2a, y2a)
        d2 = self.point_to_line_distance(x2b, y2b, x1a, y1a, x2a, y2a)
        if not (d1 <= self.distance_threshold and d2 <= self.distance_threshold):
            return False
        
        # Calculate gap with projection
        def calc_gap(x1a,y1a,x2a,y2a, x1b,y1b,x2b,y2b):
            len_a = math.sqrt((x2a - x1a)**2 + (y2a - y1a)**2)
            len_b = math.sqrt((x2b - x1b)**2 + (y2b - y1b)**2)
            
            if len_b > len_a:
                x1a,y1a,x2a,y2a,x1b,y1b,x2b,y2b = x1b,y1b,x2b,y2b,x1a,y1a,x2a,y2a
            
            dx = x2a - x1a
            dy = y2a - y1a
            norm = math.sqrt(dx**2 + dy**2) + 1e-6
            
            def proj(x, y):
                return ((x - x1a) * dx + (y - y1a) * dy) / norm
            
            p1 = proj(x1a, y1a)
            p2 = proj(x2a, y2a)
            p3 = proj(x1b, y1b)
            p4 = proj(x2b, y2b)
            
            seg1_start, seg1_end = min(p1, p2), max(p1, p2)
            seg2_start, seg2_end = min(p3, p4), max(p3, p4)
            
            if seg1_end >= seg2_start and seg2_end >= seg1_start:
                return 0  # overlap or touching
            
            if seg2_start > seg1_end:
                gap = seg2_start - seg1_end
            else:
                gap = seg1_start - seg2_end
            
            gap_pixels = abs(gap) * norm / max(1, abs(p2 - p1))
            return gap_pixels
        
        gap = calc_gap(x1a,y1a,x2a,y2a, x1b,y1b,x2b,y2b)
        if gap > self.gap_threshold:
            return False
        
        return True
    
    def remove_duplicate_lines(self, lines):
        """Remove duplicate and opposite-direction lines, preserving small segments."""
        import numpy as np
        if lines is None or len(lines) == 0:
            return None
        
        unique_lines = []
        used = [False] * len(lines)
        
        for i in range(len(lines)):
            if used[i]:
                continue
            
            x1, y1, x2, y2 = self.normalize_line(*lines[i][:4])
            group = [np.array([x1, y1, x2, y2])]
            used[i] = True
            to_check = [0]
            
            while to_check:
                curr_idx = to_check.pop()
                curr_line = group[curr_idx]
                x1a, y1a, x2a, y2a = curr_line
                
                for j in range(len(lines)):
                    if not used[j]:
                        x1b, y1b, x2b, y2b = self.normalize_line(*lines[j][:4])
                        
                        if self.lines_are_collinear(x1a, y1a, x2a, y2a, x1b, y1b, x2b, y2b):
                            used[j] = True
                            group.append(np.array([x1b, y1b, x2b, y2b]))
                            to_check.append(len(group) - 1)
            
            if group:
                all_points = []
                for line in group:
                    all_points.append((line[0], line[1]))
                    all_points.append((line[2], line[3]))
                
                first_line = group[0]
                angle = self.get_line_angle(first_line[0], first_line[1], first_line[2], first_line[3])
                
                if angle < 45:  # horizontal
                    all_x = [p[0] for p in all_points]
                    all_y = [p[1] for p in all_points]
                    merged_x1 = min(all_x)
                    merged_x2 = max(all_x)
                    merged_y1 = merged_y2 = sum(all_y) / len(all_y) if all_y else 0
                else:  # vertical
                    all_x = [p[0] for p in all_points]
                    all_y = [p[1] for p in all_points]
                    merged_y1 = min(all_y)
                    merged_y2 = max(all_y)
                    merged_x1 = merged_x2 = sum(all_x) / len(all_x) if all_x else 0
                
                length = np.sqrt((merged_x2 - merged_x1)**2 + (merged_y2 - merged_y1)**2)
                if length >= self.min_segment_length:
                    unique_lines.append([merged_x1, merged_y1, merged_x2, merged_y2])
        
        return np.array(unique_lines).reshape(-1, 1, 4) if unique_lines else None
    
    def normalize_line_direction(self, lines, reference_direction=None):
        """Ensure all lines follow consistent direction (left-to-right, top-to-bottom)."""
        import numpy as np
        if lines is None or len(lines) == 0:
            return lines
        
        normalized = []
        for line in lines:
            x1, y1, x2, y2 = line[0] if len(line.shape) > 1 else line
            
            dx = abs(x2 - x1)
            dy = abs(y2 - y1)
            if dx > dy:  # horizontal
                if x2 < x1:
                    x1, y1, x2, y2 = x2, y2, x1, y1
            else:  # vertical
                if y2 < y1:
                    x1, y1, x2, y2 = x2, y2, x1, y1
            
            normalized.append([x1, y1, x2, y2])
        
        return np.array(normalized).reshape(-1, 1, 4) if normalized else None

class ImprovedLineConnectionManager:
    """Improved connection of lines to boxes, ensuring endpoints snap exactly to edges."""

    def __init__(self, tolerance=0.035, endpoint_extension=0.08, max_distance=0.18):
        self.tolerance = tolerance
        self.endpoint_extension = endpoint_extension
        self.max_distance = max_distance

    @staticmethod
    def calculate_distance(x1, y1, x2, y2):
        """Calculate Euclidean distance."""
        import numpy as np
        return np.sqrt((x2 - x1)**2 + (y2 - y1)**2)

    @staticmethod
    def point_on_box_edge(x, y, box, tolerance=0.035):
        """Check if point lies exactly on or near a box edge."""
        box_x1, box_y1 = box['topX'], box['topY']
        box_x2, box_y2 = box['bottomX'], box['bottomY']
        
        # Check if point is on or near any box edge (not inside the box)
        is_near_left = abs(x - box_x1) <= tolerance and box_y1 <= y <= box_y2
        is_near_right = abs(x - box_x2) <= tolerance and box_y1 <= y <= box_y2
        is_near_top = abs(y - box_y1) <= tolerance and box_x1 <= x <= box_x2
        is_near_bottom = abs(y - box_y2) <= tolerance and box_x1 <= x <= box_x2
        
        return is_near_left or is_near_right or is_near_top or is_near_bottom

    def snap_to_box_edge(self, endpoint_x, endpoint_y, box, direction_x, direction_y, max_extension):
        """Snap endpoint to box edge along the line direction."""
        import numpy as np
        xmin, ymin = box['topX'], box['topY']
        xmax, ymax = box['bottomX'], box['bottomY']
        
        # Box edges
        edges = [
            (xmin, ymin, xmin, ymax),  # left
            (xmax, ymin, xmax, ymax),  # right  
            (xmin, ymin, xmax, ymin),  # top
            (xmin, ymax, xmax, ymax)   # bottom
        ]
        
        best_intersection = None
        min_distance = float('inf')
        
        for ex1, ey1, ex2, ey2 in edges:
            # Line equation: P = endpoint + t * direction
            # Edge equation: point between (ex1,ey1) and (ex2,ey2)
            denom = (direction_x * (ey2 - ey1)) - (direction_y * (ex2 - ex1))
            
            if abs(denom) < 1e-6:  # parallel
                continue
                
            t = ((ex1 - endpoint_x) * (ey2 - ey1) - (ey1 - endpoint_y) * (ex2 - ex1)) / denom
            s = ((ex1 - endpoint_x) * direction_y - (ey1 - endpoint_y) * direction_x) / denom
            
            # t >= 0 (forward direction), s in [0,1] (on edge)
            if t >= 0 and 0 <= s <= 1:
                ix = endpoint_x + t * direction_x
                iy = endpoint_y + t * direction_y
                
                # Verify intersection is on edge
                if min(ex1, ex2) <= ix <= max(ex1, ex2) and min(ey1, ey2) <= iy <= max(ey1, ey2):
                    distance = self.calculate_distance(endpoint_x, endpoint_y, ix, iy)
                    if distance <= max_extension and distance < min_distance:
                        min_distance = distance
                        best_intersection = (ix, iy)
        
        return best_intersection if best_intersection else (endpoint_x, endpoint_y)

    def extend_line_to_box(self, other_x, other_y, endpoint_x, endpoint_y, box):
        """Snap line endpoint to the nearest box edge using line direction."""
        import numpy as np
        
        # Calculate line direction from other endpoint to this endpoint
        dx = endpoint_x - other_x
        dy = endpoint_y - other_y
        length = np.sqrt(dx**2 + dy**2)
        
        if length < 1e-6:  # degenerate case
            return endpoint_x, endpoint_y
        
        # Normalize direction
        direction_x = dx / length
        direction_y = dy / length
        
        # Try to snap to box edge along direction
        snapped_x, snapped_y = self.snap_to_box_edge(
            endpoint_x, endpoint_y, box, direction_x, direction_y, self.endpoint_extension
        )
        
        # Validate snapped point is on box edge
        if self.point_on_box_edge(snapped_x, snapped_y, box, self.tolerance):
            return snapped_x, snapped_y
        
        # Fallback: project to closest box edge without extending beyond box
        xmin, ymin = box['topX'], box['topY']
        xmax, ymax = box['bottomX'], box['bottomY']
        
        # Find closest point on box boundary (edge only, not inside)
        candidates = []
        
        # Left edge
        proj_y = endpoint_y
        if ymin <= proj_y <= ymax:
            candidates.append((xmin, proj_y))
        
        # Right edge  
        proj_y = endpoint_y
        if ymin <= proj_y <= ymax:
            candidates.append((xmax, proj_y))
        
        # Top edge
        proj_x = endpoint_x
        if xmin <= proj_x <= xmax:
            candidates.append((proj_x, ymin))
        
        # Bottom edge
        proj_x = endpoint_x
        if xmin <= proj_x <= xmax:
            candidates.append((proj_x, ymax))
        
        if candidates:
            # Choose candidate in direction of line
            best_candidate = None
            min_dist = float('inf')
            
            for cx, cy in candidates:
                dx_c = cx - endpoint_x
                dy_c = cy - endpoint_y
                dot = direction_x * dx_c + direction_y * dy_c
                
                if dot >= 0:  # same direction or on edge
                    dist = self.calculate_distance(endpoint_x, endpoint_y, cx, cy)
                    if dist <= self.endpoint_extension and dist < min_dist:
                        min_dist = dist
                        best_candidate = (cx, cy)
            
            if best_candidate and self.point_on_box_edge(best_candidate[0], best_candidate[1], box, self.tolerance):
                return best_candidate
        
        # Final fallback: keep original endpoint if it's already on edge
        if self.point_on_box_edge(endpoint_x, endpoint_y, box, self.tolerance):
            return endpoint_x, endpoint_y
        
        return endpoint_x, endpoint_y

    def find_connected_box(self, x, y, boxes):
        """Find the best connected box for a line endpoint."""
        min_distance = float('inf')
        best_box = None

        for box in boxes:
            # Check if endpoint is near box edge (not inside box)
            if self.point_on_box_edge(x, y, box, self.tolerance):
                box_cx = (box['topX'] + box['bottomX']) / 2
                box_cy = (box['topY'] + box['bottomY']) / 2
                dist = self.calculate_distance(x, y, box_cx, box_cy)
                if dist < min_distance:
                    min_distance = dist
                    best_box = box
                continue

            # Check distance to box edges
            box_x1, box_y1 = box['topX'], box['topY']
            box_x2, box_y2 = box['bottomX'], box['bottomY']

            # Distance to each edge
            edge_distances = [
                self.calculate_distance(x, y, box_x1, y),  # left
                self.calculate_distance(x, y, box_x2, y),  # right
                self.calculate_distance(x, y, x, box_y1), # top
                self.calculate_distance(x, y, x, box_y2)  # bottom
            ]
            
            min_edge_dist = min(edge_distances)
            if min_edge_dist < min_distance and min_edge_dist <= self.max_distance:
                min_distance = min_edge_dist
                best_box = box

        return best_box

    def connect_lines_to_boxes(self, lines, boxes):
        """Connect lines to boxes, ensuring endpoints snap exactly to box edges."""
        connections = []

        for line in lines:
            orig_start_x, orig_start_y = line['startX'], line['startY']
            orig_end_x, orig_end_y = line['endX'], line['endY']

            # Find source box for start endpoint
            source_box = self.find_connected_box(orig_start_x, orig_start_y, boxes)
            start_x, start_y = orig_start_x, orig_start_y
            
            if source_box:
                # Try to snap start endpoint to source box edge
                start_x, start_y = self.extend_line_to_box(
                    orig_end_x, orig_end_y, orig_start_x, orig_start_y, source_box
                )

            # Find target box for end endpoint
            target_box = self.find_connected_box(orig_end_x, orig_end_y, boxes)
            end_x, end_y = orig_end_x, orig_end_y
            
            if target_box:
                # Try to snap end endpoint to target box edge
                end_x, end_y = self.extend_line_to_box(
                    orig_start_x, orig_start_y, orig_end_x, orig_end_y, target_box
                )

            # Only create connection if:
            # 1. Both source and target boxes found
            # 2. Different boxes
            # 3. Endpoints are on respective box edges
            # 4. Line follows original broken segment path (not rerouted)
            if (source_box and target_box and source_box['id'] != target_box['id'] and
                self.point_on_box_edge(start_x, start_y, source_box, self.tolerance) and
                self.point_on_box_edge(end_x, end_y, target_box, self.tolerance)):
                
                # Verify line doesn't cross inside boxes (stays on broken path)
                line_length = self.calculate_distance(start_x, start_y, end_x, end_y)
                orig_length = self.calculate_distance(orig_start_x, orig_start_y, orig_end_x, orig_end_y)
                
                # Only accept if line length is reasonable (not much longer than original)
                if line_length <= orig_length * 1.5:  # allow 50% extension max
                    connection = {
                        'source_component': {'id': source_box['id']},
                        'target_component': {'id': target_box['id']},
                        'line_details': {
                            'id': line['id'],
                            'startX': start_x,
                            'startY': start_y,
                            'endX': end_x,
                            'endY': end_y,
                            'original_startX': orig_start_x,
                            'original_startY': orig_start_y,
                            'original_endX': orig_end_x,
                            'original_endY': orig_end_y
                        }
                    }
                    connections.append(connection)

        # Deduplicate connections - prefer connections that follow original line path
        unique_connections = []
        seen_pairs = set()
        
        for conn in connections:
            src_id = conn['source_component']['id']
            tgt_id = conn['target_component']['id']
            pair = tuple(sorted([src_id, tgt_id]))
            
            if pair not in seen_pairs:
                unique_connections.append(conn)
                seen_pairs.add(pair)

        return unique_connections

def improve_line_detection_pipeline(lines_raw, boxes, img_width, img_height):
    """
    Complete improved line detection and connection pipeline.
    Ensures lines follow broken segments and end exactly on box edges.
    """
    
    lines_px = []
    for line in lines_raw:
        x1 = int(line['startX'] * img_width)
        y1 = int(line['startY'] * img_height)
        x2 = int(line['endX'] * img_width)
        y2 = int(line['endY'] * img_height)
        lines_px.append([x1, y1, x2, y2])
    
    import numpy as np
    
    # Step 1: Remove duplicates and merge collinear segments (preserve broken line paths)
    optimizer = LineDetectionOptimizer(
        angle_tolerance=4,      # Stricter angle matching
        distance_threshold=18,  # Distance for collinearity
        min_overlap=0.5,
        gap_threshold=80,       # Merge small gaps in broken lines
        min_segment_length=5    # Detect very small segments
    )
    lines_px = optimizer.remove_duplicate_lines(np.array(lines_px))
    
    # Step 2: Normalize line directions
    lines_px = optimizer.normalize_line_direction(lines_px)
    
    # Step 3: Convert back to normalized coordinates
    cleaned_lines = []
    if lines_px is not None:
        for idx, line in enumerate(lines_px):
            x1, y1, x2, y2 = line[0] if len(line.shape) > 1 else line
            length = np.sqrt((x2 - x1)**2 + (y2 - y1)**2)
            if length >= optimizer.min_segment_length:
                # Preserve original line path information
                cleaned_lines.append({
                    'id': f"L-{idx}",
                    'startX': x1 / img_width,
                    'startY': y1 / img_height,
                    'endX': x2 / img_width,
                    'endY': y2 / img_height,
                    'length': length / np.sqrt(img_width**2 + img_height**2)  # normalized length
                })
    
    # Step 4: Connect lines to boxes - ensure lines follow original broken paths
    connector = ImprovedLineConnectionManager(
        tolerance=0.035,        # Edge detection tolerance
        endpoint_extension=0.08, # Max extension to reach box edge
        max_distance=0.18       # Max distance from line end to box
    )
    connections = connector.connect_lines_to_boxes(cleaned_lines, boxes)
    
    return cleaned_lines, connections

@app.route('/')
def index():
    return "status 200", 200

def get_snowflake_connection(config_path="config.json"):
    # Load config
    with open(config_path, "r") as f:
        config = json.load(f)["snowflake"]

    # Load private key
    with open(config["private_key_path"], "rb") as key:
        p_key = serialization.load_pem_private_key(
            key.read(),
            password=None,
            backend=default_backend()
        )

    # Convert to DER format for connector
    pkb = p_key.private_bytes(
        encoding=serialization.Encoding.DER,
        format=serialization.PrivateFormat.PKCS8,
        encryption_algorithm=serialization.NoEncryption()
    )

    # Connect to Snowflake using key-pair authentication
    conn = snowflake.connector.connect(
        user=config["user"],
        account=config["account"],
        warehouse=config["warehouse"],
        database=config["database"],
        schema=config["schema"],
        role=config["role"],
        private_key=pkb
    )
    return conn

def convert_row_to_taginfo_model(cursor, row):
    """Converts a pyodbc row to a dictionary mimicking TagInfoModel."""
    # TagInfoModel mapping: [TAG ID] -> Tag_ID, [File Name] -> File_Name, etc.
    taginfo_mapping = {
        "TAG_NO": "tagNo",
        "FILE_NAME": "file_Name",
        "URL": "url",
        "FILE_TYPE": "file_Type",
        "FILE_DESCRIPTION": "file_Description",
        "CREATED_DATE": "created_Date"
    }
    row_dict = convert_row_to_dict_with_mapping(cursor, row, taginfo_mapping)
    if not row_dict:
        return None
    # Apply date formatting if the key exists and value is not None
    created_date = row_dict.get("Created_Date")
    if created_date and isinstance(created_date, datetime):
        row_dict["Created_Date"] = created_date.strftime("%m/%d/%Y %H:%M:%S") 
    else:
        row_dict["Created_Date"] = "" 

    # Convert other values to string
    for key, value in row_dict.items():
        if key != 'Created_Date': 
            if value is not None:
                row_dict[key] = str(value)
            else:
                row_dict[key] = ""

    return row_dict

def convert_row_to_dict_with_mapping(cursor, row, mapping):
    """Converts a pyodbc row to a dictionary using a provided mapping."""
    if row is None:
        return None
    
    # Get column description from the cursor (column names and metadata)
    description = cursor.description
    row_dict = {}

    # Loop through each column in the row
    for i, col in enumerate(description):
        sql_col_name = col[0]

        # Map the SQL column name to the model property name
        model_prop_name = mapping.get(sql_col_name, sql_col_name)
        value = row[i]
        row_dict[model_prop_name] = value

    return row_dict

# --- pdf to image Utilities ---

def enhance_image(img, contrast_factor=1.8, brightness_factor=1.2):
    enhancer_contrast = ImageEnhance.Contrast(img)
    img = enhancer_contrast.enhance(contrast_factor)
    enhancer_brightness = ImageEnhance.Brightness(img)
    img = enhancer_brightness.enhance(brightness_factor)
    return img

def pdf_to_images(pdf_path, output_dir="images", zoom_factor=3.0, contrast=1.8, brightness=1.2):
    try:
        os.makedirs(output_dir, exist_ok=True)
        image_paths = []
        doc = fitz.open(pdf_path)
        for page_num in range(doc.page_count):  #All Pages
        # for page_num in range(1):  # Only first page
            page = doc[page_num]
            mat = fitz.Matrix(zoom_factor, zoom_factor)
            pix = page.get_pixmap(matrix=mat)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            img = enhance_image(img, contrast, brightness)
            base_name = Path(pdf_path).stem
            image_name = f"{base_name}_page{page_num+1}.png"
            image_path = os.path.join(output_dir, image_name)
            img.save(image_path, "PNG", quality=100)
            image_paths.append(image_path)
            logger.info(f"Saved PDF page {page_num+1} as image: {image_path}")
        doc.close()
        return image_paths
    except Exception as e:
        logger.error(f"PDF conversion failed for {pdf_path}: {e}", exc_info=True)
        raise

#------------------------ pdf to image utilities end ------------------------

TAG_PATTERNS = {
    # Line Numbers
    "Line": re.compile(
        r'^[A-Z0-9-]+_[0-9]+"?\s?[A-Z]*-\d+-\d+$'          
    ),

    # Instrument Tags (Enhanced)
    "Instrument": re.compile(
    r'^[A-Z0-9-]+_[A-Z]{2,}-\d+(-[A-Z0-9]+)?$'
),

    # Equipment Tags
    "Equipment": re.compile(
       r'^[A-Z0-9-]+_[A-Z]-\d+(-[A-Z0-9]+)?$'
    ),

    # Cables
    "Cable": re.compile(
        r'^[A-Z0-9-]+_(CBL|CABLE|CAB|CB)-?\d+'
    ),
}

def classify_tag_type(tag_no: str, file_type: str) -> str:
    tag_no = tag_no.strip().upper()
    file_type = file_type.strip().upper()
    print("TagNo and FileType:", tag_no, file_type)

    # Detect only true invalid / system words
    invalid_patterns = [
        r'^[A-Z0-9-]+_REV\d+$',              
        r'^[A-Z0-9-]+_LOCALINSTRUMENT$',    
    ]
    if any(re.match(p, tag_no) for p in invalid_patterns):
        return "Unknown"

    # --- Rule 1: File-type special cases ---
    if file_type.startswith("A7001"):  
        func_code_match = re.search(r'_([A-Z/]+)-', tag_no)
        if func_code_match:
            func_code = func_code_match.group(1)
            return "Instrument" if len(func_code) >= 2 else "Equipment"
        return "Instrument"

    elif file_type.startswith("A6"):  
        return "Cable"

    elif file_type.startswith("A7"):   
        if TAG_PATTERNS["Cable"].search(tag_no):
            return "Cable"
        return "Instrument"

    # Rule 2: Pattern-based classification
    if TAG_PATTERNS["Line"].search(tag_no):
        return "Line"
    elif TAG_PATTERNS["Cable"].search(tag_no):
        return "Cable"
    elif TAG_PATTERNS["Instrument"].search(tag_no):
        return "Instrument"
    elif TAG_PATTERNS["Equipment"].search(tag_no):
        return "Equipment"

    # --- Rule 3: File-type fallback rules ---
    if file_type.startswith(("A4", "B5")): 
        func_code_match = re.search(r'_([A-Z/]+)-', tag_no)
        if func_code_match:
            func_code = func_code_match.group(1)
            return "Instrument" if len(func_code) >= 2 else "Equipment"
        return "Line"

    elif file_type.startswith("A42"):  
        return "Equipment"

    elif file_type.startswith("A3"):  
        return "Line"


    return "Unknown"

def insert_extracted_data_to_database(contents, original_filepath):
    if not contents:
        logger.info("No content extracted, skipping database insert.")
        return

    file_type_description_map = {
        'A1': 'Jacket Structural',
        'A2': 'Deck Structural',
        'A3': 'Supports',
        'A4': 'General Arrangement',
        'A5': 'Piping Layout',
        'A6': 'Electrical Drawing',
        'A7': 'Instrumentation Drawing',
        'A8': 'Safety Drawing',
        'B5': 'Isometrics',
        'A42': 'Plotplans',
        'A7001': 'C&E Drawing'
    }

    def is_valid_plant_tag(tag):
        if all(part.isdigit() for part in tag.split('-')):
            return False
        if tag.replace('-', '').isalpha() or tag.upper() in {'BY-PASS', 'BYPASS', 'PASS', 'NO'}:
            return False
        if not (any(c.isalpha() for c in tag) and any(c.isdigit() for c in tag)):
            return False
        if tag.upper().startswith(('E17', 'L7')):
            return False
        return True

    try:
        logger.info(f"Starting database insert for {original_filepath}")
        conn = get_snowflake_connection()
        with conn.cursor() as cursor:
            filename = os.path.basename(original_filepath)
            parts = filename.split("_")
            document_id = "_".join(parts[:2])
            logger.info(f"DocumentID: {document_id}")

            # Detect File Type
            file_type_match = re.search(r'(A[1-8]\d*|B5|A42|KA45\d*)', filename, re.IGNORECASE)
            full_file_type = file_type_match.group(0).upper() if file_type_match else None

            if full_file_type and (
                full_file_type.startswith("A40") or
                full_file_type.startswith("A41") or
                full_file_type.startswith("KA45")
            ):
                logger.info(f"Skipping P&ID file (A40/A41/KA45): {filename}")
                return

            if full_file_type is None:
                file_type = "Unknown"
                file_description = "Uncategorized Drawing"
            elif full_file_type.startswith("A7001"):
                file_type = "A7001"
                file_description = file_type_description_map.get("A7001", "C&E Drawing")
            elif full_file_type.startswith("A42"):
                file_type = "A42"
                file_description = file_type_description_map.get("A42", "Plotplans")
            elif full_file_type.startswith("A7"):
                file_type = "A7"
                file_description = file_type_description_map.get("A7", "Instrumentation Drawing")
            elif full_file_type.startswith("A"):
                file_type = full_file_type[:2]
                file_description = file_type_description_map.get(file_type, "Unknown")
            else:
                file_type = full_file_type
                file_description = file_type_description_map.get(file_type, "Unknown")

            pid_value = filename
            file_prefix = parts[0]
            query_base = """
                INSERT INTO ENI_NL_DEV_TAGREGISTRY.DEV.TBL_TAG_INFO
                ("TAG_NO", "FILE_NAME", "URL", "FILE_TYPE", "CREATED_DATE", "FILE_DESCRIPTION", "TAGTYPE")
                VALUES {}
            """

            pattern_regular = re.compile(
                r'^[A-Z&]{1,3}-\d{1,5}[A-Z]?(?:-\d+[A-Z]?)?(?:-[A-Z0-9]+)?(?:/[A-Z]+)?$'
            )
            pattern_pipe = re.compile(
                r'^(\d{1,2}(?:/\d{1,2})?"-)[A-Z0-9]{2,6}(?:-[A-Z0-9]{2,6}){2,4}(?:-[A-Z]{1,2})?$'
            )

            normalized_contents = [tag.strip("()") for tag in contents]
            unique_contents = list(dict.fromkeys(normalized_contents))

            filtered_regular = [
                tag for tag in unique_contents
                if pattern_regular.match(tag) and is_valid_plant_tag(tag)
                
            ]
            
            # ✅ Expand tags containing slash "/" (applies only to filtered_regular)
            expanded_regular = []
            for tag in filtered_regular:
                if "/" in tag:
                    # Example: E17-A_F-1158A/B
                    base_part, suffix_part = tag.split("/", 1)
                    # Split into the part before slash and expand both variants
                    match = re.match(r"(.+?)([A-Z]+)$", base_part)
                    if match:
                        prefix, last_letters = match.groups()
                        expanded_regular.append(f"{prefix}{last_letters}")  
                        expanded_regular.append(f"{prefix}{suffix_part}")  
                    else:
                        # Fallback if it doesn't fit the expected pattern
                        expanded_regular.append(base_part)
                        expanded_regular.append(suffix_part)
                else:
                    expanded_regular.append(tag)

            # Replace filtered_regular with expanded version
            filtered_regular = expanded_regular
            
            remaining_for_pipe = [tag for tag in unique_contents if tag not in filtered_regular]
            filtered_pipe = [
                tag for tag in remaining_for_pipe
                if pattern_pipe.match(tag) and is_valid_plant_tag(tag)
            ]
            all_filtered_tags = filtered_regular + filtered_pipe

            if all_filtered_tags:
                cursor.execute("""
                    SELECT "TAG_NO" FROM ENI_NL_DEV_TAGREGISTRY.DEV.TBL_TAG_INFO
                    WHERE "FILE_NAME" = %s
                """, (pid_value,))
                existing_tags = {row[0] for row in cursor.fetchall()}

                new_tags = []
                for tag in all_filtered_tags:
                    modified_tag = f"{file_prefix}_{tag}"
                    if modified_tag not in existing_tags:
                        new_tags.append(tag)
                    else:
                        logger.debug(f"Skipping duplicate tag: {modified_tag}")

                all_filtered_tags = new_tags

            if not all_filtered_tags:
                logger.info(f"No new unique tags to insert for {filename}")
                return

            batch_size = 100
            insert_values = []
            for tag in all_filtered_tags:
                tag = re.sub(r'(\d+)([A-Z])($|-)', r'\1-\2\3', tag)
                modified_tag = f"{file_prefix}_{tag}"
                logger.debug("Classify input -> tag: %s, file: %s", modified_tag, pid_value)
                tag_type = classify_tag_type(modified_tag, file_type)
                logger.debug("Classify result -> %s -> %s", modified_tag, tag_type)
                
                # modified_tag = f"{file_prefix}_{tag}"
                # logger.debug("Classify input -> tag: %s, file: %s", modified_tag, pid_value)
                # tag = re.sub(r'(\d+)([A-Z])($|-)', r'\1-\2\3', tag)
                # tag_type = classify_tag_type(modified_tag, file_type)
                # logger.debug("Classify result -> %s -> %s", modified_tag, tag_type)

                logger.info(f"FileType: {file_type} | TagNo: {modified_tag} | TagType: {tag_type}")
                insert_values.append((
                    modified_tag,  
                    pid_value,     
                    f"https://eni-dev.pimshosting.com/dcs-documents-details?Domain=NLAsset&DocID={document_id}",
                    file_type,     
                    datetime.now(),
                    file_description, 
                    tag_type       
                ))

            for i in range(0, len(insert_values), batch_size):
                batch = insert_values[i:i + batch_size]
                placeholders = ', '.join(['(%s, %s, %s, %s, %s, %s, %s)'] * len(batch))
                query = query_base.format(placeholders)
                flat_values = tuple(item for sublist in batch for item in sublist)
                cursor.execute(query, flat_values)

            logger.info(f"✅ Inserted {len(insert_values)} new tags for {filename}")
    except Exception as ex:
        logger.error(f"Database insert error: {ex}", exc_info=True)

MAX_WORKERS = 10  

def process_single_page(args):
    """
    Process one page: OCR + annotation.
    Returns words and path to annotated image.
    """
    client, image_path, page_idx, original_filepath, file_ext, upload_folder = args

    try:
        with open(image_path, "rb") as f:
            poller = client.begin_analyze_document("prebuilt-read", f)
            result = poller.result()

        # Extract words that have bounding polygons
        words_with_poly = [
            word for page in result.pages
            for word in page.words
            if hasattr(word, 'polygon') and word.polygon
        ]
        words_text = [word.content for word in words_with_poly]

        # Annotate the image
        with Image.open(image_path) as img:
            draw = ImageDraw.Draw(img)

            for word in words_with_poly:
                poly = [(p.x, p.y) for p in word.polygon]
                draw.polygon(poly, outline="red", width=3)

                x_min = int(min(p.x for p in word.polygon))
                y_min = int(min(p.y for p in word.polygon))
                draw.text((x_min, y_min - 12), word.content, fill="red")

            stem = Path(original_filepath).stem
            if file_ext.lower() == ".pdf":
                annotated_name = f"annotated_{stem}_page{page_idx + 1}.png"
            else:
                annotated_name = f"annotated_{Path(original_filepath).name}"

            annotated_path = Path(upload_folder) / annotated_name
            img.save(annotated_path, "PNG")

        return {
            "success": True,
            "words": words_text,
            "annotated_path": str(annotated_path)
        }

    except Exception as ex:
        logger.error(
            f"OCR/annotation failed on page {page_idx + 1} of {original_filepath}: {ex}",
            exc_info=True
        )
        return {
            "success": False,
            "words": [],
            "annotated_path": None
        }

@app.route("/api/Home/upload-files", methods=["POST"])
def upload_files_endpoint():
    logger.info("Starting file upload processing")

    upload_folder = app.config["UPLOAD_FOLDER"]
    Path(upload_folder).mkdir(parents=True, exist_ok=True)

    # Optional: clean old annotated files
    for old_file in Path(upload_folder).glob("annotated_*.png"):
        try:
            old_file.unlink()
        except:
            pass

    if "files" not in request.files:
        return jsonify({"message": "No file part"}), 400

    files = request.files.getlist("files")
    if not files or all(f.filename == "" for f in files):
        return jsonify({"message": "No files selected"}), 400

    uploaded_paths = []
    succeeded_files = []
    failed_files = []
    seen = set()

    # === 1. Save and validate uploaded files ===
    for file in files:
        if not file or not file.filename:
            continue

        filename = file.filename
        if filename in seen:
            continue
        seen.add(filename)

        ext = Path(filename).suffix.lower()
        if ext not in {".pdf", ".png", ".jpg", ".jpeg"}:
            failed_files.append({"fileName": filename, "reason": "Invalid file type"})
            continue

        # Size validation
        file.seek(0, 2)
        if file.tell() > 200 * 1024 * 1024:  # 200 MB limit
            failed_files.append({"fileName": filename, "reason": "File too large"})
            file.seek(0)
            continue
        file.seek(0)

        filepath = Path(upload_folder) / filename
        try:
            file.save(str(filepath))
            uploaded_paths.append(filepath)
            succeeded_files.append({"fileName": filename})
        except Exception as e:
            failed_files.append({"fileName": filename, "reason": str(e)})

    if not uploaded_paths:
        return jsonify({"message": "No valid files to process"}), 400

    # === 2. Initialize Azure client once ===
    try:
        client = DocumentAnalysisClient(
            endpoint=FORM_RECOGNIZER_ENDPOINT,
            credential=AzureKeyCredential(FORM_RECOGNIZER_API_KEY)
        )
    except Exception as e:
        logger.error(f"Azure client init failed: {e}")
        return jsonify({"message": "OCR service unavailable"}), 500

    all_annotated_paths = []

    # === 3. Process each document with parallel page processing ===
    for original_filepath in uploaded_paths:
        original_filename = original_filepath.name
        file_ext = original_filepath.suffix.lower()

        # Convert PDF to images (or treat image as a single page)
        if file_ext == ".pdf":
            try:
                image_paths = pdf_to_images(
                    str(original_filepath),
                    output_dir=upload_folder
                )
            except Exception:
                logger.error(f"PDF conversion failed: {original_filename}")
                failed_files.append({"fileName": original_filename, "reason": "PDF conversion failed"})
                continue
        else:
            image_paths = [str(original_filepath)]

        # Prepare parallel tasks
        tasks = [
            (client, img_path, idx, str(original_filepath), file_ext, upload_folder)
            for idx, img_path in enumerate(image_paths)
        ]

        document_all_words = []

        # Parallel page OCR + annotation
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = [executor.submit(process_single_page, task) for task in tasks]

            for future in as_completed(futures):
                result = future.result()
                if result["success"]:
                    document_all_words.extend(result["words"])
                    if result["annotated_path"]:
                        all_annotated_paths.append(result["annotated_path"])

        # ONE DB insert per document (prevents duplicates)
        if document_all_words:
            logger.info(
                f"Inserting tags from {original_filename} "
                f"({len(image_paths)} pages, {len(set(document_all_words))} unique words)"
            )
            insert_extracted_data_to_database(document_all_words, str(original_filepath))

    # === 4. Return response ===
    annotated_urls = [f"/uploads/{Path(p).name}" for p in all_annotated_paths]

    return jsonify({
        "message": "Files processed successfully",
        "annotatedImages": annotated_urls,
        "succeededFiles": succeeded_files,
        "failedFiles": failed_files
    }), 200

# Get Tag Info Data Endpoint
@app.route('/api/Home/GetDrawingScannerData', methods=['GET'])
def get_table_info_data_endpoint():
    """Retrieve tags and related details from Snowflake TAGINFO table."""
    try:
        logger.info("Fetching TAGINFO data from Snowflake")
        conn = get_snowflake_connection()
        with conn.cursor() as cursor:
            query = """
                SELECT 
                    "S_NO", 
                    "TAG_NO", 
                    "FILE_NAME", 
                    "URL", 
                    "FILE_TYPE",
                    "TAGTYPE",
                    "FILE_DESCRIPTION", 
                    "CREATED_DATE"
                FROM ENI_NL_DEV_TAGREGISTRY.DEV.TBL_TAG_INFO
                WHERE 
                    (
                        "FILE_TYPE" LIKE 'A%' 
                        AND "FILE_TYPE" != 'A4' 
                        AND "FILE_NAME" NOT ILIKE '%A40%' 
                        AND "FILE_NAME" NOT ILIKE '%A41%' 
                        AND "FILE_NAME" NOT ILIKE '%KA45%'
                    )
                    OR "FILE_TYPE" IN ('B5', 'A42', 'Unknown')
                ORDER BY "CREATED_DATE" DESC;
            """
            cursor.execute(query)
            rows = cursor.fetchall()

            taginfo_list = [convert_row_to_taginfo_model(cursor, row) for row in rows]
            cursor.execute('SELECT COUNT(*) FROM ENI_NL_DEV_TAGREGISTRY.DEV.TBL_TAG_INFO')
            total_count = cursor.fetchone()[0]

            logger.info(f"Retrieved {len(taginfo_list)} records from Snowflake")
            return jsonify({"data": taginfo_list, "totalCount": total_count}), 200

    except Exception as ex:
        logger.error(f"Failed to fetch TAGINFO data from Snowflake: {ex}", exc_info=True)
        return jsonify({"error": "Internal server error", "details": str(ex)}), 500

# Get PID Tag Info Data Endpoint
@app.route('/api/Home/GetPIDData', methods=['GET'])
def get_table_PIDinfo_data_endpoint():
    """Retrieve tags and related details from Snowflake TAGINFO table."""
    try:
        logger.info("Fetching TAGINFO data from Snowflake")
        conn = get_snowflake_connection()
        with conn.cursor() as cursor:
            query = """SELECT 
                    "S_NO",
                    "TAG_NO", 
                    "FILE_NAME", 
                    "URL", 
                    "FILE_TYPE", 
                    "FILE_DESCRIPTION", 
                    "CREATED_DATE",
                    "TAGTYPE"
                FROM ENI_NL_DEV_TAGREGISTRY.DEV.TBL_TAG_INFO
                WHERE 
                    "FILE_TYPE" IN ('A4')
                    AND ("TAGTYPE" != 'Unknown' OR "TAGTYPE" IS NULL)
                ORDER BY "CREATED_DATE" DESC;

            """
            cursor.execute(query)
            rows = cursor.fetchall()
           
            # Convert all rows to dictionaries
            taginfo_list = [convert_row_to_taginfo_model(cursor, row) for row in rows]
           
            cursor.execute('SELECT COUNT(*) FROM ENI_NL_DEV_TAGREGISTRY.DEV.TBL_TAG_INFO')
            total_count = cursor.fetchone()[0]
 
            logger.info(f"Retrieved {len(taginfo_list)} records from Snowflake")
            return jsonify({"data": taginfo_list, "totalCount": total_count}), 200
 
    except Exception as ex:
        logger.error(f"Failed to fetch TAGINFO data from Snowflake: {ex}", exc_info=True)
        return jsonify({"error": "Internal server error", "details": str(ex)}), 500

# Get Image Base64 Endpoint
@app.route('/api/Home/get-image-base64', methods=['GET'])
def get_image_base64_endpoint():
    """Retrieves locally saved annotated images as Base64 strings."""
    logger.info("Retrieving annotated images as base64")
    image_directory = app.config['UPLOAD_FOLDER']
 
    if not os.path.exists(image_directory):
        logger.warning(f"Image directory not found: {image_directory}")
        # Return empty list as C# returns empty list if no data/error
        return jsonify({"images": []}), 200
    try:
        # Find all annotated files in the directory
        image_files = [
            os.path.join(image_directory, f)
            for f in os.listdir(image_directory)
            if os.path.isfile(os.path.join(image_directory, f)) and f.lower().startswith("annotated_")
            and os.path.splitext(f)[1].lower() in ['.jpg', '.jpeg', '.png'] # Filter by common image extensions
        ]
 
        if not image_files:
             logger.info("No annotated image files found.")
             return jsonify({"images": []}), 200
 
        images_base64 = []
        for image_path in image_files:
            try:
                with open(image_path, "rb") as image_file:
                    bytes_content = image_file.read()
                    base64_string = base64.b64encode(bytes_content).decode('utf-8')
                    # Detect mime type based on extension
                    extension = os.path.splitext(image_path)[1].lower()
                    if extension in ['.jpg', '.jpeg']:
                        mime_type = 'image/jpeg'
                    elif extension == '.png':
                        mime_type = 'image/png'
                    else:
                        mime_type = 'application/octet-stream' # Should not happen with filter above
 
                    images_base64.append(f"data:{mime_type};base64,{base64_string}")
 
            except Exception as file_ex:
                logger.error(f"Error reading or encoding image file {image_path}: {file_ex}")
        logger.info(f"Retrieved {len(images_base64)} annotated images")
        return jsonify({"images": images_base64}), 200
 
    except Exception as ex:
        logger.error(f"Failed to get images base64: {ex}", exc_info=True)
        # C# returns "no data" on error (implying 200?), but 500 is more appropriate for internal error.
        return jsonify({"error": "Internal server error", "details": str(ex)}), 500
 
def connect_overlapping_lines_fast(lines, img_width, img_height, angle_threshold=5, distance_threshold=20, gap_threshold=30):
    """
    Fast line connection using spatial indexing and optimized algorithms
    
    Args:
        lines: List of line dictionaries with startX, startY, endX, endY (normalized coords)
        img_width: Image width in pixels
        img_height: Image height in pixels
        angle_threshold: Maximum angle difference in degrees
        distance_threshold: Maximum perpendicular distance
        gap_threshold: Maximum gap to bridge
    
    Returns:
        List of connected line dictionaries
    """
    logger.info(f"Connecting {len(lines)} lines with optimized method")
    if not lines or len(lines) == 0:
        return []
    
    import numpy as np
    from collections import defaultdict
    
    # Convert normalized coords to pixels once
    lines_px = []
    for line in lines:
        x1 = line['startX'] * img_width
        y1 = line['startY'] * img_height
        x2 = line['endX'] * img_width
        y2 = line['endY'] * img_height
        
        # Calculate angle and length
        angle = np.arctan2(y2-y1, x2-x1) * 180 / np.pi
        length = np.sqrt((x2-x1)**2 + (y2-y1)**2)
        
        lines_px.append({
            'x1': x1, 'y1': y1, 'x2': x2, 'y2': y2,
            'angle': angle, 'length': length,
            'cx': (x1+x2)/2, 'cy': (y1+y2)/2  # center point
        })
    
    # Spatial grid for fast neighbor lookup
    grid_size = max(gap_threshold * 2, 50)
    grid = defaultdict(list)
    
    for idx, line in enumerate(lines_px):
        # Add line to grid cells it touches
        min_x = int(min(line['x1'], line['x2']) / grid_size)
        max_x = int(max(line['x1'], line['x2']) / grid_size)
        min_y = int(min(line['y1'], line['y2']) / grid_size)
        max_y = int(max(line['y1'], line['y2']) / grid_size)
        
        for gx in range(min_x, max_x + 1):
            for gy in range(min_y, max_y + 1):
                grid[(gx, gy)].append(idx)
    
    def get_neighbors(line_idx):
        """Get potential neighbor indices using spatial grid"""
        line = lines_px[line_idx]
        neighbors = set()
        
        # Check cells around line endpoints
        for x, y in [(line['x1'], line['y1']), (line['x2'], line['y2'])]:
            gx = int(x / grid_size)
            gy = int(y / grid_size)
            
            # Check 3x3 grid around point
            for dx in [-1, 0, 1]:
                for dy in [-1, 0, 1]:
                    neighbors.update(grid.get((gx+dx, gy+dy), []))
        
        neighbors.discard(line_idx)
        return list(neighbors)
    
    def can_connect_fast(idx1, idx2):
        """Fast connection check"""
        l1 = lines_px[idx1]
        l2 = lines_px[idx2]
        
        # Quick angle check
        angle_diff = abs(l1['angle'] - l2['angle'])
        if angle_diff > 180:
            angle_diff = 360 - angle_diff
        if angle_diff > angle_threshold:
            return False
        
        # Quick distance check using center points
        center_dist = np.sqrt((l1['cx']-l2['cx'])**2 + (l1['cy']-l2['cy'])**2)
        max_possible_dist = (l1['length'] + l2['length']) / 2 + gap_threshold
        if center_dist > max_possible_dist:
            return False
        
        # Check minimum endpoint distance
        endpoints = [
            (l1['x1'], l1['y1']), (l1['x2'], l1['y2']),
            (l2['x1'], l2['y1']), (l2['x2'], l2['y2'])
        ]
        
        min_gap = float('inf')
        for i in range(2):
            for j in range(2, 4):
                gap = np.sqrt((endpoints[i][0]-endpoints[j][0])**2 + 
                             (endpoints[i][1]-endpoints[j][1])**2)
                min_gap = min(min_gap, gap)
        
        if min_gap > gap_threshold:
            return False
        
        # Simplified collinearity check
        # Check if endpoints of one line are close to the other line
        def point_line_dist(px, py, x1, y1, x2, y2):
            dx, dy = x2-x1, y2-y1
            if dx == 0 and dy == 0:
                return np.sqrt((px-x1)**2 + (py-y1)**2)
            t = max(0, min(1, ((px-x1)*dx + (py-y1)*dy) / (dx*dx + dy*dy)))
            return np.sqrt((px-(x1+t*dx))**2 + (py-(y1+t*dy))**2)
        
        d1 = point_line_dist(l2['x1'], l2['y1'], l1['x1'], l1['y1'], l1['x2'], l1['y2'])
        d2 = point_line_dist(l2['x2'], l2['y2'], l1['x1'], l1['y1'], l1['x2'], l1['y2'])
        
        return (d1 + d2) / 2 <= distance_threshold
    
    # Union-Find for fast component detection
    parent = list(range(len(lines_px)))
    
    def find(x):
        if parent[x] != x:
            parent[x] = find(parent[x])
        return parent[x]
    
    def union(x, y):
        px, py = find(x), find(y)
        if px != py:
            parent[px] = py
            return True
        return False
    
    # Build connections using spatial indexing
    for i in range(len(lines_px)):
        neighbors = get_neighbors(i)
        for j in neighbors:
            if i < j and can_connect_fast(i, j):
                union(i, j)
    
    # Group connected components
    components = defaultdict(list)
    for i in range(len(lines_px)):
        components[find(i)].append(i)
    
    # Merge each component into single line
    connected_lines = []
    for comp_indices in components.values():
        if len(comp_indices) == 1:
            # Single line, no merging needed
            idx = comp_indices[0]
            line = lines_px[idx]
            connected_lines.append({
                'startX': line['x1'] / img_width,
                'startY': line['y1'] / img_height,
                'endX': line['x2'] / img_width,
                'endY': line['y2'] / img_height
            })
        else:
            # Merge multiple lines - find extremities
            all_points = []
            for idx in comp_indices:
                line = lines_px[idx]
                all_points.extend([
                    (line['x1'], line['y1']),
                    (line['x2'], line['y2'])
                ])
            
            # Find two farthest points
            max_dist = 0
            best_pair = (all_points[0], all_points[1])
            
            # Sample points to reduce computation
            sample_size = min(len(all_points), 20)
            sampled = all_points[::max(1, len(all_points)//sample_size)]
            
            for i in range(len(sampled)):
                for j in range(i+1, len(sampled)):
                    dist = (sampled[i][0]-sampled[j][0])**2 + (sampled[i][1]-sampled[j][1])**2
                    if dist > max_dist:
                        max_dist = dist
                        best_pair = (sampled[i], sampled[j])
            
            connected_lines.append({
                'startX': best_pair[0][0] / img_width,
                'startY': best_pair[0][1] / img_height,
                'endX': best_pair[1][0] / img_width,
                'endY': best_pair[1][1] / img_height
            })
    
    # Re-assign IDs
    for idx, line in enumerate(connected_lines):
        line['id'] = f"CL-{idx}"
    
    print(f"Line connection: {len(lines)} segments -> {len(connected_lines)} connected lines")
    
    return connected_lines

# ============================================================================
def yolo_model_line_detection(input_image_path, output_dir, model_path="best.pt"):
    """
    Process engineering drawing to detect objects and lines with optimized line detection
    
    Args:
        input_image_path: Path to input image
        output_dir: Directory to save outputs
        model_path: Path to YOLO model file
    
    Returns:
        output_image_path: Path to processed output image
    """
    
    # === Optimized Line Detection Helper Functions ===
    def preprocess_image_optimized(image):
        """Optimized preprocessing - balance of speed and accuracy"""
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
        # CLAHE for contrast enhancement
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        enhanced = clahe.apply(gray)
        
        # Light blur to reduce noise
        blurred = cv2.GaussianBlur(enhanced, (3, 3), 0)
        
        # Combine Otsu and Adaptive thresholding for better line detection
        _, thresh1 = cv2.threshold(blurred, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        thresh2 = cv2.adaptiveThreshold(blurred, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
                                      cv2.THRESH_BINARY_INV, 11, 2)
        combined = cv2.bitwise_or(thresh1, thresh2)
        
        # Morphological closing to connect broken lines
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
        processed = cv2.morphologyEx(combined, cv2.MORPH_CLOSE, kernel)
        
        return processed

    def detect_lines_optimized(preprocessed):
        """Three-pass line detection optimized for engineering drawings"""
        all_lines = []
        
        # Pass 1: Main structural lines (strict parameters)
        lines1 = cv2.HoughLinesP(
            preprocessed, rho=1, theta=np.pi/180, threshold=50,
            minLineLength=30, maxLineGap=10
        )
        if lines1 is not None:
            all_lines.extend(lines1)
        
        # Pass 2: Medium lines (moderate parameters)
        lines2 = cv2.HoughLinesP(
            preprocessed, rho=1, theta=np.pi/180, threshold=30,
            minLineLength=20, maxLineGap=15
        )
        if lines2 is not None:
            all_lines.extend(lines2)
        
        # Pass 3: Catch missed thin/broken lines (sensitive parameters)
        lines3 = cv2.HoughLinesP(
            preprocessed, rho=1, theta=np.pi/180, threshold=20,
            minLineLength=15, maxLineGap=25
        )
        if lines3 is not None:
            all_lines.extend(lines3)
        
        return np.array(all_lines) if all_lines else None

    def filter_duplicate_lines(lines, min_length=10, distance_threshold=12):
        """Fast duplicate removal optimized for performance"""
        if lines is None or len(lines) == 0:
            return None
        
        filtered = []
        
        for line in lines:
            x1, y1, x2, y2 = line[0] if len(line.shape) > 1 else line
            
            # Skip very short lines (likely noise)
            length = np.sqrt((x2-x1)**2 + (y2-y1)**2)
            if length < min_length:
                continue
            
            # Fast duplicate check - only compare with recent lines for speed
            is_duplicate = False
            check_count = min(len(filtered), 15)  # Check last 15 lines only
            
            for i in range(max(0, len(filtered) - check_count), len(filtered)):
                ex1, ey1, ex2, ey2 = filtered[i]
                
                # Fast Manhattan distance check
                dist = abs(x1-ex1) + abs(y1-ey1) + abs(x2-ex2) + abs(y2-ey2)
                if dist < distance_threshold:
                    is_duplicate = True
                    break
            
            if not is_duplicate:
                filtered.append([x1, y1, x2, y2])
        
        return np.array(filtered).reshape(-1, 1, 4) if filtered else None

    def detect_lines_in_tile(preprocessed, x_off, y_off, global_img, tile_idx):
        """Optimized line detection in single tile"""
        
        # Detect lines with optimized multi-pass approach
        lines = detect_lines_optimized(preprocessed)
        
        # Filter duplicates and noise
        if lines is not None:
            lines = filter_duplicate_lines(lines)
        
        structured_lines = []
        if lines is not None:
            for l_idx, line in enumerate(lines):
                x1, y1, x2, y2 = line[0] if len(line.shape) > 1 else line
                # Convert tile coordinates to global image coordinates
                X1, Y1, X2, Y2 = x1 + x_off, y1 + y_off, x2 + x_off, y2 + y_off
                # Draw line on global image (red color)
                cv2.line(global_img, (X1, Y1), (X2, Y2), (0, 0, 255), 1)
                # Store normalized coordinates
                structured_lines.append({
                    "id": f"T{tile_idx}-L{l_idx}",
                    "startX": X1 / global_img.shape[1],
                    "startY": Y1 / global_img.shape[0],
                    "endX": X2 / global_img.shape[1],
                    "endY": Y2 / global_img.shape[0]
                })
        return structured_lines

    def split_into_tiles(image, tile_size=768, overlap=40):
        """Optimized tile splitting - balanced size for performance"""
        H, W = image.shape[:2]
        step = tile_size - overlap
        tiles = []
        for y in range(0, H, step):
            for x in range(0, W, step):
                x_end, y_end = min(x + tile_size, W), min(y + tile_size, H)
                tiles.append(((x, y), image[y:y_end, x:x_end]))
        return tiles

    # === YOLO Post-processing Functions ===
    def non_max_suppression(detections, iou_threshold=0.35):
        """Removes overlapping detections"""
        if len(detections) == 0:
            return []
        boxes = np.array(detections)
        keep_detections = []
        for cls in np.unique(boxes[:, 5]):
            cls_boxes = boxes[boxes[:, 5] == cls]
            order = np.argsort(-cls_boxes[:, 4])
            cls_boxes = cls_boxes[order]
            while len(cls_boxes) > 0:
                best_box = cls_boxes[0]
                keep_detections.append(best_box)
                if len(cls_boxes) == 1:
                    break
                xx1 = np.maximum(best_box[0], cls_boxes[1:, 0])
                yy1 = np.maximum(best_box[1], cls_boxes[1:, 1])
                xx2 = np.minimum(best_box[2], cls_boxes[1:, 2])
                yy2 = np.minimum(best_box[3], cls_boxes[1:, 3])
                w = np.maximum(0, xx2 - xx1)
                h = np.maximum(0, yy2 - yy1)
                intersection = w * h
                union = (
                    (best_box[2] - best_box[0]) * (best_box[3] - best_box[1]) +
                    (cls_boxes[1:, 2] - cls_boxes[1:, 0]) * (cls_boxes[1:, 3] - cls_boxes[1:, 1]) -
                    intersection
                )
                iou = intersection / (union + 1e-6)
                cls_boxes = cls_boxes[1:][iou < iou_threshold]
        return np.array(keep_detections).tolist()

    def deduplicate_boxes(detections, dist_thresh=30):
        """Removes duplicate detections based on center distance"""
        unique = []
        for box in detections:
            x1, y1, x2, y2, conf, cls = box
            cx = (x1 + x2) / 2
            cy = (y1 + y2) / 2
            is_duplicate = False
            for u in unique:
                ux1, uy1, ux2, uy2, uconf, ucls = u
                ucx = (ux1 + ux2) / 2
                ucy = (uy1 + uy2) / 2
                if abs(cx - ucx) < dist_thresh and abs(cy - ucy) < dist_thresh and cls == ucls:
                    is_duplicate = True
                    if conf > uconf:
                        unique.remove(u)
                        unique.append(box)
                    break
            if not is_duplicate:
                unique.append(box)
        return unique

    def sliding_window_inference(image, model, tile_size=640, overlap=100, conf_thres=0.35):
        """Performs YOLO detection using sliding window approach"""
        H, W = image.shape[:2]
        step = tile_size - overlap
        all_detections = []
        
        for y in range(0, H, step):
            for x in range(0, W, step):
                y_end, x_end = min(y + tile_size, H), min(x + tile_size, W)
                tile = image[y:y_end, x:x_end]
                
                # Pad small tiles
                if tile.shape[0] < tile_size//2 or tile.shape[1] < tile_size//2:
                    padded = np.zeros((tile_size, tile_size, 3), dtype=np.uint8)
                    padded[:tile.shape[0], :tile.shape[1]] = tile
                    tile = padded
                
                # Run YOLO on tile
                results = model.predict(source=tile, conf=conf_thres, imgsz=tile_size, verbose=False)
                for r in results:
                    if r.boxes is not None:
                        for box in r.boxes:
                            cls = int(box.cls)
                            conf = float(box.conf)
                            x1, y1, x2, y2 = box.xyxy[0].tolist()
                            # Convert tile coordinates to global coordinates
                            x1 += x; x2 += x; y1 += y; y2 += y
                            all_detections.append([x1, y1, x2, y2, conf, cls])
                            
        return all_detections

    # === Main Processing Pipeline ===
    
    # Create output directory
    import os
    os.makedirs(output_dir, exist_ok=True)
    filename = Path(input_image_path).stem
    
    # Load YOLO model
    print("Loading YOLO model...")
    model = YOLO(model_path)
    
    # Load input image
    print(f"Loading image: {input_image_path}")
    full_img = cv2.imread(input_image_path)
    if full_img is None:
        raise ValueError(f"Could not load image: {input_image_path}")
    
    H, W = full_img.shape[:2]
    print(f"Image dimensions: {W} x {H}")

    # Class names for detected objects
    class_names = [
        "Air Intake","Angle Valve","Auto Reset","Auto Vent","Ball Valve","Bellow","Birdscreen",
        "Blind Flange","Blower","Break Line","Butterfly Valve","Cabinet or Chamber","Cap",
        "Catalyst","Check Valve","Choke Valve","Compressor or Vessel","Concentrate Drum",
        "Concentric Reducer","Connection Point","Continued on DWG","Cooler or Heater",
        "Coriolis Meter","DCS","Diaphragm Control Valve","Drain","Eccentric Reducer",
        "Electrical Heat Tracing and Insulation","Expension Beldow","Filling Hose","Filter",
        "Flame Arrestor","Flange Connection","Flexible Connection","Flow Arrow","Gate Valve",
        "Globe Valve","Installed in Control Panel","Instrument in Back of Panel",
        "Instrument on Local Panel","Insulation","Isolation Spacer","Line Number",
        "Local Instrument","Local Reset","Lubricator","Measuring Point","Mechanical Interlock",
        "Mono Valve","Needle Valve","Not Continued on DWG","Open Drain System","Open Vent",
        "Personal Protection","Positive Displacement Meter","Pressure Regulator",
        "Pulsation Dampner","Pump","Radar Type Level Meter","Relief or Safety Valve",
        "Remote Seal","Safeguarding System","Sample Connection","Sampling Probe",
        "Shut Down Ball Valve","Silencer","Slope Direction","Solenoid Valve","Spec Break",
        "Special Piping Item","Spectacle Blind","Steam Genrator","Stream Number","Tag","Tank",
        "Temporary Strainer","Thermostatic Valve","Three Way Valve","To Closed Drain System",
        "Ultrasonic Meter","Ultrasonic Sensor","Variable Speed drive","Y Type Strainer"
    ]

    # === Step 1: Run YOLO Detection ===
    print("Running YOLO detection...")
    detections = sliding_window_inference(full_img, model, tile_size=1824, overlap=0, conf_thres=0.25)
    print(f"Initial detections: {len(detections)}")
    
    # === Step 2: Post-process YOLO Results ===
    print("Post-processing YOLO detections...")
    detections = non_max_suppression(detections, iou_threshold=0.5)
    detections = deduplicate_boxes(detections, dist_thresh=30)
    print(f"Final detections after post-processing: {len(detections)}")

    # === Step 3: Draw YOLO Detections ===
    print("Drawing YOLO bounding boxes...")
    for (x1, y1, x2, y2, conf, cls) in detections:
        cv2.rectangle(full_img, (int(x1), int(y1)), (int(x2), int(y2)), (0, 255, 0), 2)

    # Save YOLO detection results
    output_yolo_detection = os.path.join(output_dir, f"{filename}_yolo_only.png")
    cv2.imwrite(output_yolo_detection, full_img)
    print(f"Saved YOLO-only image: {output_yolo_detection}")

    # === Step 4: Create Cleaned Image for Line Detection ===
    print("Creating cleaned image for line detection...")
    cleaned_img_for_lines = full_img.copy()
    for (x1, y1, x2, y2, _, _) in detections:
        # Fill detected object areas with WHITE to mask them out
        cv2.rectangle(cleaned_img_for_lines, (int(x1), int(y1)), (int(x2), int(y2)), (255, 255, 255), -1)
        # Add white border around objects
        cv2.rectangle(cleaned_img_for_lines, (int(x1), int(y1)), (int(x2), int(y2)), (255, 255, 255), 2)

    # Save cleaned image for debugging
    cleaned_image_path = os.path.join(output_dir, f"{filename}_cleaned_for_lines.png")
    cv2.imwrite(cleaned_image_path, cleaned_img_for_lines)
    print(f"Saved cleaned image: {cleaned_image_path}")

    # === Step 5: OPTIMIZED LINE DETECTION PIPELINE ===
    print("Starting optimized line detection...")
    
    # Split cleaned image into optimized tiles
    tiles = split_into_tiles(cleaned_img_for_lines, tile_size=768, overlap=40)
    structured_lines = []
    
    # Process each tile for line detection
    total_tiles = len(tiles)
    for tile_idx, ((x_off, y_off), tile) in enumerate(tiles, start=1):
        if tile_idx % 10 == 0:
            print(f"Processing tile {tile_idx}/{total_tiles}")
        
        # Preprocess tile with optimized method
        preprocessed_tile = preprocess_image_optimized(tile)
        
        # Detect lines in this tile
        tile_lines = detect_lines_in_tile(preprocessed_tile, x_off, y_off, full_img, tile_idx)
        structured_lines.extend(tile_lines)

    print(f"Detected {len(structured_lines)} line segments")

    # === Step 6: Save Results ===
    


   
    ############################################
    # Save YOLO detection JSON - FIXED TO MATCH MAIN CODE FORMAT
    detections_json = {
        "processed_image": input_image_path,
        "text_data": []
    }
    
    for idx, (x1, y1, x2, y2, conf, cls) in enumerate(detections):
        detections_json["text_data"].append({
            "topX": float(x1 / W),
            "topY": float(y1 / H),
            "bottomX": float(x2 / W),
            "bottomY": float(y2 / H),
            "id": f"{filename}-{idx}",
            "label": class_names[int(cls)],
            "score": float(conf),
            "Document_ID": os.path.basename(input_image_path),
            "text_associated": class_names[int(cls)]
        })
    
    detections_json_path = os.path.join(output_dir, f"{filename}_Detections.json")
    with open(detections_json_path, "w") as f:
        json.dump(detections_json, f, indent=4)
    print(f"Saved detections JSON: {detections_json_path}")

    # Save final processed image (with both YOLO boxes and detected lines)
    output_image_path = os.path.join(output_dir, f"{filename}_final_processed.png")
    cv2.imwrite(output_image_path, full_img)
    lines_json_path = os.path.join(output_dir, f'{filename}_Lines_detect.json')
    
    print("Connecting overlapping line segments...")
    structured_lines = connect_overlapping_lines_fast(structured_lines, W, H)
    
    # ✅ NEW: Apply improved line detection and deduplication
    print("Applying improved line detection and deduplication...")
    boxes_for_connection = detections_json['text_data']  # Use detected boxes
    cleaned_lines, connections = improve_line_detection_pipeline(
        structured_lines, 
        boxes_for_connection, 
        W, 
        H
    )
    structured_lines = cleaned_lines  # Use cleaned lines for further processing
    
    print(f"Line optimization: {len(detections_json['text_data'])} segments → {len(cleaned_lines)} cleaned lines")
    
    with open(lines_json_path, 'w') as f:
        json.dump(structured_lines, f, indent=4)
    print(f"Saved final processed image: {output_image_path}")
    
    # Apply long line detection and connection
   # result_new = detect_and_connect_lines(output_image_path, f"{filename}_final_processed_2.png")
    
    return output_image_path

# ============================================================================
# REGEX PATTERNS FOR LABEL VALIDATION
# ============================================================================
LABEL_PATTERNS = {
    "Compressor or Vessel": (
     r'(?:[A-Z]\s+\d{3,4}\s+\d[A-Z]|[A-Z]\s+\d{3,4}\s+[A-Z]|[A-Z]\s+\d{2,4}\s+\d|[A-Z]\s+\d{2,4}[A-Z]?)'
    ),
    "Line Number": (
    r'\d+[A-Z]?\s+\d\s+\d+|[A-Z]{1,4}\s+\d{2,5}(?:\s+[A-Z0-9]|\s+\d{1,2})?|[A-Z]{2,3}\s+[A-Z]{1,2}\s+\d{1,3}|[A-Z]{2,4}\s+[A-Z]\s+\d\s+\d{2}|[A-Z]{2,3}\s+[A-Z]\s+\d{3}\s+\d|\d\s+\d{5}\s+\d\s+\d{3}$'                            
 
),
    "Local Instrument": (
        r'[A-Z]{2,4}(?:\s+\d{4}\s+\d|\s+\d{4}\s+[A-Z]|\s+\d{4,5}[A-Z])?|[A-Z]{3}\s+\d{4}[A-Z]\s+\d'
    ),
    "DCS": (
       r'(?:[A-Z]+\s+[A-Z]+\s+\d{4,5}[A-Z]?\s+\d{1,2}|[A-Z]{1,5}\s+\d{4,5}[A-Z]?\s+\d{1,2})'
    ),
    "Instrument on Local Panel": (
        r'[A-Z]{2,4}|\d\s+[A-Z]{3,4}|[A-Z]{2,3}\s+\d{4}'
    ),
    "Instrument in Back of Panel": (
        r'[A-Z]{2,4}|\d\s+[A-Z]{3,4}'
    ),
    "Mechanical Interlock": (
        r'MIL\s+\d{4,5}\s+\d|MIL\s+\d{4}[A-Z]\s+\d\s+[A-Z]|MIL\s+\d{2}\s+\d{2}[A-Z]\s+\d|'
        r'MIL\s+\d{4}[A-Z]|MIL\s+\d{3}[A-Z]{2}\s+\d|[A-Z]{2}\s+\d{2,3}|MIL\s+\d{4}\s+\d{3}'
    ),
    "Installed in Control Panel": (
        r'[A-Z]{2,4}(?:\s+\d{4}\s+\d|\s+\d{4}\s+[A-Z]|\s+\d{4,5}[A-Z]?)?|[A-Z]{3}\s+\d{4}[A-Z]\s+\d'
    ),
    # "Tag": (
    #     r'[A-Z0-9]{2,}(?:[-\s]+[A-Z0-9]+)*'
    # ),
    "Blower": (
        r'[A-Z]\s+\d{2,4}(?:\s*[A-Z])?|[A-Z]\s+[A-Z]\s+\d{2,4}|[A-Z]\s+\d{3,4}\s+\d[A-Z]'
    ),
    "Tank": (
        r'[A-Z]\s+\d{2,4}(?:\s*[A-Z])?|[A-Z]\s+[A-Z]\s+\d{2,4}|[A-Z]\s+\d{3,4}\s+\d[A-Z]'
    ),
    "Cabinet or Chamber": (
        r'[A-Z0-9]{1,4}(?:[-\s]+[A-Z0-9]+)*'
    )
}

def check_regex_match(text, label):
    """
    Search for a regex pattern match anywhere in the text and return the matched portion with spaces replaced by hyphens.
    
    Args:
        text: Extracted OCR text (cleaned and normalized)
        label: Label type
    
    Returns:
        tuple: (bool, str) - (True if match found, matched text with spaces replaced by hyphens or empty string if no match)
    """
    if not text or not label in LABEL_PATTERNS:
        return False, ""

    # Compile pattern and search
    pattern = re.compile(LABEL_PATTERNS[label],re.IGNORECASE)
    match = pattern.search(text)
    
    if match:
        matched_text = match.group(0).strip()
        # Replace spaces with hyphens in the matched text
        matched_text = re.sub(r'\s+', '-', matched_text)
        return True, matched_text
    return False, ""

# ============================================================================
# MAIN OCR FUNCTION
# ============================================================================
#-------------- Open AI Integration for P&ID Tag's Extraction-------------------------

client = AzureOpenAI(api_version=azure_openai_version,
                     api_key=azure_openai_key,
                     base_url=f"{api_base}/openai/deployments/{deployment_name}")

def get_image_base64_data_url(image_path: str) -> str:
    if not os.path.exists(image_path):
        raise FileNotFoundError(image_path)
    ext = os.path.splitext(image_path)[1].lower()
    mime = "image/png" if ext == ".png" else "image/jpeg"
    with open(image_path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("utf-8")
    return f"data:{mime};base64,{b64}"


def calculate_distance_tag(box1: Dict, box2: Dict) -> float:
    """Calculate Euclidean distance between centers of two bounding boxes"""
    center1_x = (box1['topX'] + box1['bottomX']) / 2
    center1_y = (box1['topY'] + box1['bottomY']) / 2
    center2_x = (box2['topX'] + box2['bottomX']) / 2
    center2_y = (box2['topY'] + box2['bottomY']) / 2
    
    return np.sqrt((center1_x - center2_x)**2 + (center1_y - center2_y)**2)

def find_closest_tag(target_box: Dict, all_boxes: List[Dict]) -> Dict:
    """Find the closest tag to the target bounding box"""
    tag_boxes = [box for box in all_boxes if box.get('text_associated', '').lower() == 'tag']
    
    if not tag_boxes:
        return None
    
    min_distance = float('inf')
    closest_tag = None
    
    for tag_box in tag_boxes:
        distance = calculate_distance_tag(target_box, tag_box)
        if distance < min_distance:
            min_distance = distance
            closest_tag = tag_box
    
    return closest_tag

def calculate_distance(x1: float, y1: float, x2: float, y2: float) -> float:
    """Calculate Euclidean distance between two points."""
    return ((x2 - x1) ** 2 + (y2 - y1) ** 2) ** 0.5

def point_in_box(x: float, y: float, box: Dict, tolerance: float = 0.01) -> bool:
    """Check if a point is inside or near a bounding box."""
    return (box['topX'] - tolerance <= x <= box['bottomX'] + tolerance and
            box['topY'] - tolerance <= y <= box['bottomY'] + tolerance)

def find_nearest_box(x: float, y: float, boxes: List[Dict], max_distance: float = 0.05) -> Dict:
    """Find the nearest box to a given point."""
    min_distance = float('inf')
    nearest_box = None
    
    for box in boxes:
        # Calculate distance to box center
        box_center_x = (box['topX'] + box['bottomX']) / 2
        box_center_y = (box['topY'] + box['bottomY']) / 2
        distance = calculate_distance(x, y, box_center_x, box_center_y)
        
        if distance < min_distance and distance <= max_distance:
            min_distance = distance
            nearest_box = box
    
    return nearest_box

def connect_lines_to_boxes(lines: List[Dict], boxes: List[Dict], tolerance: float = 0.02) -> List[Dict]:
    """
    Connect lines to boxes based on their coordinates.
    
    Args:
        lines: List of line objects with startX, startY, endX, endY
        boxes: List of box objects with topX, topY, bottomX, bottomY
        tolerance: Distance tolerance for matching points to boxes
    
    Returns:
        List of connection objects with source_component, target_component, and line_details
    """
    connections = []
    
    for line in lines:
        start_x = line['startX']
        start_y = line['startY']
        end_x = line['endX']
        end_y = line['endY']
        
        # Find boxes connected to start and end points
        source_box = None
        target_box = None
        
        # Check if start point is in/near any box
        for box in boxes:
            if point_in_box(start_x, start_y, box, tolerance):
                source_box = box
                break
        
        # If not in a box, find nearest box
        if source_box is None:
            source_box = find_nearest_box(start_x, start_y, boxes, max_distance=tolerance * 2)
        
        # Check if end point is in/near any box
        for box in boxes:
            if point_in_box(end_x, end_y, box, tolerance):
                target_box = box
                break
        
        # If not in a box, find nearest box
        if target_box is None:
            target_box = find_nearest_box(end_x, end_y, boxes, max_distance=tolerance * 2)
        
        # Create connection if both source and target found
        if source_box and target_box and source_box['id'] != target_box['id']:
            connection = {
                'source_component': {'id': source_box['id']},
                'target_component': {'id': target_box['id']},
                'line_details': {
                    'id': line['id'],
                    'startX': start_x,
                    'startY': start_y,
                    'endX': end_x,
                    'endY': end_y
                }
            }
            connections.append(connection)
    
    return connections

def create_objects_with_connections_format(components: List[Dict], connections: List[Dict]) -> Dict:
    """Create the objects with connections format, keeping only first line between components."""
    if not components:
        return {"objects": []}
    
    components_by_id = {comp['id']: comp for comp in components}

    # Group lines between two components, keeping only the first one
    grouped_lines = {}
    for connection in connections:
        source_id = connection['source_component']['id']
        target_id = connection['target_component']['id']
        line_details = connection['line_details']

        pair_key = (source_id, target_id)
        
        # Only store the first line between each pair of components
        if pair_key not in grouped_lines:
            grouped_lines[pair_key] = {
                'source': components_by_id.get(source_id),
                'target': components_by_id.get(target_id),
                'line': line_details  # Store single line only
            }

    # Build objects map - include ALL components
    objects_map = {comp['id']: {
        "topX": comp.get('topX', 0),
        "topY": comp.get('topY', 0),
        "bottomX": comp.get('bottomX', 0),
        "bottomY": comp.get('bottomY', 0),
        "id": comp['id'],
        "label": comp.get('label', ''),
        "score": comp.get('score', 0),
        "text_associated": comp.get('text_associated', ''),
        "document_id": comp.get("Document_ID", ''),
        "connections": [],
        "asset_id": comp.get('asset_id', ''),
        "pid_id": comp.get('pid_id', ''),
        "device_id": comp.get('device_id', ''),
        "timeseries_id": comp.get('timeseries_id', ''),
        "document_name": comp.get('document_name', '')
    } for comp in components}

    # Add connections to each component with via array containing single line
    for pair_key, group in grouped_lines.items():
        source_id, target_id = pair_key
        
        if source_id not in objects_map or target_id not in objects_map:
            continue
        
        if group['target'] is None:
            continue
            
        target_comp = group['target']
        line = group['line']  # Single line
        
        # Create via array with single line and its length
        start_x = line.get('startX')
        start_y = line.get('startY')
        end_x = line.get('endX')
        end_y = line.get('endY')
        
        # Calculate length
        length = calculate_distance(start_x, start_y, end_x, end_y)
        
        via_array = [{
            "startX": start_x,
            "startY": start_y,
            "endX": end_x,
            "endY": end_y,
            "length": length
        }]
        
        objects_map[source_id]['connections'].append({
            "id": target_id,
            "label": target_comp.get('label', ''),
            "bbox": {
                "topX": target_comp.get('topX', 0),
                "topY": target_comp.get('topY', 0),
                "bottomX": target_comp.get('bottomX', 0),
                "bottomY": target_comp.get('bottomY', 0)
            },
            "text_associated": target_comp.get('text_associated', ''),
            "via": via_array
        })

    final_objects = list(objects_map.values())
    return {"objects": final_objects}

def load_json_file(file_path: str) -> Union[Dict, List, None]:
    """Load JSON file and return the data."""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        print(f"✓ Loaded {file_path}")
        return data
    except FileNotFoundError:
        print(f"✗ Error: File not found - {file_path}")
        return None
    except json.JSONDecodeError as e:
        print(f"✗ Error: Invalid JSON in {file_path} - {e}")
        return None
    except Exception as e:
        print(f"✗ Error loading {file_path}: {e}")
        return None

def save_json_file(data: Dict, output_path: str,filename, indent: int = 2):
    """Save data to JSON file."""
    try:
        # with open(output_path, 'w', encoding='utf-8') as f:
        #     json.dump(data, f, indent=indent, ensure_ascii=False)
        # print(f"✓ Saved output to {output_path}")

        output_path = os.path.join(output_path, f'{filename}_neo4j_graph_data.json')
        with open(output_path, 'w', encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)

       # logger.info(f"Saved relationship data to {output_path}")
        
        return True
    except Exception as e:
        print(f"✗ Error saving {output_path}: {e}")
        return False

def extract_components_from_json(data: Union[Dict, List]) -> List[Dict]:
    """Extract components/boxes from JSON data."""
    if isinstance(data, list):
        return data
    
    if isinstance(data, dict):
        # Try common keys for components
        for key in ['text_data', 'components', 'objects', 'detections', 'boxes']:
            if key in data and isinstance(data[key], list):
                return data[key]
        
        # Check for any list value
        for value in data.values():
            if isinstance(value, list) and len(value) > 0:
                return value
    
    return []

def extract_lines_from_json(data: Union[Dict, List]) -> List[Dict]:
    """Extract lines from JSON data."""
    if isinstance(data, list):
        return data
    
    if isinstance(data, dict):
        # Try common keys for lines
        for key in ['lines', 'connections', 'edges', 'links']:
            if key in data and isinstance(data[key], list):
                return data[key]
        
        # Check for any list value
        for value in data.values():
            if isinstance(value, list) and len(value) > 0:
                return value
    
    return []

def finding_relationships(lines_file: str, boxes_file: str, output_file: str, filename,tolerance: float = 0.02,):
    """
    Main function to process PID data by connecting lines to boxes.
    """
    logger.info(f"Starting relationship processing for {filename} with tolerance={tolerance}")
    # Load lines data
    print("\n[1/5] Loading lines data...")
    lines_data = load_json_file(lines_file)
    if lines_data is None:
        return False
    
    # Load boxes data
    print("\n[2/5] Loading boxes data...")
    boxes_data = load_json_file(boxes_file)
    if boxes_data is None:
        return False
    
    # Extract lines and boxes
    lines = extract_lines_from_json(lines_data)
    boxes = extract_components_from_json(boxes_data)
    
    if not lines:
        print("✗ Error: No lines found in the input file")
        return False
    
    if not boxes:
        print("✗ Error: No boxes found in the input file")
        return False
    
    print(f"  - Found {len(lines)} lines")
    print(f"  - Found {len(boxes)} boxes")
    
    # Connect lines to boxes
    print(f"\n[3/5] Connecting lines to boxes (tolerance={tolerance})...")
    connections = connect_lines_to_boxes(lines, boxes, tolerance)
    print(f"  - Created {len(connections)} connections")

    
    # Count unique component pairs
    unique_pairs = set()
    for conn in connections:
        pair = (conn['source_component']['id'], conn['target_component']['id'])
        unique_pairs.add(pair)
    print(f"  - Unique component pairs: {len(unique_pairs)}")
    
    # Process and merge data
    print("\n[4/5] Processing and merging data (keeping one line per connection)...")
    try:
        merged_data = create_objects_with_connections_format(boxes, connections)
    except Exception as e:
        print(f"✗ Error during processing: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    # Statistics
    total_objects = len(merged_data['objects'])
    objects_with_connections = sum(1 for obj in merged_data['objects'] if obj['connections'])
    total_connections = sum(len(obj['connections']) for obj in merged_data['objects'])
    total_via_segments = sum(
        len(conn['via']) 
        for obj in merged_data['objects'] 
        for conn in obj['connections']
    )
    
    print(f"  - Total objects: {total_objects}")
    print(f"  - Objects with connections: {objects_with_connections}")
    print(f"  - Total connections: {total_connections}")
    print(f"  - Total via segments: {total_via_segments}")
    
    # Save the output
    print("\n[5/5] Saving merged data...")
    success = save_json_file(merged_data, output_file,filename)
    
    if success:
        print("\n" + "=" * 70)
        print("✓ Processing completed successfully!")
        print("=" * 70)
        return True
    else:
        return False

############################# saikumar merge code end##################################################
driver_1 = GraphDatabase.driver(NEO4J_URI,auth=(NEO4J_USER,NEO4J_PASSWORD))
 
# def clear_neo4j_database(driver):
#         with driver.session() as session:
#             session.run("MATCH (n) DETACH DELETE n")

def sending_json_to_neo4j(filename):
    logger.info(f"Started pushing data to Neo4j for {filename}")
    class Neo4jImporter:
        def __init__(self, uri, user, password):
            try:
                # Try neo4j+s:// first (CA-signed certificates, advised for Azure)
                uri_s = uri.replace("neo4j+ssc://", "neo4j+s://") if "neo4j+ssc://" in uri else f"neo4j+s://{uri.split('://')[-1]}"
                self.driver = GraphDatabase.driver(uri_s, auth=(user, password))
                self.driver.verify_connectivity()
                logger.info(f"Neo4j driver initialized with {uri_s}")
            except Exception as e:
                logger.warning(f"Failed to connect with {uri_s}: {e}. Falling back to neo4j+ssc://")
                # Fallback to neo4j+ssc:// (self-signed certificates, works locally)
                try:
                    uri_ssc = uri if "neo4j+ssc://" in uri else f"neo4j+ssc://{uri.split('://')[-1]}"
                    self.driver = GraphDatabase.driver(uri_ssc, auth=(user, password))
                    self.driver.verify_connectivity()
                    logger.info(f"Neo4j driver initialized with {uri_ssc}")
                except Exception as e:
                    logger.error(f"Failed to initialize Neo4j driver with {uri_ssc}: {e}", exc_info=True)
                    raise
       
        def close(self):
            self.driver.close()
 
        # APPROACH 1: Store via points as JSON string
        def create_objects_and_connections_json_string(self, data):
            with self.driver.session() as session:
                for obj in data["objects"]:
                    session.execute_write(self._create_object_node, obj)
               
                for obj in data["objects"]:
                    if "connections" in obj:
                        for connection in obj["connections"]:
                            session.execute_write(self._create_connection_with_json_string, obj, connection)
       
        @staticmethod
        def _create_connection_with_json_string(tx, source_obj, connection):
            target_query = """
            MERGE (target:Object {id: $target_id})
            SET target.label = $target_label,
                target.text_associated = $target_text,
                target.topX = $target_topX,
                target.topY = $target_topY,
                target.bottomX = $target_bottomX,
                target.bottomY = $target_bottomY,
                target.document_id = $document_id,
                target.score = coalesce(target.score, 0.0),
                target.asset_id = coalesce(target.asset_id, ""),
                target.pid_id = coalesce(target.pid_id, ""),
                target.device_id = coalesce(target.device_id, ""),
                target.timeseries_id = coalesce(target.timeseries_id, ""),
                target.document_name = coalesce(target.document_name, "")
            """
            tx.run(target_query,
                target_id=connection["id"],
                target_label=connection["label"],
                target_text=connection["text_associated"],
                target_topX=connection["bbox"]["topX"],
                target_topY=connection["bbox"]["topY"],
                target_bottomX=connection["bbox"]["bottomX"],
                target_bottomY=connection["bbox"]["bottomY"],
                document_id=source_obj["document_id"])
           
            via_json = json.dumps(connection.get("via", []))
           
            rel_query = """
            MATCH (source:Object {id: $source_id})
            MATCH (target:Object {id: $target_id})
            MERGE (source)-[r:CONNECTED_TO]-(target)
            SET r.via_json = $via_json
            """
           
            tx.run(rel_query,
                source_id=source_obj["id"],
                target_id=connection["id"],
                via_json=via_json)
 
        # APPROACH 2: Create separate VIA_POINT nodes
        def create_objects_connections_with_via_nodes(self, data):
            with self.driver.session() as session:
                for obj in data["objects"]:
                    session.execute_write(self._create_object_node, obj)
               
                for obj in data["objects"]:
                    if "connections" in obj:
                        for connection in obj["connections"]:
                            session.execute_write(self._create_connection_with_via_nodes, obj, connection)
       
        @staticmethod
        def _create_connection_with_via_nodes(tx, source_obj, connection):
            target_query = """
            MERGE (target:Object {id: $target_id})
            SET target.label = $target_label,
                target.text_associated = $target_text,
                target.topX = $target_topX,
                target.topY = $target_topY,
                target.bottomX = $target_bottomX,
                target.bottomY = $target_bottomY,
                target.document_id = $document_id,
                target.score = coalesce(target.score, 0.0),
                target.asset_id = coalesce(target.asset_id, ""),
                target.pid_id = coalesce(target.pid_id, ""),
                target.device_id = coalesce(target.device_id, ""),
                target.timeseries_id = coalesce(target.timeseries_id, ""),
                target.document_name = coalesce(target.document_name, "")
            """
            tx.run(target_query,
                target_id=connection["id"],
                target_label=connection["label"],
                target_text=connection["text_associated"],
                target_topX=connection["bbox"]["topX"],
                target_topY=connection["bbox"]["topY"],
                target_bottomX=connection["bbox"]["bottomX"],
                target_bottomY=connection["bbox"]["bottomY"],
                document_id=source_obj["document_id"])
           
            rel_query = """
            MATCH (source:Object {id: $source_id})
            MATCH (target:Object {id: $target_id})
            MERGE (source)-[r:CONNECTED_TO]-(target)
            """
           
            tx.run(rel_query,
                source_id=source_obj["id"],
                target_id=connection["id"])
           
            if "via" in connection:
                for i, via in enumerate(connection["via"]):
                    via_point_id = f"{source_obj['id']}_to_{connection['id']}_via_{i}"
                   
                    via_query = """
                    MERGE (via:ViaPoint {id: $via_id})
                    SET via.startX = $startX,
                        via.startY = $startY,
                        via.endX = $endX,
                        via.endY = $endY,
                        via.length = $length,
                        via.order = $order
                    """
                   
                    tx.run(via_query,
                        via_id=via_point_id,
                        startX=via["startX"],
                        startY=via["startY"],
                        endX=via["endX"],
                        endY=via["endY"],
                        length=via["length"],
                        order=i)
                   
                    via_rel_query = """
                    MATCH (source:Object {id: $source_id})
                    MATCH (target:Object {id: $target_id})
                    MATCH (via:ViaPoint {id: $via_id})
                    MATCH (source)-[conn:CONNECTED_TO]-(target)
                    MERGE (conn)-[:HAS_VIA_POINT]->(via)
                    """
                   
                    tx.run(via_rel_query,
                        source_id=source_obj["id"],
                        target_id=connection["id"],
                        via_id=via_point_id)
 
        # APPROACH 3: Flatten via points into separate properties
        def create_objects_and_connections_flattened(self, data):
            with self.driver.session() as session:
                for obj in data["objects"]:
                    session.execute_write(self._create_object_node, obj)
               
                for obj in data["objects"]:
                    if "connections" in obj:
                        for connection in obj["connections"]:
                            session.execute_write(self._create_connection_flattened, obj, connection)
       
        @staticmethod
        def _create_connection_flattened(tx, source_obj, connection):
            target_query = """
            MERGE (target:Object {id: $target_id})
            SET target.label = $target_label,
                target.text_associated = $target_text,
                target.topX = $target_topX,
                target.topY = $target_topY,
                target.bottomX = $target_bottomX,
                target.bottomY = $target_bottomY,
                target.document_id = $document_id,
                target.score = coalesce(target.score, 0.0),
                target.asset_id = coalesce(target.asset_id, ""),
                target.pid_id = coalesce(target.pid_id, ""),
                target.device_id = coalesce(target.device_id, ""),
                target.timeseries_id = coalesce(target.timeseries_id, ""),
                target.document_name = coalesce(target.document_name, "")
            """
            tx.run(target_query,
                target_id=connection["id"],
                target_label=connection["label"],
                target_text=connection["text_associated"],
                target_topX=connection["bbox"]["topX"],
                target_topY=connection["bbox"]["topY"],
                target_bottomX=connection["bbox"]["bottomX"],
                target_bottomY=connection["bbox"]["bottomY"],
                document_id=source_obj["document_id"])
           
            via_startX = []
            via_startY = []
            via_endX = []
            via_endY = []
            via_lengths = []
           
            if "via" in connection:
                for via in connection["via"]:
                    via_startX.append(via["startX"])
                    via_startY.append(via["startY"])
                    via_endX.append(via["endX"])
                    via_endY.append(via["endY"])
                    via_lengths.append(via["length"])
           
            rel_query = """
            MATCH (source:Object {id: $source_id})
            MATCH (target:Object {id: $target_id})
            MERGE (source)-[r:CONNECTED_TO]->(target)
            ON CREATE SET
                r.via_startX = $via_startX,
                r.via_startY = $via_startY,
                r.via_endX   = $via_endX,
                r.via_endY   = $via_endY,
                r.via_lengths = $via_lengths
            ON MATCH SET
                r.via_startX = coalesce(r.via_startX, []) + $via_startX,
                r.via_startY = coalesce(r.via_startY, []) + $via_startY,
                r.via_endX   = coalesce(r.via_endX, []) + $via_endX,
                r.via_endY   = coalesce(r.via_endY, []) + $via_endY,
                r.via_lengths = coalesce(r.via_lengths, []) + $via_lengths
            """
            tx.run(rel_query,
                source_id=source_obj["id"],
                target_id=connection["id"],
                via_startX=via_startX,
                via_startY=via_startY,
                via_endX=via_endX,
                via_endY=via_endY,
                via_lengths=via_lengths)
 
        @staticmethod
        def _create_object_node(tx, obj):
            query = """
            MERGE (o:Object {id: $id})
            SET o.label = $label,
                o.score = $score,
                o.text_associated = $text_associated,
                o.topX = $topX,
                o.topY = $topY,
                o.bottomX = $bottomX,
                o.bottomY = $bottomY,
                o.asset_id = $asset_id,
                o.pid_id = $pid_id,
                o.device_id = $device_id,
                o.timeseries_id = $timeseries_id,
                o.document_name = $document_name,
                o.document_id = $document_id
            """
            tx.run(query,
                id=obj["id"],
                label=obj["label"],
                score=obj["score"],
                text_associated=obj["text_associated"],
                document_id=obj["document_id"],
                topX=obj["topX"],
                topY=obj["topY"],
                bottomX=obj["bottomX"],
                bottomY=obj["bottomY"],
                asset_id=obj.get("asset_id", ""),
                pid_id=obj.get("pid_id", ""),
                device_id=obj.get("device_id", ""),
                timeseries_id=obj.get("timeseries_id", ""),
                document_name=obj.get("document_name", ""))
 
    def import_data_approach_3(json_data, NEO4J_URI, NEO4J_USER, NEO4J_PASSWORD):
        """Approach 3: Flatten via points into arrays"""
        importer = Neo4jImporter(NEO4J_URI, NEO4J_USER, NEO4J_PASSWORD)
        try:
            importer.create_objects_and_connections_flattened(json_data)
            logger.info(f"Data imported successfully to Neo4j for {filename}")
        except Exception as e:
            logger.error(f"Failed to import data to Neo4j for {filename}: {e}", exc_info=True)
            raise
        finally:
            importer.close()
 
    try:
        your_json_data = "images"
        n4j = os.path.join(your_json_data, f"{filename}_neo4j_graph_data.json")
        with open(n4j, "r", encoding="utf-8") as f:
            data = json.load(f)
       
        # Trim environment variables to avoid whitespace issues
        NEO4J_URI = os.getenv("NEO4J_URI", "").strip()
        NEO4J_USER = os.getenv("NEO4J_USER", "").strip()
        NEO4J_PASSWORD = os.getenv("NEO4J_PASSWORD", "").strip()
        import_data_approach_3(data, NEO4J_URI, NEO4J_USER, NEO4J_PASSWORD)
    except FileNotFoundError as e:
        logger.error(f"Neo4j graph data JSON not found: {n4j}", exc_info=True)
        raise
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in Neo4j graph data: {n4j}", exc_info=True)
        raise
    except Exception as e:
        logger.error(f"Failed to push data to Neo4j for {filename}: {e}", exc_info=True)
        raise

def P_and_ID_insert_extracted_data_database(filename):
    logger.info(f"Started inserting data to Snowflake for {filename}")
    input_database_json = "images"
    json_path = os.path.join(input_database_json, f"{filename}_Json_results_OCR.json")

    try:
        data = load_json(json_path)
    except FileNotFoundError as e:
        logger.error(f"JSON file not found: {json_path}", exc_info=True)
        raise
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in file: {json_path}", exc_info=True)
        raise

    labels_of_interest = {
        "Local Instrument",
        "DCS",
        "Line Number",
        "Mechanical Interlock",
        "Instrument on Local Panel",
        "Instrument in Back of Panel",
        "Installed in Control Panel",
        "Tag",
        "Blower",
        "Tank",
        "Compressor or Vessel",
        "Cabinet or Chamber",
        "Pump"
    }

    image = data["processed_image"]
    image_name = os.path.basename(image)
    parts = image_name.split('_')
    document_id = "_".join(parts[:2]) if len(parts) >= 2 else parts[0]
    prefix_name = parts[0]
    file_type = 'A4'
    now = get_est_now()
    time = now.strftime("%Y-%m-%d %H:%M:%S")

    seen_rows = set()
    insert_values = []

    for item in data.get("text_data", []):
        label = item.get("text_associated")
        if label not in labels_of_interest:
            continue

        suffix_name = item.get("label", "")
        # if label == "Line Number":
        #     suffix_name = re.sub(r'["()]', '', suffix_name)
        #     parts = [part.strip() for part in suffix_name.split(",")]
        #     suffix_name = "-".join(parts)
        # else:
        #     suffix_name = suffix_name.replace(",", "-").replace(" ", "")

        tag_variant = f'{prefix_name}_{suffix_name}'

        # ✅ Expand tags containing "/" into separate variants
        # expanded_tags = []
        # if "/" in modified_tag:
        #     # Example: E17-A_F-1158A/B
        #     base_part, suffix_part = modified_tag.split("/", 1)
        #     match = re.match(r"(.+?)([A-Z]+)$", base_part)
        #     if match:
        #         prefix, last_letters = match.groups()
        #         expanded_tags.append(f"{prefix}{last_letters}")  # E17-A_F-1158A
        #         expanded_tags.append(f"{prefix}{suffix_part}")   # E17-A_F-1158B
        #     else:
        #         # fallback if unexpected format
        #         expanded_tags.append(base_part)
        #         expanded_tags.append(suffix_part)
        # else:
        #     expanded_tags.append(modified_tag)

        # Classify and prepare rows for each expanded tag
        # for tag_variant in expanded_tags:
        # tag_variant = re.sub(r'(\d+)([A-Z])($|-)', r'\1-\2\3', modified_tag)
        tag_type = classify_tag_type(tag_variant, file_type)
        logger.info(f"P&ID Classify -> {tag_variant} -> {tag_type}")

        row = (
            tag_variant,
            image_name,
            f"https://eni-dev.pimshosting.com/dcs-documents-details?Domain=NLAsset&DocID={document_id}",
            file_type,
            time,
            "P&ID",
            tag_type
        )
        row_key = (tag_variant, image_name, document_id)
        if row_key not in seen_rows:
            seen_rows.add(row_key)
            insert_values.append(row)

    if not insert_values:
        logger.info(f"No values to insert for {filename}")
        return

    logger.info(f"Prepared {len(insert_values)} unique rows for insertion")

    try:
        conn = get_snowflake_connection()
        query = """
        INSERT INTO ENI_NL_DEV_TAGREGISTRY.DEV.TBL_TAG_INFO
        ("TAG_NO", "FILE_NAME", "URL", "FILE_TYPE", "CREATED_DATE", "FILE_DESCRIPTION", "TAGTYPE")
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        try:
            with conn.cursor() as cursor:
                cursor.executemany(query, insert_values)
            conn.commit()
            logger.info(f"✅ Inserted {len(insert_values)} rows into Snowflake for {filename}")
        finally:
            conn.close()
    except Exception as e:
        logger.error(f"Failed to insert data into Snowflake for {filename}: {e}", exc_info=True)
        raise

#------------------------------------------------------------------------(CKR CODE STARTS)----------------------------------------------------------------------------------
# Utility: small logger helper

# Utility: small logger helper
def log(msg):
    print(msg)

# -----------------------------
# === helper utils ============
# -----------------------------
def compute_tile_starts(length, tile, overlap):
    step = tile - overlap
    starts = list(range(0, max(1, length - tile + 1), step))
    if not starts or starts[-1] + tile < length:
        starts.append(max(0, length - tile))
    return starts

# -----------------------------
# === Cropping / Layout =======
# -----------------------------
def crop_image(image_path, layout_model_path, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    model = YOLO(layout_model_path)

    img = cv2.imread(image_path)
    if img is None:
        raise FileNotFoundError(image_path)

    results = model(img)[0]
    for b in results.boxes:
        x1, y1, x2, y2 = map(int, b.xyxy[0].tolist())
        cropped = img[y1:y2, x1:x2]

        # cv2.imwrite(os.path.join(output_dir, "layout_detected.png"), img)
        cv2.imwrite(os.path.join(output_dir, "cropped.png"), cropped)
        log(f"Saved cropped image at {os.path.join(output_dir,'cropped.png')}")
        return cropped
    # if no boxes found, return original
    cv2.imwrite(os.path.join(output_dir, "cropped.png"), img)
    return img


# merge detections
def intersects(a, b):
    return not (a[2] < b[0] or a[0] > b[2] or a[3] < b[1] or a[1] > b[3])

def merge_envelope(boxes):
    x1 = min(b[0] for b in boxes)
    y1 = min(b[1] for b in boxes)
    x2 = max(b[2] for b in boxes)
    y2 = max(b[3] for b in boxes)
    conf = max(b[4] for b in boxes)
    cls = boxes[0][5]
    return [x1, y1, x2, y2, conf, cls]

def merge_clusters(detections, target_class=None):
    # Filter by class if needed
    if target_class is not None:
        dets = [d for d in detections if d[5] == target_class]
    else:
        dets = detections[:]

    if not dets:
        return []

    # Build graph adjacency
    n = len(dets)
    adj = [[] for _ in range(n)]

    for i in range(n):
        for j in range(i+1, n):
            if intersects(dets[i], dets[j]):
                adj[i].append(j)
                adj[j].append(i)

    visited = [False]*n
    merged_boxes = []

    # DFS find clusters
    for i in range(n):
        if not visited[i]:
            stack = [i]
            cluster = []
            visited[i] = True
            while stack:
                u = stack.pop()
                cluster.append(dets[u])
                for v in adj[u]:
                    if not visited[v]:
                        visited[v] = True
                        stack.append(v)
            # Merge cluster
            merged_boxes.append(merge_envelope(cluster))

    # Return merged clusters + untouched non-target class detections
    if target_class is None:
        return merged_boxes
    else:
        other = [d for d in detections if d[5] != target_class]
        return merged_boxes + other


def remove_duplicate_bboxes(dets, dist_thresh=25):
    """
    dets = [x1, y1, x2, y2, conf, cls]
    """
    cleaned = []
    used = set()

    def centroid(box):
        x1, y1, x2, y2 = box[:4]
        return ((x1+x2)/2, (y1+y2)/2)

    for i in range(len(dets)):
        if i in used:
            continue

        A = dets[i]
        cA = centroid(A)
        group = [A]
        used.add(i)

        for j in range(i+1, len(dets)):
            if j in used:
                continue

            B = dets[j]
            if A[5] != B[5]:   # must be same class
                continue

            cB = centroid(B)
            dist = np.linalg.norm(np.array(cA) - np.array(cB))

            if dist < dist_thresh:
                group.append(B)
                used.add(j)

        # keep larger bbox
        best = max(group, key=lambda b: (b[2]-b[0]) * (b[3]-b[1]))
        cleaned.append(best)

    return cleaned

# -----------------------------
# === Component detection =====
# -----------------------------
def detect_components(image, model_path, imgsz=2528, conf_thres=0.25):
    model = YOLO(model_path)
    results = model.predict(
        source=image,
        conf=conf_thres,
        imgsz=imgsz,
        verbose=False
    )

    detections = []
    for r in results:
        for box in r.boxes:
            x1, y1, x2, y2 = box.xyxy[0].tolist()
            conf = float(box.conf)
            cls = int(box.cls)
            detections.append([x1, y1, x2, y2, conf, cls])

    return detections, image

# -----------------------------
# === Mask / line detection ==
# -----------------------------
def detect_line_mask(image, model_path, tile, overlap, conf,
                       area_thresh=80, perim_thresh=40, min_dim=20,
                       poly_thickness=2, do_skeletonize=True):
    """
    Build mask from model polygons, remove small specks while keeping thin long lines.
    """
    model = YOLO(model_path)
    H, W = image.shape[:2]
    raw_mask = np.zeros((H, W), dtype=np.uint8)

    xs = compute_tile_starts(W, tile, overlap)
    ys = compute_tile_starts(H, tile, overlap)

    # Fill polygons
    for y in ys:
        for x in xs:
            tile_img = image[y:y+tile, x:x+tile]
            results = model.predict(tile_img, conf=conf, verbose=False)
            for r in results:
                if r.masks is None:
                    continue
                for poly in r.masks.xy:
                    pts = np.array([[int(px+x), int(py+y)] for px, py in poly], dtype=np.int32)
                    if pts.shape[0] < 3:
                        continue
                    try:
                        cv2.fillPoly(raw_mask, [pts], 255)
                    except Exception:
                        cv2.fillPoly(raw_mask, [pts.reshape((-1,1,2))], 255)

    # Morphological opening to remove tiny specks
    kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (3,3))
    raw_mask = cv2.morphologyEx(raw_mask, cv2.MORPH_OPEN, kernel, iterations=1)

    # Contours and filtering
    contours, _ = cv2.findContours(raw_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
    kept_contours = []
    for cnt in contours:
        area = cv2.contourArea(cnt)
        perim = cv2.arcLength(cnt, closed=False)
        x,y,w,h = cv2.boundingRect(cnt)
        if (area >= area_thresh) or (perim >= perim_thresh) or (max(w,h) >= min_dim):
            kept_contours.append(cnt)

    final_mask = np.zeros_like(raw_mask)
    if kept_contours:
        cv2.drawContours(final_mask, kept_contours, -1, 255, thickness=poly_thickness, lineType=cv2.LINE_AA)

    # optional skeletonize (thinning)
    if do_skeletonize:
        try:
            final_mask = cv2.ximgproc.thinning(final_mask)
        except Exception:
            pass

    return final_mask

# -----------------------------
# === Segments extraction ====
# -----------------------------
def extract_segments(mask):
    H, W = mask.shape
    visited = set()
    points = np.column_stack(np.where(mask == 255))
    segments = []

    def bfs(px, py):
        q = deque([(px, py)])
        visited.add((px, py))
        seg = []
        while q:
            x, y = q.popleft()
            seg.append((x, y))
            for nx, ny in [(x+1,y),(x-1,y),(x,y+1),(x,y-1),(x+1,y+1),
                           (x-1,y-1),(x+1,y-1),(x-1,y+1)]:
                if 0 <= nx < H and 0 <= ny < W:
                    if mask[nx, ny] == 255 and (nx, ny) not in visited:
                        visited.add((nx, ny))
                        q.append((nx, ny))
        return seg

    for px, py in points:
        if (px, py) not in visited:
            segments.append(bfs(px, py))

    return segments

# -----------------------------
# === Build lines json ========
# -----------------------------
def build_lines_json(segments):
    lines = []
    for i, seg in enumerate(segments):
        if len(seg) < 2:
            continue
        coords = []
        coords.append([seg[0][1], seg[0][0]])
        for idx, (x, y) in enumerate(seg[1:-1]):
            if idx % 10 == 0:
                coords.append([y, x])
        coords.append([seg[-1][1], seg[-1][0]])
        lines.append({
            "line_id": f"L{i+1}",
            "line_coordinates": coords
        })
    return lines



def build_components_json(detections, class_names_map, filename):
    """
    detections: [x1,y1,x2,y2,conf,cls,src]
    class_names_map: {
        "model_1": names_1,
        "model_2": names_2
    }
    """
    comps = []

    for idx, (x1, y1, x2, y2, conf, cls, src) in enumerate(detections):
        class_name = class_names_map[src][int(cls)]

        comps.append({
            "id": idx,
            "name": class_name,
            "bbox": [int(x1), int(y1), int(x2), int(y2)],
            "center": [
                int((x1 + x2) / 2),
                int((y1 + y2) / 2)
            ],
            "text_associated": class_name,
            "document_id": filename,
            # "source_model": src,          # ✅ optional but useful
            "confidence": round(conf, 3)  # ✅ optional but recommended
        })

    return comps

def build_components_json_temp(detections, class_names_map, filename):
    """
    detections: [x1,y1,x2,y2,conf,cls,src]
    class_names_map: {
        "model_1": names_1,
        "model_2": names_2
    }
    """
    comps = []

    for idx, (x1, y1, x2, y2, conf, cls, src) in enumerate(detections):
        class_name = class_names_map[src][int(cls)]

        # --- bbox extraction ---
        topX = float(x1)
        topY = float(y1)
        bottomX = float(x2)
        bottomY = float(y2)

        comps.append({
            "id": idx,
            "name": class_name,

            # original bbox (kept)
            "bbox": [int(x1), int(y1), int(x2), int(y2)],

            # extracted coordinates (NEW)
            "topX": topX,
            "topY": topY,
            "bottomX": bottomX,
            "bottomY": bottomY,

            "center": [
                int((x1 + x2) / 2),
                int((y1 + y2) / 2)
            ],
            "text_associated": class_name,
            "document_id": filename,
            # "source_model": src,          # optional
            "confidence": round(conf, 3)
        })

    return comps


# -----------------------------
# === Simple touch connections
# -----------------------------
def component_touches_line(line_coords, component, margin=5):
    x1, y1, x2, y2 = component["bbox"]
    x1 -= margin; y1 -= margin; x2 += margin; y2 += margin
    for (x, y) in line_coords:
        if x1 <= x <= x2 and y1 <= y <= y2:
            return True
    return False




def build_simple_touch_connections_1(components, lines, margin=5):
    final_connections = []

    # Build lookup: id → full component object
    comp_map = {c["id"]: c for c in components}

    for line in lines:
        line_id = line["line_id"]
        coords = line["line_coordinates"]

        touched = []
        for comp in components:
            if component_touches_line(coords, comp, margin):
                touched.append(comp)

        if len(touched) < 2:
            continue

        # Create pairwise connections
        for i in range(len(touched)):
            for j in range(i+1, len(touched)):
                compA = touched[i]
                compB = touched[j]

                final_connections.append({
                    "from": compA,
                    "to": compB,
                    "line_id": line_id           # NEW FORMAT
                })

    return final_connections

# -----------------------------
# === LLM OCR helpers ==========
# -----------------------------
def get_image_base64_data_url(path):
    """
    Simple helper - returns a data URL (base64) for the image.
    Replace with your preferred upload method if needed.
    """
    import base64
    with open(path, "rb") as f:
        b = f.read()
    mime = "image/png"
    b64 = base64.b64encode(b).decode("ascii")
    return f"data:{mime};base64,{b64}"


def extract_label_with_llm(cropped_roi_path: str) -> str:
    """
    Sends a single cropped label image to LLM and returns normalised tag.
    Replace client and model name as per your LLM provider.
    """
    prompt = """
    You are an expert in P&ID and instrumentation tag extraction.

    Your tasks:

    1. Read ONLY the tag text that is explicitly and clearly visible in the image.
    - Do NOT infer, assume, autocomplete, or hallucinate any part of a tag.
    - If no complete and valid tag text is visible in the image, do NOT extract anything and return nothing.

    2. Normalize the extracted tag using strict rules:
    - Insert a hyphen between letters and numbers (E10 → E-10).
    - Insert a hyphen between number → letter boundaries (1450E → 1450-E).
    - Keep multi-letter codes intact (SSV, PIT, TIT, TE, PI, etc.).
    - For line numbers: remove the hyphen between the leading number group and the following letter code (e.g., 90-DSH → 90DSH).
    - Include the pipe size in inches ONLY if the pipe size and line number are visibly part of the same tag in the image, exactly as shown, with a space after the inches (8" 90DSH-1-020).
    - Do NOT add a pipe size, line number, or any other element that is not visibly present in the same tag text.
    - Do not confuse letters with numbers:
        - D is never 0.
        - S is never 5.
        - Multi-letter codes such as DS, SH, and DSH must always be preserved as letters.
    - If a tag contains a combined suffix in the form A/B at the end (e.g., 1158A/B), split it into two normalized tags, inserting a hyphen before each suffix (1158-A and 1158-B).
    - If a tag contains a combined suffix in the form C/D followed by additional sections (e.g., 1158C/D-1), split it into two normalized tags while preserving the trailing section (1158-C-1 and 1158-D-1).
    - Except for the explicit suffix-splitting rules above, do NOT create multiple tags.
    - Do NOT add or remove characters except as required by these normalization rules.

    3. Output rules:
    - Respond ONLY with the normalized tag (or two tags when suffix-splitting applies).
    - If no valid tag is visible in the image, output nothing.
    - No JSON.
    - No explanation.
    - No backticks.
    
        """

    if client is None:
        # No LLM client configured — fallback to empty
        log("LLM client is not configured. Skipping OCR for " + cropped_roi_path)
        return ""

    try:
        data_url = get_image_base64_data_url(cropped_roi_path)
    except Exception as e:
        log(f"Cannot encode ROI {cropped_roi_path}: {e}")
        return ""

    try:
        # Replace this call with your provider's chat completion call & adapt parsing
        resp = client.chat.completions.create(
            model="gpt-4o",  # update as needed
            messages=[{
                "role": "user",
                "content": [
                    {"type": "image_url", "image_url": {"url": data_url}},
                    {"type": "text", "text": prompt}
                ]
            }],
            temperature=0.0,
            max_tokens=50,
        )
        tag = resp.choices[0].message.content.strip()
        log(f"LLM → {tag} for {cropped_roi_path}")
        return tag
    except Exception as e:
        log(f"LLM failed for {cropped_roi_path}: {e}")
        return ""

def extract_component_labels_2(components, cropped_image_path, filename):

    def clean_label(text):
        if not text:
            return None
        text = text.strip()
        return text if text else None


    labels_of_interest = {
        "Local Instrument","DCS","Line Number","Mechanical Interlock",
        "Instrument on Local Panel","Instrument in Back of Panel",
        "Installed in Control Panel","Cabinet or Chamber"
    }

    mixed_names = ["Tank", "Compressor or Vessel"]
    outside_only_names = ["Blower", "Pump"]

    DIST_THRESHOLD = 600

    img = cv2.imread(cropped_image_path)
    if img is None:
        log("ERROR: Could not load image for OCR.")
        return components

    H, W = img.shape[:2]
    crop_dir = "outputs/component_label_crops"
    os.makedirs(crop_dir, exist_ok=True)

    # ----------------------------
    # Precompute centers + defaults
    # ----------------------------
    for comp in components:
        x1, y1, x2, y2 = comp["bbox"]
        comp["center"] = [(x1 + x2) // 2, (y1 + y2) // 2]

        comp["topX"] = x1 / W
        comp["topY"] = y1 / H
        comp["bottomX"] = x2 / W
        comp["bottomY"] = y2 / H

        comp["id"] = f"{filename}-{comp['id']}"
        comp["text_associated"] = comp["name"]
        comp["label"] = comp["name"]
        comp["label_finalized"] = False

    tag_items = [c for c in components if c["name"] == "Tag"]

    # =====================================================
    # PHASE 1 — labels_of_interest (INSIDE ONLY)
    # =====================================================
    for comp in components:
        if comp["label_finalized"]:
            continue

        if comp["name"] not in labels_of_interest:
            continue

        x1, y1, x2, y2 = comp["bbox"]
        roi = img[y1:y2, x1:x2]

        crop_path = os.path.join(crop_dir, f"inside_{comp['id']}.png")
        cv2.imwrite(crop_path, roi)

        label = clean_label(extract_label_with_llm(crop_path))
        if label:
            comp["label"] = label

        comp["label_finalized"] = True

    # =====================================================
    # PHASE 2 — mixed_names (INSIDE → TAG)
    # =====================================================
    for comp in components:
        if comp["label_finalized"]:
            continue

        if comp["name"] not in mixed_names:
            continue

        x1, y1, x2, y2 = comp["bbox"]
        cx, cy = comp["center"]

        # ---- inside first
        roi = img[y1:y2, x1:x2]
        crop_path = os.path.join(crop_dir, f"inside_{comp['id']}.png")
        cv2.imwrite(crop_path, roi)

        inside_label = clean_label(extract_label_with_llm(crop_path))

        # ---- fallback to nearest tag
        outside_label = None
        if not inside_label and tag_items:
            nearest, best_dist = None, 1e9
            for tag in tag_items:
                tx, ty = tag["center"]
                d = ((cx - tx)**2 + (cy - ty)**2) ** 0.5
                if d < best_dist:
                    best_dist = d
                    nearest = tag

            if nearest and best_dist < DIST_THRESHOLD:
                tx1, ty1, tx2, ty2 = nearest["bbox"]
                roi = img[ty1:ty2, tx1:tx2]
                crop_path = os.path.join(crop_dir, f"outside_{comp['id']}.png")
                cv2.imwrite(crop_path, roi)
                outside_label = clean_label(extract_label_with_llm(crop_path))

        comp["label"] = inside_label or outside_label or comp["label"]
        comp["label_finalized"] = True

    # =====================================================
    # PHASE 3 — outside_only_names (TAG ONLY)
    # =====================================================
    for comp in components:
        if comp["label_finalized"]:
            continue

        if comp["name"] not in outside_only_names:
            continue

        cx, cy = comp["center"]
        outside_label = None

        if tag_items:
            nearest, best_dist = None, 1e9
            for tag in tag_items:
                tx, ty = tag["center"]
                d = ((cx - tx)**2 + (cy - ty)**2) ** 0.5
                if d < best_dist:
                    best_dist = d
                    nearest = tag

            if nearest and best_dist < DIST_THRESHOLD:
                tx1, ty1, tx2, ty2 = nearest["bbox"]
                roi = img[ty1:ty2, tx1:tx2]
                crop_path = os.path.join(crop_dir, f"outside_{comp['id']}.png")
                cv2.imwrite(crop_path, roi)
                outside_label = clean_label(extract_label_with_llm(crop_path))

        comp["label"] = outside_label or comp["label"]
        comp["label_finalized"] = True

    # ----------------------------
    # Cleanup
    # ----------------------------
    for comp in components:
        comp.pop("label_finalized", None)

    
    return components

# OCR Multi threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import hashlib

LLM_CACHE = {}

def image_hash(path):
    with open(path, "rb") as f:
        return hashlib.md5(f.read()).hexdigest()

def llm_cached(path):
    h = image_hash(path)
    if h in LLM_CACHE:
        return path, LLM_CACHE[h]

    label = extract_label_with_llm(path)
    LLM_CACHE[h] = label
    return path, label


def extract_component_labels_fast(components, cropped_image_path, filename):

    labels_of_interest = {
        "Local Instrument","DCS","Line Number","Mechanical Interlock",
        "Instrument on Local Panel","Instrument in Back of Panel",
        "Installed in Control Panel","Cabinet or Chamber"
    }
    mixed_names = ["Tank", "Compressor or Vessel","Tank 2"]
    outside_only_names = ["Blower", "Pump","Cooler 2","Filter","cooler 3","Positive Displacement Meter"]

    DIST_THRESHOLD = 600

    img = cv2.imread(cropped_image_path)
    if img is None:
        log("ERROR: Could not load image for OCR.")
        return components

    H, W = img.shape[:2]
    crop_dir = "images/component_label_crops"
    os.makedirs(crop_dir, exist_ok=True)

    crop_requests = []   # (comp, crop_path)
    tag_items = []

    # ----------------------------
    # Precompute geometry
    # ----------------------------
    for comp in components:
        x1, y1, x2, y2 = comp["bbox"]
        comp["center"] = [(x1 + x2)//2, (y1 + y2)//2]
        comp["id"] = f"{filename}-{comp['id']}"
        comp["label"] = comp["name"]
        comp["_final"] = False

        if comp["name"] == "Tag":
            tag_items.append(comp)

    # =====================================================
    # PHASE 1 — labels_of_interest (INSIDE)
    # =====================================================
    for comp in components:
        if comp["_final"]:
            continue
        if comp["name"] not in labels_of_interest:
            continue

        x1,y1,x2,y2 = comp["bbox"]
        roi = img[y1:y2, x1:x2]
        path = os.path.join(crop_dir, f"inside_{comp['id']}.png")
        cv2.imwrite(path, roi)
        crop_requests.append((comp, "inside", path))
        comp["_final"] = True

    # =====================================================
    # PHASE 2 — mixed_names (INSIDE → TAG)
    # =====================================================
    for comp in components:
        if comp["_final"]:
            continue
        if comp["name"] not in mixed_names:
            continue

        # inside
        x1,y1,x2,y2 = comp["bbox"]
        roi = img[y1:y2, x1:x2]
        path = os.path.join(crop_dir, f"inside_{comp['id']}.png")
        cv2.imwrite(path, roi)
        crop_requests.append((comp, "inside", path))

        # outside tag fallback
        cx,cy = comp["center"]
        nearest, best = None, 1e9
        for tag in tag_items:
            tx,ty = tag["center"]
            d = ((cx-tx)**2 + (cy-ty)**2)**0.5
            if d < best:
                best, nearest = d, tag

        if nearest and best < DIST_THRESHOLD:
            tx1,ty1,tx2,ty2 = nearest["bbox"]
            roi = img[ty1:ty2, tx1:tx2]
            path = os.path.join(crop_dir, f"outside_{comp['id']}.png")
            cv2.imwrite(path, roi)
            crop_requests.append((comp, "outside", path))

        comp["_final"] = True

    # =====================================================
    # PHASE 3 — outside_only_names (TAG ONLY)
    # =====================================================
    for comp in components:
        if comp["_final"]:
            continue
        if comp["name"] not in outside_only_names:
            continue

        cx,cy = comp["center"]
        nearest, best = None, 1e9
        for tag in tag_items:
            tx,ty = tag["center"]
            d = ((cx-tx)**2 + (cy-ty)**2)**0.5
            if d < best:
                best, nearest = d, tag

        if nearest and best < DIST_THRESHOLD:
            tx1,ty1,tx2,ty2 = nearest["bbox"]
            roi = img[ty1:ty2, tx1:tx2]
            path = os.path.join(crop_dir, f"outside_{comp['id']}.png")
            cv2.imwrite(path, roi)
            crop_requests.append((comp, "outside", path))

        comp["_final"] = True

    # =====================================================
    # THREADED LLM EXECUTION (🔥 only change)
    # =====================================================
    llm_results = {}

    with ThreadPoolExecutor(max_workers=6) as executor:
        futures = [executor.submit(llm_cached, p) for _,_,p in crop_requests]

        for f in as_completed(futures):
            path, label = f.result()
            llm_results[path] = label

    # =====================================================
    # ASSIGN RESULTS (rule-preserving)
    # =====================================================
    for comp, kind, path in crop_requests:
        label = llm_results.get(path)
        if not label:
            continue

        if kind == "inside":
            comp["label"] = label
        elif kind == "outside" and comp["label"] == comp["name"]:
            comp["label"] = label

    # Cleanup
    for comp in components:
        comp.pop("_final", None)

    return components

# -----------------------------
# === Neo4j push (batch) ======
# -----------------------------


def push_to_neo4j(components, lines, connections, uri, user, pwd, filename):
    driver = GraphDatabase.driver(uri, auth=(user, pwd))

    with driver.session() as session:

        # -------------------------------------------------------
        # 1️⃣ PUSH COMPONENTS
        # -------------------------------------------------------
        q_comp = """
        UNWIND $rows AS row
        MERGE (o:Object {id: row.id})
        SET 
            o.label = row.label,              
            o.caption = row.label,
            o.text_associated = row.text_associated,
            
            o.topX = row.topX,
            o.topY = row.topY,
            o.bottomX = row.bottomX,
            o.bottomY = row.bottomY,
            o.document_id = row.document_id
        """

        comp_rows = []
        for comp in components:
            comp_rows.append({
                "id": comp["id"],
                "label": comp.get("label", ""),
                "text_associated": comp.get("text_associated", ""),
                # "name": comp.get("name", ""),
                "topX": comp.get("topX", 0.0),
                "topY": comp.get("topY", 0.0),
                "bottomX": comp.get("bottomX", 0.0),
                "bottomY": comp.get("bottomY", 0.0),
                "document_id": comp.get("document_id", f"{filename}")
            })

        session.run(q_comp, {"rows": comp_rows})

        # -------------------------------------------------------
        # 2️⃣ PUSH LINES (ONLY LINE_ID)
        # -------------------------------------------------------
        q_line = """
        UNWIND $rows AS row
        MERGE (l:Line {line_id: row.line_id})
        """

        line_rows = []
        for l in lines:
            line_rows.append({
                "line_id": l["line_id"]   # NO COORDINATES
            })

        session.run(q_line, {"rows": line_rows})

        # -------------------------------------------------------
        # 3️⃣ PUSH CONNECTIONS
        # -------------------------------------------------------
        q_conn = """
        UNWIND $rows AS row
        MATCH (c1:Object {id: row.from_id})
        MATCH (c2:Object {id: row.to_id})
        MERGE (c1)-[r:CONNECTED_TO {line_id: row.line_id}]->(c2)
        SET r.document_id = row.document_id
        """

        conn_rows = []
        for c in connections:
            conn_rows.append({
                "from_id": c["from"]["id"],
                "to_id": c["to"]["id"],
                "line_id": c["line_id"],
                "document_id": c["from"]["document_id"]
            })

        session.run(q_conn, {"rows": conn_rows})

    driver.close()
    log("✔ Neo4j Push Complete")



def to_py(obj):
    """
    Recursively converts numpy types (np.int64, np.float32, np.ndarray)
    into Python-native types for JSON serialization.
    """
    if isinstance(obj, dict):
        return {k: to_py(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [to_py(v) for v in obj]
    elif isinstance(obj, tuple):
        return tuple(to_py(v) for v in obj)
    elif hasattr(obj, "item"):
        # numpy scalar
        return obj.item()
    elif isinstance(obj, np.ndarray):
        # convert array to list
        return obj.tolist()
    else:
        return obj

def send_image_array(img_array):
    """Converts a NumPy image array to PNG bytes and returns a Flask response."""
    _, buffer = cv2.imencode(".png", img_array)
    io_buf = io.BytesIO(buffer.tobytes())
    io_buf.seek(0)
    return send_file(io_buf, mimetype="image/png")

def load_json(filepath):
    with open(filepath, 'r') as file:
        return json.load(file)
    
def P_and_ID_insert_extracted_data_database_CKR(filename):
    logger.info(f"Started inserting data to Snowflake for {filename}")

    # JSON path
    json_path = os.path.join("images", f"{filename}_components_with_text.json")

    # Load JSON
    try:
        data = load_json(json_path)   # EXPECTS A LIST OF ITEMS
    except Exception as e:
        logger.error(f"Failed loading JSON: {json_path}", exc_info=True)
        raise

    # Labels we want to insert
    labels_of_interest = {
        "Local Instrument",
        "DCS",
        "Line Number",
        "Mechanical Interlock",
        "Instrument on Local Panel",
        "Instrument in Back of Panel",
        "Installed in Control Panel",
        "Tag",
        "Blower",
        "Tank",
        "Compressor or Vessel",
        "Cabinet or Chamber",
        "Pump",
        "Filter",
        "Cooler 2",
        "Tank 2",
        "cooler 3",
        "Positive Displacement Meter"
    }

    insert_values = []
    seen_rows = set()

    # Get image name from JSON id
    first_item = data[0]
    image_name = first_item["document_id"] + ".jpg"
    parts = first_item["document_id"].split("_")
    document_id = "_".join(parts[:2])
    prefix_name = parts[0]

    now = get_est_now()
    time = now.strftime("%Y-%m-%d %H:%M:%S")
    file_type = "A4"

    # LOOP OVER JSON ITEMS
    for item in data:
        label_type = item.get("text_associated")

        if label_type not in labels_of_interest:
            continue

        # 🔧 CHANGE START — split multi-tag labels into separate rows
        raw_label = item.get("label", "")

        # Handle cases like "P-1293-A \nP-1293-B"
        suffix_tags = [
            t.strip()
            for t in raw_label.splitlines()
            if t.strip()
        ]

        # If label is empty or invalid, skip
        if not suffix_tags:
            continue

        for suffix_name in suffix_tags:
            tag_variant = f"{prefix_name}_{suffix_name}"

            tag_type = classify_tag_type(tag_variant, file_type)

            row = (
                tag_variant,                                                   # TAG_NO
                image_name,                                                    # FILE_NAME
                f"https://pims.infra01.net/dcs-documents-details-custom?Domain=NLAsset&DocID={document_id}",
                file_type,                                                     # FILE_TYPE
                time,                                                          # CREATED_DATE
                "P&ID",                                                        # FILE_DESCRIPTION
                tag_type                                                       # TAGTYPE
            )

            row_key = (tag_variant, image_name, document_id)
            if row_key not in seen_rows:
                seen_rows.add(row_key)
                insert_values.append(row)
        # 🔧 CHANGE END

    if not insert_values:
        logger.info(f"No insertable items found in {filename}")
        return

    logger.info(f"Prepared {len(insert_values)} unique rows for insertion")

    try:
        conn = get_snowflake_connection()
        query = """
        INSERT INTO ENI_NL_DEV_TAGREGISTRY.DEV.TBL_TAG_INFO
        ("TAG_NO", "FILE_NAME", "URL", "FILE_TYPE", "CREATED_DATE", "FILE_DESCRIPTION", "TAGTYPE")
        VALUES (%s, %s, %s, %s, %s, %s, %s)
        """

        try:
            with conn.cursor() as cursor:
                cursor.executemany(query, insert_values)
            conn.commit()
            logger.info(f"✅ Inserted {len(insert_values)} rows into Snowflake for {filename}")
        finally:
            conn.close()

    except Exception as e:
        logger.error(f"Snowflake Insert Failed for {filename}: {e}", exc_info=True)
        raise


# -----------------------------
# === Main pipeline ===========
# -----------------------------

import time
def run_pipeline(
    image_path,
    output_dir,
    layout_model,
    component_model_1,
    component_model_2,
    line_model,
    class_names_1,
    class_names_2,
    neo4j_uri,
    neo4j_user,
    neo4j_pwd,
    filename
):

    # -------------------------------------
    # STEP 1 — Crop layout region
    # -------------------------------------
    def start_time():
        return time.time()
    def total_time(start_time):
        elapsed = time.time() - start_time  # total seconds
        minutes = int(elapsed // 60)
        seconds = int(elapsed % 60)
        logger.info(f"Time Taken: {minutes} minute(s) {seconds} second(s)")
    full_time=start_time()
    start_time_1=start_time()
    print("Crop image")
    cropped = crop_image(image_path, layout_model, output_dir)
    cropped_path = os.path.join(output_dir, f"{filename}_cropped.png")
    cv2.imwrite(cropped_path, cropped)
    total_time(start_time_1)
    
    # -----------------------------
    # Model 1 Inference (Large Components)
    # -----------------------------
    start_time_2=start_time()
    print("Detect components")
    detections_1,full_img_1 = detect_components(cropped, component_model_1)
    detections_1 = remove_duplicate_bboxes(detections_1)
    detections_1 = merge_clusters(detections_1, target_class=1)


    # tag model source
    detections_1 = [d + ["model_1"] for d in detections_1]

    EXCLUDED_CLASS_NAME = "Compressor or Vessel"
    EXCLUDED_CLASS_ID = class_names_2.index(EXCLUDED_CLASS_NAME)

    # -----------------------------
    # Model 2 Inference (Lines / Symbols)
    # -----------------------------
    # Detect using model_2
    detections_2,full_img = detect_components(cropped, component_model_2)

    # ❌ Remove "Compressor or Vessel" from model_2 detections
    detections_2 = [
        d for d in detections_2
        if d[5] != EXCLUDED_CLASS_ID
    ]

    # Remove duplicates AFTER filtering
    detections_2 = remove_duplicate_bboxes(detections_2)

    # Tag source
    detections_2 = [d + ["model_2"] for d in detections_2]


    # -----------------------------
    # Merge All Detections
    # -----------------------------
    all_detections = detections_1 + detections_2

    # Debug component boxes
    debug_img = full_img.copy()
    for (x1, y1, x2, y2, conf, cls,src) in all_detections:
        cv2.rectangle(debug_img, (int(x1), int(y1)), (int(x2), int(y2)), (0,255,0), 2)
    cv2.imwrite(os.path.join(output_dir, f"{filename}_YOLO_Component_detection.png"), debug_img)

    total_time(start_time_2)
    # -------------------------------------
    # STEP 3 — Build component objects (YOLO only)
    # -------------------------------------
    start_time_3=start_time()
    print("Build Component Objects")
    class_names_map = {
    "model_1": class_names_1,
    "model_2": class_names_2
    }

    components_raw = build_components_json_temp(
        all_detections,        # merged detections
        class_names_map,
        filename
    )

    total_time(start_time_3)
    # components = build_components_json(
    #     detections,
    #     class_names,
    #     filename
    # )

    # -------------------------------------
    # STEP 4 — Line detection mask
    # -------------------------------------
    print("Line Detection")
    start_time_4=start_time()
    tile = 640
    overlap = 100
    conf = 0.05


    mask = detect_line_mask(cropped, line_model, tile, overlap, conf)

    # Create overlay
    line_layer = debug_img.copy()
    line_layer[mask == 255] = (0,255,255)

    blended = cv2.addWeighted(debug_img, 0.7, line_layer, 0.3, 0)
    cv2.imwrite(os.path.join(output_dir, f"{filename}_line_overlay_image.png"), blended)

    total_time(start_time_4)
    # -------------------------------------
    # STEP 5 — Extract line segments + JSON
    # -------------------------------------
    segs = extract_segments(mask)
    lines = build_lines_json(segs)

    with open(os.path.join(output_dir, f"{filename}_lines.json"), "w", encoding="utf-8") as f:
        json.dump(to_py(lines), f, indent=2, ensure_ascii=False)

    # -------------------------------------
    # STEP 6 — Build touch-based connections
    # -------------------------------------

    print("Building connections")
    start_time_4=start_time()
    try:
        # print(components_raw)
        # print(lines)
        connections = build_simple_touch_connections_1(components_raw, lines, margin=5)
        if not connections:
            print("connections is empty")
        else:
            print("connections has items")
    except Exception as e:
        print("-------------",e)
    total_time(start_time_4)
    # -------------------------------------
    # STEP 7 — OCR extraction (updates components)
    # -------------------------------------
    print("OCR....")
    start_time_5=start_time()
    # components = extract_component_labels_2(components_raw, cropped_path, filename)
    components = extract_component_labels_fast(components_raw, cropped_path, filename)

    total_time(start_time_5)
    # 
    
    with open(os.path.join(output_dir, f"{filename}_Neo4j_connections.json"), "w", encoding="utf-8") as f:
        json.dump(to_py(connections), f, indent=2)

    
    # Save OCR-updated component list
    with open(os.path.join(output_dir, f"{filename}_components_with_text.json"), "w", encoding="utf-8") as f:
        json.dump(components, f, indent=2, ensure_ascii=False)

    with open(os.path.join(output_dir, f"{filename}_components_Raw.json"), "w", encoding="utf-8") as f:
        json.dump(components_raw, f, indent=2, ensure_ascii=False)
    print("Database insertion")
    start_time_6=start_time()
    P_and_ID_insert_extracted_data_database_CKR(filename)
    total_time(start_time_6)
    # -------------------------------------
    # STEP 8 — Neo4j push
    # -------------------------------------
    print("Neo4j pushing")
    start_time_7=start_time()
    push_to_neo4j(
        components,
        lines,
        connections,
        neo4j_uri,
        neo4j_user,
        neo4j_pwd,
        filename
    )
    total_time(start_time_7)
    logger.info("Total Time.....")
    total_time(full_time)

    # -------------------------------------
    # DONE
    # -------------------------------------
    return components, lines, connections, blended


class_names_list_1 = [
        'Blower', 'Compressor or Vessel', 'Cooler 2', 'Equipment',
        'Filter', 'Pump', 'Tank', 'Tank 2', 'cooler 3'
    ]

class_names_list_2=[
  "Air Intake","Angle Valve","Auto Reset","Auto Vent","Ball Valve","Bellow","Birdscreen","Blind Flange","Break Line",
  "Butterfly Valve","Cap","Catalyst","Check Valve","Choke Valve","Compressor or Vessel","Concentrate Drum","Concentric Reducer",
  "Connection Point","Continued on DWG","Coriolis Meter","DCS","Diaphragm Control Valve","Drain","Eccentric Reducer",
  "Electrical Heat Tracing and Insulation","Expension Beldow","Filling Hose","Filter","Flame Arrestor","Flange Connection",
  "Flexible Connection","Flow Arrow","Gate Valve","Globe Valve","Installed in Control Panel","Instrument in Back of Panel",
  "Instrument on Local Panel","Insulation","Isolation Spacer","Line Number","Local Instrument","Local Reset","Lubricator",
  "Measuring Point","Mechanical Interlock","Mono Valve","Needle Valve","Not Continued on DWG","Open Drain System","Open Vent",
  "Personal Protection","Positive Displacement Meter","Pressure Regulator","Pulsation Dampner","Pump","Radar Type Level Meter",
  "Relief or Safety Valve","Remote Seal","Safeguarding System","Sample Connection","Sampling Probe","Shut Down Ball Valve",
  "Silencer","Slope Direction","Solenoid Valve","Spec Break","Special Piping Item","Spectacle Blind","Steam Genrator",
  "Stream Number","Tag","Temporary Strainer","Thermostatic Valve","Three Way Valve","To Closed Drain System","Ultrasonic Meter",
  "Ultrasonic Sensor","Variable Speed drive","Y Type Strainer"
    ]

MODEL_PATH = "best.pt" 
OUTPUT_DIR = "images"
P_and_D_OUTPUT_DIR="images"

#------------------------------------------------------------------------(CKR CODE END)----------------------------------------------------------------------------------  

# app = func.FunctionApp(http_auth_level=func.AuthLevel.FUNCTION)
 



logger = logging.getLogger(__name__)

def _enhance_for_ocr(img: Image.Image) -> Image.Image:
    """
    Apply OCR-optimized enhancement exactly as in ocr_pdf_to_dark_images.
    """
    img = ImageEnhance.Sharpness(img).enhance(2.0)
    img = ImageEnhance.Contrast(img).enhance(2.5)
    img = ImageEnhance.Brightness(img).enhance(0.85)
    img = ImageOps.autocontrast(img, cutoff=2)
    return img

def pdf_to_images_pid(pdf_path: str,
                  output_dir: str = "images",
                  zoom_factor: float = 3.0,
                  contrast: float = 1.8,
                  brightness: float = 1.2) -> list[str]:
    """
    Convert PDF to images with **exact same visual output** as ocr_pdf_to_dark_images(dpi=300, darken=True).
    
    - Uses DPI = 300 (hardcoded for OCR quality)
    - Applies identical enhancement pipeline
    - Keeps original filename format: {name}_page1.png
    - Uses logger.info/error instead of print
    - Fully compatible with your Flask route
    """
    try:
        os.makedirs(output_dir, exist_ok=True)
        image_paths: list[str] = []

        doc = fitz.open(pdf_path)
        base_name = Path(pdf_path).stem

        # Fixed DPI=300 for OCR (equivalent to zoom ≈ 4.1667)
        dpi = 300
        zoom = dpi / 72
        mat = fitz.Matrix(zoom, zoom)

        logger.info(f"Converting PDF '{base_name}' ({doc.page_count} page(s)) at {dpi} dpi with OCR enhancement")

        for page_num in range(doc.page_count):
            page = doc[page_num]
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

            # Apply exact same enhancement as ocr_pdf_to_dark_images
            img = _enhance_for_ocr(img)

            # Same filename format as original code
            image_name = f"{base_name}_page{page_num+1}.png"
            image_path = os.path.join(output_dir, image_name)
            img.save(image_path, "PNG", quality=95)
            image_paths.append(image_path)

            logger.info(f"Saved PDF page {page_num+1} as image: {image_path}")

        doc.close()
        return image_paths

    except Exception as e:
        logger.error(f"PDF conversion failed for {pdf_path}: {e}", exc_info=True)
        raise


# image_path="E17-A_A4010-0001_08_AB_2020-09-11_02_page1.png"
# output_dir="outputs"
layout_model="layout.pt"
component_model_1="large_comp_detection.pt"
component_model_2="yolo_model_with_tag.pt"
line_model="Line_model.pt"
# class_names=class_names_list
# neo4j_uri="neo4j+ssc://b491cc6f.databases.neo4j.io"
# neo4j_user="neo4j"
# neo4j_pwd="zwUACNHJtJwnUw2tKVncwkWNigGYsFEtSD4HiwoGFmg"

# Suchit neo4j creds
# neo4j_uri="neo4j+ssc://11eed1fb.databases.neo4j.io"
# neo4j_user="neo4j"
# neo4j_pwd="3vOY6r_IXz3p9QA9oLeq7QrCnMiPaDYvdS-pwkW7l_E"


@app.route('/api/Home/imageprocess', methods=["POST"])

def imageprocess():
    try:
        # ------------------------------------
        # 1. GET IMAGE FROM REQUEST
        # ------------------------------------
        file = request.files.get("image")
        if not file:
            return jsonify({"error": "No image file provided"}), 400

        filename = file.filename.replace(".png", "").replace(".jpg", "")
        temp_dir = tempfile.gettempdir()
        input_path = os.path.join(temp_dir, file.filename)
        file.save(input_path)

        logger.info(f"Uploaded image saved to: {input_path}")

        # ------------------------------------
        # 2. IF PDF → CONVERT FIRST PAGE TO PNG
        # ------------------------------------
        if filename.lower().endswith(".pdf"):
            image_paths = pdf_to_images_pid(input_path, output_dir=temp_dir)
            if not image_paths:
                return jsonify({"error": "PDF → Image conversion failed"}), 500

            input_path = image_paths[0]
            filename = os.path.basename(input_path).replace(".png", "")

        # ------------------------------------
        # 3. RUN THE NEW PIPELINE (YOUR UPDATED CODE)
        # ------------------------------------
        # components, lines, connections,overlay_image = run_pipeline(
        #     input_path,
        #     P_and_D_OUTPUT_DIR,
        #     layout_model,
        #     component_model,
        #     line_model,
        #     class_names,
        #     neo4j_uri,
        #     neo4j_user,
        #     neo4j_pwd
        # )

        # Run pipeline
        components, lines, connections,overlay_image = run_pipeline(
            input_path, P_and_D_OUTPUT_DIR,
            layout_model, component_model_1,component_model_2, line_model,
            class_names_list_1,class_names_list_2,
            NEO4J_URI, NEO4J_USER, NEO4J_PASSWORD,filename
    )

        log("Pipeline finished.")

        logger.info(f"Pipeline completed for {filename}")
        logger.info(f"Components: {len(components)}, Lines: {len(lines)}, Connections: {len(connections)}")

        


        # ------------------------------------
        # 4. SAVE FINAL PIPELINE OUTPUT AS IMAGE
        # ------------------------------------
        # final_image_path = os.path.join("images", f"{filename}_final_processed.png")
        # cv2.imwrite(final_image_path, cv2.imread(input_path))

        # ------------------------------------
        # 5. OCR TEXT EXTRACTION
        # ------------------------------------
        # Component_Text_Extraction(
        #     model_name="gpt-4o",
        #     full_image_path=final_image_path,
        #     filename=filename
        # )

        # json_file_path = os.path.join(P_and_D_OUTPUT_DIR, f"{filename}_Json_results_OCR.json")
        # process_json(json_file_path)

        logger.info("OCR + Tag assignment completed")

        # ------------------------------------
        # 6. RELATIONSHIP DETECTION (OCR TEXT + YOLO LINES)
        # ------------------------------------
        original_image = cv2.imread(input_path)
        img_area = original_image.shape[0] * original_image.shape[1]

        if img_area < 2048**2:
            tolerance = 0.015
        elif img_area > 4096**2:
            tolerance = 0.03
        else:
            tolerance = 0.02

        # finding_relationships(
        #     os.path.join("images", f"{filename}_Lines_detect.json"),
        #     os.path.join("images", f"{filename}_Json_results_OCR.json"),
        #     "images",
        #     filename,
        #     tolerance=tolerance
        # )

        logger.info("Relationship mapping completed")

        # ------------------------------------
        # 7. SAVE FINAL RESULTS TO DATABASE + NEO4J
        # ------------------------------------
        # P_and_ID_insert_extracted_data_database(filename)
        # sending_json_to_neo4j(filename)

        logger.info("Snowflake + Neo4j update completed")

        # ------------------------------------
        # 8. RETURN PROCESSED IMAGE TO CLIENT
        # ------------------------------------
        return send_image_array(overlay_image)

    except Exception as e:
        logger.error(f"Processing failed: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500



# ----------------------- Tag Registry Code ---------------------------------

# Get Current Time in EST
def get_est_now():
    utc_now = datetime.utcnow().replace(tzinfo=pytz.utc)
    est_tz = pytz.timezone('America/New_York')
    est_now = utc_now.astimezone(est_tz)
    return est_now

# columns Mapping
equipment_column_mapping = {
    'status': 'STATUS','type': 'TYPE','tagNo': 'TAG_NO','asset': 'ASSET','function': 'FUNCTION','functionDescription': 'FUNCTION_DESCRIPTION',
    'sequence': 'SEQUENCE','suffix1': 'SUFFIX_1','suffix2': 'SUFFIX_2','description': 'DESCRIPTION','specification': 'SPECIFICATION','pedGrpCatMod': 'PED_GRP_CAT_MOD',
    'supplier': 'SUPPLIER','flowRateNm3H': 'FLOW_RATE_NM3_HR','capacityM3': 'CAPACITY_M3','dutyKw': 'DUTY_KW','designPressure': 'DESIGN_PRESSURE_BARG',
    'testPressBarg': 'TEST_PRESS_BARG','designTemperature': 'DESIGN_TEMPERATURE_DEGC','medium': 'MEDIUM','powerKw': 'POWER_KW',
    'pimsDocument': 'PIMS_DOCUMENT','document': 'DOCUMENT','relatedTags':'RELATED_TAGS','maximoWo':'MAXIMO_WO','relatedTags1': 'RELATED_TAGS_1','relatedTags2': 'RELATED_TAGS_2','relatedTags3': 'RELATED_TAGS_3',
    'lengthTTMm': 'LENGTH_T_T_MM', 'widthMm': 'WIDTH_MM', 'heightMm': 'HEIGHT_MM','diaIdMm': 'DIA_ID_MM','dryWtTe': 'DRY_WT_TE','operationalWeightTe': 'OPERATIONAL_WEIGHT_TE',
    'dateUpdated': 'DATE_UPDATED','modifiedBy':'MODIFIED_BY','modifiedDate':'MODIFIED_DATE','remarks': 'REMARKS','file_Type': 'FILE_TYPE','url': 'URL','file_Name': 'FILE_NAME','file_Description': 'FILE_DESCRIPTION',
}

instrument_column_mapping = {
    'status': 'STATUS','type': 'TYPE','tagNo': 'TAG_NO', 'asset': 'ASSET','function': 'FUNCTION', 'instrumentType': 'INSTRUMENT_TYPE_FUNCTION_DESCRIPTION',
    'sequence': 'SEQUENCE',  'suffix1': 'SUFFIX_1', 'suffix2': 'SUFFIX_2','description': 'DESCRIPTION', 'loopId': 'LOOP_ID','deviceType': 'DEVICE_TYPE','atex': 'ATEX',
    'deckModuleRoom': 'DECK_MODULE_ROOM', 'packagePanel': 'PACKAGE_PANEL','system': 'ICSS_SYSTEM','cabinet': 'ICSS_CABINET','ioType': 'ICSS_IO_TYPE','signalType': 'ICSS_SIGNAL_TYPE',
    'pimsDocument':'PIMS_DOCUMENT','document':'DOCUMENT','relatedTags': 'RELATED_TAGS','ceDiagram':'CE_DIAGRAM', 'locationDrawing': 'LOCATION_DRAWING','layout':'LAYOUT','loopDrawing': 'LOOP_DRAWING','dataSheet': 'DATA_SHEET','hookupDrawing1': 'HOOKUP_DRAWING_1','hookupDrawing2': 'HOOKUP_DRAWING_2','specification': 'SPECIFICATION',
    'maximoWo': 'MAXIMO_WO','manufacturer': 'MANUFACTURER','model': 'MODEL','calibratedRange': 'CALIBRATED_RANGE','unit': 'UNIT','ll': 'TRIP_LL','maximoWo':'MAXIMO_WO',
    'l': 'ALARM_L','ctrL': 'CTR_L','ctrN': 'CTR_N','ctrH': 'CTR_H','h': 'ALARM_H','hh': 'TRIP_HH','range': 'OUTPUT_RANGE_MA','setPointL': 'OUTPUT_SET_POINT_L_MA',
    'setPointH': 'OUTPUT_SET_POINT_H_MA', 'file_Type': 'FILE_TYPE','url': 'URL','file_Name': 'FILE_NAME','file_Description': 'FILE_DESCRIPTION','dateUpdated': 'DATE_UPDATED',
    'remarks': 'REMARKS'
}

line_column_mapping = {
    'status': 'STATUS','tagType': 'TAG_TYPE','tagNo': 'TAG_NO','asset': 'ASSET','diameter': 'DIAMETER','classSpecification': 'CLASS_SPECIFICATION','fluidCode': 'FLUID_CODE_SERVICE',
    'sequence': 'SEQUENCE','insulation': 'INSULATION','insulationThicknessMm': 'INSULATION_THICKNESS_MM','fromLocation': 'FROM_LOCATION','toLocation': 'TO_LOCATION',
    'phase': 'OPERATION_PHASE','operationPressureBarg': 'OPERATION_PRESSURE_BARG','operationTemperatureDegc': 'OPERATION_TEMPERATURE_DEGC','assessmentVapPressureBarg': 'ASSESSMENT_VAP_PRESSURE_BARG',
    'assessmentMinDesTempDegc': 'ASSESSMENT_MIN_DES_TEMP_DEGC','assessmentMaxDesTempDegc': 'ASSESSMENT_MAX_DES_TEMP_DEGC','assessmentDesPressBarg': 'ASSESSMENT_DES_PRESS_BARG',
    'wtOrSch': 'PIPE_WT_OR_SCH','pedCritical': 'PED_CRITICAL','hazardousCategory': 'HAZARDOUS_CATEGORY','pedTable': 'PED_TABLE','pedCat': 'PED_CAT','pedMod': 'PED_MOD',
    'pedGroup': 'PED_GROUP','kviTable': 'KVI_TABLE','kvlYN': 'KVL_Y_N','testPressure': 'TEST_PRESSURE','inspectionNumber': 'INSPECTION_NUMBER',
    'pimsDocument': 'PIMS_DOCUMENT','pid2': 'PID_2','pid3':'PID_3','document':'DOCUMENT','relatedTags': 'RELATED_TAGS',
    'file_Type': 'FILE_TYPE','url': 'URL','file_Name': 'FILE_NAME','file_Description': 'FILE_DESCRIPTION','dateUpdated': 'DATE_UPDATED',
    'remarks': 'REMARKS'
}

cable_column_mapping = {
    'status': 'STATUS','tagType': 'TAG_TYPE','tagNo': 'TAG_NO','asset': 'ASSET','prefix': 'PREFIX','sequence': 'SEQUENCE',
    'suffix1': 'SUFFIX_1','suffix2': 'SUFFIX_2','description': 'DESCRIPTION','fromLocation': 'FROM_LOCATION','toLocation': 'TO_LOCATION',
    'size': 'SIZE','colour': 'COLOUR','length': 'LENGTH','pimsDocument':'PIMS_DOCUMENT','document':'DOCUMENT','relatedTags': 'RELATED_TAGS','instrumentDrawing1':'INSTRUMENT_DRAWING_1','instrumentDrawing2':'INSTRUMENT_DRAWING_2','nrpaca':'NRPACA','nrcoca':'NRCOCA',
    'dateUpdated': 'DATE_UPDATED','file_Type': 'FILE_TYPE','url': 'URL','file_Name': 'FILE_NAME','file_Description': 'FILE_DESCRIPTION','remarks': 'REMARKS'
}

# Table Mapping
TABLE_MAPPING = {
    'Equipment': {
        'main': 'ENI_NL_DEV_TAGREGISTRY.DEV.TBL_EQUIPMENT_TAGS',
        'history': 'ENI_NL_DEV_TAGREGISTRY.DEV.TBL_EQUIPMENT_TAGS_HISTORY',
        'mapping': equipment_column_mapping
    },
    'Instrument': {
        'main': 'ENI_NL_DEV_TAGREGISTRY.DEV.TBL_INSTRUMENT_TAGS',
        'history': 'ENI_NL_DEV_TAGREGISTRY.DEV.TBL_INSTRUMENT_TAGS_HISTORY',
        'mapping': instrument_column_mapping
    },
    'Line': {
        'main': 'ENI_NL_DEV_TAGREGISTRY.DEV.TBL_LINE_TAGS',
        'history': 'ENI_NL_DEV_TAGREGISTRY.DEV.TBL_LINE_TAGS_HISTORY',
        'mapping': line_column_mapping
    },
    'Cable': {
        'main': 'ENI_NL_DEV_TAGREGISTRY.DEV.TBL_CABLE_TAGS',
        'history': 'ENI_NL_DEV_TAGREGISTRY.DEV.TBL_CABLE_TAGS_HISTORY',
        'mapping': cable_column_mapping
    },
    'Cable - Electrical': {   
        'main': 'ENI_NL_DEV_TAGREGISTRY.DEV.TBL_CABLE_TAGS',
        'history': 'ENI_NL_DEV_TAGREGISTRY.DEV.TBL_CABLE_TAGS_HISTORY',
        'mapping': cable_column_mapping
    },
    'Cable - Instrument': {   
        'main': 'ENI_NL_DEV_TAGREGISTRY.DEV.TBL_CABLE_TAGS',
        'history': 'ENI_NL_DEV_TAGREGISTRY.DEV.TBL_CABLE_TAGS_HISTORY',
        'mapping': cable_column_mapping
    }
}

#_______________user managment____________________________

# Get User By Email API
@app.route('/api/Home/GetUserByEmail', methods=['GET'])
def get_user_by_email_endpoint():
    """
    Retrieve user details by email address.
    """
    email = request.args.get('email')
    logger.info("Received GetUserByEmail request.")
    if not email:
        logger.info("Email query parameter is missing.")
        return jsonify({"error": "Email query parameter is required."}), 400

    try:
        conn = get_snowflake_connection()
        with conn.cursor() as cursor:
            # query = (
            #     'SELECT "email", "roleType", "isActive" '
            #     'FROM ENI_NL_DEV_TAGREGISTRY.DEV.TBL_USERS '
            #     'WHERE "email" = %s AND "isActive" = 1'
            # )
            query = (
                'SELECT "email", "roleType", "isActive" '
                'FROM ENI_NL_DEV_TAGREGISTRY.DEV.TBL_USERS '
                'WHERE LOWER("email") = LOWER(%s) AND "isActive" = 1'
            )
            cursor.execute(query, (email,))
            rows = cursor.fetchall()
            column_names = [desc[0] for desc in cursor.description]
            user_list = [dict(zip(column_names, row)) for row in rows]

        logger.info(f"User lookup for email '{email}' returned {len(user_list)} result(s).")
        return jsonify(user_list), 200

    except Exception as ex:
        logger.info(f"Failed to fetch user data: {ex}", exc_info=True)
        return jsonify({"error": "Internal server error", "details": str(ex)}), 500

# Add User API
@app.route('/api/Home/AddUser', methods=['POST'])
def add_user_endpoint():
    """
    Adds a new user to the system if the email does not already exist.
    """
    data = request.get_json()
    user_data = data.get('currentUser')

    if not user_data or 'email' not in user_data or 'roleType' not in user_data:
        logger.info("Invalid request body: missing email or roleType.")
        return jsonify({"success": False, "error": "Invalid request body. Requires email and roleType."}), 400

    email = user_data.get('email').lower()
    role_type = user_data.get('roleType')
    try:
        # Use context manager for DB connection and cursor
        conn = get_snowflake_connection()
        with conn.cursor() as cursor:
            # Check if the email already exists and is active
            check_query = (
                'SELECT "email" FROM ENI_NL_DEV_TAGREGISTRY.DEV.TBL_USERS '
                'WHERE "email" = %s AND "isActive" = 1'
            )
            cursor.execute(check_query, (email,))
            existing_user = cursor.fetchone()

            if existing_user:
                logger.info(f"Attempted to add existing user: {email}")
                return jsonify({"success": False, "message": "Email already exists."}), 200

            # Insert the user if not exists
            insert_query = (
                'INSERT INTO ENI_NL_DEV_TAGREGISTRY.DEV.TBL_USERS("email", "roleType", "isActive") '
                'VALUES (%s, %s, 1)'
            )
            cursor.execute(insert_query, (email, role_type))
            conn.commit()
            logger.info(f"User {email} added successfully with role {role_type}.")

        return jsonify({"success": True}), 200

    except Exception as ex:
        logger.info(f"Failed to add user: {ex}", exc_info=True)
        return jsonify({"success": False, "error": str(ex)}), 400
  
# Edit User API  
@app.route('/api/Home/EditUser', methods=['PUT'])
def edit_user_endpoint():
    """
    Updates an existing user's roleType in the system.
    """
    user_data = request.get_json()
    if not user_data or 'email' not in user_data or 'roleType' not in user_data:
        logger.info("Invalid request body: missing email or roleType.")
        return jsonify({"success": False, "error": "Invalid request body. Requires email and roleType."}), 400

    email = user_data.get('email')
    role_type = user_data.get('roleType')
    try:
        # Use context manager for DB connection and cursor
        conn = get_snowflake_connection()
        with conn.cursor() as cursor:
            update_query = (
                'UPDATE ENI_NL_DEV_TAGREGISTRY.DEV.TBL_USERS '
                'SET "roleType" = %s '
                'WHERE "email" = %s AND "isActive" = 1'
            )
            cursor.execute(update_query, (role_type, email))
            conn.commit()
            if cursor.rowcount == 0:
                logger.info(f"No active user found with email: {email}")
                return jsonify({"success": False, "error": "User not found or not active."}), 404

        logger.info(f"User {email} updated successfully to role {role_type}.")
        return jsonify({"success": True}), 200

    except Exception as ex:
        logger.info(f"Failed to edit user: {ex}", exc_info=True)
        return jsonify({"success": False, "error": "Internal server error", "details": str(ex)}), 500
 
# Delete User API   
@app.route('/api/Home/DeleteUser', methods=['POST'])
def delete_user_endpoint():
    """
    Deactivates (soft deletes) a user by setting isActive=0 for the given email.
    """
    email = request.args.get('email')
    if not email:
        logger.info("Email query parameter is missing.")
        return jsonify({"success": False, "error": "Email query parameter is required."}), 400

    try:
        conn = get_snowflake_connection()
        with conn.cursor() as cursor:
            # Soft delete: set isActive=0 instead of hard delete
            update_query = (
                'UPDATE ENI_NL_DEV_TAGREGISTRY.DEV.TBL_USERS '
                'SET "isActive" = 0 '
                'WHERE "email" = %s AND "isActive" = 1'
            )
            cursor.execute(update_query, (email,))
            conn.commit()
            if cursor.rowcount == 0:
                logger.info(f"No active user found with email: {email}")
                return jsonify({"success": False, "error": "User not found or already inactive."}), 404

        logger.info(f"User {email} deactivated successfully.")
        return jsonify({"success": True}), 200

    except Exception as ex:
        logger.info(f"Failed to deactivate user: {ex}", exc_info=True)
        return jsonify({"success": False, "error": "Internal server error", "details": str(ex)}), 500
 
# Get All Users API       
@app.route('/api/Home/GetAllUser', methods=['GET'])
def get_all_user_endpoint():
    """
    Retrieves all active users from the database.
    """
    try:
        conn = get_snowflake_connection()
        with conn.cursor() as cursor:
            query = (
                'SELECT DISTINCT "email", "roleType", "isActive" '
                'FROM ENI_NL_DEV_TAGREGISTRY.DEV.TBL_USERS '
                'WHERE "isActive" = 1'
                'ORDER BY "roleType", LOWER("email")'
            )
            cursor.execute(query)
            rows = cursor.fetchall()
            user_list = convert_rows_to_list_of_approval_models(cursor, rows)
        logger.info(f"Fetched {len(user_list)} active users from the database.")
        return jsonify(user_list), 200

    except Exception as ex:
        logger.info(f"Failed to fetch all users: {ex}", exc_info=True)
        return jsonify({"error": "Internal server error", "details": str(ex)}), 500

#_______________tag managment____________________________

# Admin Insert API
@app.route('/api/Home/insert', methods=['POST'])
def insert_tag_endpoint():
    """
    Inserts a new tag record into the appropriate table based on tagType.
    If tag exists and action = 'Link to Drawing' → append document fields.
    """
    model_data = request.get_json() or {}
    logger.debug(f"Received request data: {model_data}")

    tag_no = model_data.get('tagNo')
    tag_type = model_data.get('tagtype') or model_data.get('tagType')
    action = model_data.get('action') or model_data.get('Action')

    if not tag_no or not tag_type or tag_type not in TABLE_MAPPING:
        logger.info("Missing or invalid tagNo/tagType.")
        return ("Tag_NO and valid tagType are required."), 400

    target_table = TABLE_MAPPING[tag_type]['main']
    column_mapping = TABLE_MAPPING[tag_type]['mapping']
    logger.info(f"Target table: '{target_table}'")

    try:
        conn = get_snowflake_connection()
        with conn.cursor() as cursor:

            # -------------------------------
            # 1️⃣ CHECK IF TAG ALREADY EXISTS
            # -------------------------------
            cursor.execute(f"SELECT COUNT(*) FROM {target_table} WHERE TAG_NO = %s", (tag_no,))
            exists = cursor.fetchone()[0]

            # -------------------------------------------------------
            # 2️⃣ HANDLE SPECIAL CASE → LINK TO DRAWING (APPEND MODE)
            # -------------------------------------------------------
            if exists > 0 and action and action.lower().strip() == "link to drawing":
                logger.info(f"Tag '{tag_no}' exists → appending document info (Link to Drawing)")

                # Fetch existing row
                cursor.execute(f"""
                    SELECT PIMS_DOCUMENT, FILE_NAME, FILE_DESCRIPTION, FILE_TYPE, URL
                    FROM {target_table}
                    WHERE TAG_NO = %s
                """, (tag_no,))
                row = cursor.fetchone()

                existing_doc, existing_file, existing_desc, existing_type, existing_url = row

                # Incoming values
                new_doc  = model_data.get("pimsDocument")
                new_file = model_data.get("file_Name")
                new_desc = model_data.get("file_Description")
                new_type = model_data.get("file_Type")
                new_url  = model_data.get("url")

                # Helper → append without duplicates
                def merge(existing, new):
                    if not new: 
                        return existing
                    if not existing:
                        return new
                    parts = [x.strip() for x in existing.split(";") if x.strip()]
                    if new not in parts:
                        return existing + "; " + new
                    return existing

                merged_doc  = merge(existing_doc, new_doc)
                merged_file = merge(existing_file, new_file)
                merged_desc = merge(existing_desc, new_desc)
                merged_type = merge(existing_type, new_type)
                merged_url  = merge(existing_url, new_url)

                update_sql = f"""
                    UPDATE {target_table}
                    SET 
                        PIMS_DOCUMENT    = %s,
                        FILE_NAME        = %s,
                        FILE_DESCRIPTION = %s,
                        FILE_TYPE        = %s,
                        URL              = %s,
                        MODIFIED_BY      = %s,
                        MODIFIED_DATE    = CURRENT_TIMESTAMP()
                    WHERE TAG_NO = %s
                """

                cursor.execute(update_sql, (
                    merged_doc, merged_file, merged_desc, merged_type, merged_url,
                    model_data.get("userEmail"), tag_no
                ))

                conn.commit()

                return jsonify({
                    "success": True,
                    "message": "Document appended (Link to Drawing)",
                    "mode": "append"
                }), 200

            # -------------------------------------------------------
            # 3️⃣ TAG EXISTS BUT ACTION IS NOT LINK TO DRAWING
            # -------------------------------------------------------
            if exists > 0:
                logger.info(f"Tag '{tag_no}' already exists → standard duplicate handling")
                return jsonify({
                    "success": False,
                    "message": f"Tag '{tag_no}' already exists. Change tag number or edit existing tag."
                }), 409

            # -------------------------------------------
            # 4️⃣ NORMAL INSERT (TAG DOES NOT EXIST)
            # -------------------------------------------
            params_dict = {
                snowflake_col: model_data[json_key]
                for json_key, snowflake_col in column_mapping.items()
                if model_data.get(json_key) not in (None, '', ' ')
            }

            now = get_est_now()
            params_dict['DATE_UPDATED'] = now
            params_dict['MODIFIED_DATE'] = now
            params_dict['MODIFIED_BY'] = model_data.get('userEmail')

            columns = list(params_dict.keys())
            values = [params_dict[col] for col in columns]

            placeholders = ', '.join(['%s'] * len(columns))
            columns_str = ', '.join(columns)

            insert_sql = f"INSERT INTO {target_table} ({columns_str}) VALUES ({placeholders})"
            cursor.execute(insert_sql, values)
            conn.commit()

        logger.info(f"Successfully inserted tag '{tag_no}' into '{target_table}'.")
        return jsonify({
            "success": True,
            "message": f"Tag '{tag_no}' inserted successfully.",
            "mode": "insert",
            "modelData": model_data
        }), 200

    except Exception as ex:
        logger.error(f"Error inserting tag: {ex}", exc_info=True)
        try:
            conn.rollback()
        except:
            pass

        return jsonify({"success": False, "error": str(ex)}), 500

# Contributor Insert API
@app.route('/api/Home/ContributorInsert', methods=['POST'])
def contributor_insert_endpoint():
    """
    Inserts a new contributor record into the appropriate history table based on tagType.
    """
    try:
        logger.info('Contributor Insert API called.')
        data = request.get_json()
        form_data = data.get('formData')
        if not form_data:
            return jsonify({'error': 'formData is required'}), 400

        tag_no = form_data.get('TAG_NO') or form_data.get('tagNo')
        if not tag_no:
            return jsonify({'error': 'TAG_NO is mandatory and cannot be empty'}), 400

        tag_type = form_data.get('tagType')
        if not tag_type or tag_type not in TABLE_MAPPING:
            return jsonify({'error': 'Invalid tagType'}), 400

        table_info = TABLE_MAPPING[tag_type]
        target_table = table_info['main']
        table_name = table_info['history']
        mapping = table_info['mapping']

        columns, values = [], []

        for key, column in mapping.items():
            value = form_data.get(key) or form_data.get(column)
            if key in ('dateUpdated', 'modifiedDate', 'modifieddate', 'MODIFIED_DATE'):
                continue  
            if column == 'TAG_NO':
                if not value:
                    return jsonify({'error': 'TAG_NO is mandatory and cannot be empty'}), 400
                columns.append(column)
                values.append(value)
            elif value not in (None, '', ' '):
                columns.append(column)
                values.append(value)

        # Add audit columns
        columns.extend(['DATE_UPDATED', 'MODIFIED_DATE', 'MODIFIED_BY'])
        now = get_est_now()
        values.extend([now, now, data.get('userEmail')])

        # Add action column
        action = data.get('action')
        if not action:
            return jsonify({'error': 'Action is required'}), 400
        columns.append('ACTION')
        values.append(action)
        
        # ✅ ADD THIS BLOCK - Set APPROVAL_STATUS to PENDING
        columns.append('APPROVAL_STATUS')
        values.append('PENDING')

        # Connect to Snowflake
        conn = get_snowflake_connection()
        try:
            with conn.cursor() as cursor:
                # ✅ Check if tag already exists in approval table with PENDING status
                check_sql = f"""
                    SELECT COUNT(*) 
                    FROM {table_name} 
                    WHERE TAG_NO = %s AND APPROVAL_STATUS = 'PENDING'
                """
                cursor.execute(check_sql, (tag_no,))
                count = cursor.fetchone()[0]

                if count > 0:
                    logger.info(f"Tag '{tag_no}' already exists in approval queue (Pending).")
                    return jsonify({
                        "success": False,
                        "message": "Tag already exists in the approvals queue."
                    }), 200

                # ✅ Otherwise safe to insert (REJECTED or new tags allowed)
                placeholders = ', '.join(['%s'] * len(columns))
                column_str = ', '.join(columns)
                sql = f"INSERT INTO {table_name} ({column_str}) VALUES ({placeholders})"
                cursor.execute(sql, values)

            conn.commit()
            
            # INSERT APPROVAL STATUS AS "PENDING" - NEW CODE STARTS HERE
            # approval_result = Inserting_Approval_Status_Data(
            #     tag_id=tag_no,
            #     status="Pending",  # Set as Pending when contributor submits
            #     comment="",
            #     action=action,
            #     admin_email="",  # Empty since not yet approved
            #     tag_type=tag_type
            # )

            # if not approval_result["success"]:
            #     logger.info(f"Failed to insert approval status: {approval_result['error']}")
            #     # Note: We don't rollback the history insert since it was already committed
            #     # You might want to handle this differently based on your business logic
            # else:
            #     logger.info(f"Approval status created as 'Pending' for TAG_ID: {tag_no}")
            # # NEW CODE ENDS HERE
            
        except Exception as db_ex:
            conn.rollback()
            logger.info(f"Snowflake error: {db_ex}", exc_info=True)
            return jsonify({'error': str(db_ex)}), 500
        finally:
            conn.close()

        return '', 200

    except Exception as e:
        logger.info(f"Unexpected error: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

# Approval Status Insertion Function
def Inserting_Approval_Status_Data(tag_id, status, comment, action, admin_email, tag_type):
    """
    Inserts approval status data into Snowflake table: TBL_TAG_REGISTRY_STATUS.
    """
    if not tag_id or not tag_type or tag_type not in TABLE_MAPPING:
        logger.info(f"Invalid input: tag_id={tag_id}, tag_type={tag_type}")
        return {"success": False, "error": "Invalid tag_id or tag_type."}

    con_email = ''
    now = get_est_now()
    created_date = now
    modified_date = None

    try:
        history_table = TABLE_MAPPING[tag_type]['history']
        with get_snowflake_connection() as conn, conn.cursor() as cursor:
            history_query = f"""
                SELECT MODIFIED_BY, MODIFIED_DATE
                FROM {history_table}
                WHERE TAG_NO = %s
                ORDER BY DATE_UPDATED DESC
                LIMIT 1
            """
            cursor.execute(history_query, (tag_id,))
            row = cursor.fetchone()
            if row:
                con_email = row[0] or ''
                modified_date = row[1]

            insert_query = """
                INSERT INTOENI_NL_DEV_TAGREGISTRY.DEV.TBL_TAG_REGISTRY_STATUS
                (TAG_ID, ACTION, CREATED_DATE, STATUS, ADMIN, APPROVED_DATE, REJECTION_COMMENT, REQUESTED_BY, TAG_TYPE)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            cursor.execute(
                insert_query,
                (tag_id, action, created_date, status, admin_email, modified_date, comment, con_email, tag_type)
            )
            conn.commit()
            logger.info(f"Approval status inserted for TAG_ID {tag_id}, TAG_TYPE {tag_type}")
            return {"success": True, "message": f"Approval status inserted for TAG_ID {tag_id}"}

    except snowflake.connector.errors.DatabaseError as db_ex:
        logger.info(f"Database error inserting approval status for TAG_ID {tag_id}: {str(db_ex)}", exc_info=True)
        return {"success": False, "error": f"Database error: {str(db_ex)}"}
    except Exception as ex:
        logger.info(f"Unexpected error inserting approval status for TAG_ID {tag_id}: {str(ex)}", exc_info=True)
        return {"success": False, "error": f"Unexpected error: {str(ex)}"}

# Contributor Approval Data Retrieval Endpoint 
@app.route('/api/Home/ContributorApprovalData', methods=['GET'])
def get_contributor_approval_data():
    """
    Retrieves paginated approval records for a given user from Snowflake TBL_TAG_REGISTRY_STATUS.
    """
    email = request.args.get('user')
    if not email:
        logger.info("Missing 'user' query parameter.")
        return jsonify({"error": "User email parameter is required."}), 400

    try:
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('pageSize', 100))
        if page < 1 or page_size < 1:
            logger.info("Invalid page or pageSize: must be positive integers.")
            return jsonify({"error": "Page and pageSize must be positive integers."}), 400

        offset = (page - 1) * page_size

        with get_snowflake_connection() as conn, conn.cursor() as cursor:
            query = """
                SELECT ACTION,TAG_ID,TAG_TYPE,REQUESTED_BY,CREATED_DATE,STATUS,
                    REJECTION_COMMENT,ADMIN,APPROVED_DATE,
                    COUNT(*) OVER() as total_count
                FROM ENI_NL_DEV_TAGREGISTRY.DEV.TBL_TAG_REGISTRY_STATUS
                WHERE REQUESTED_BY = %s
                ORDER BY CREATED_DATE DESC
                LIMIT %s OFFSET %s
            """
            cursor.execute(query, (email, page_size, offset))
            rows = cursor.fetchall()

            if not rows:
                logger.info(f"No approval records found for user: {email}")
                return jsonify({"data": [], "totalCount": 0}), 200

            total_count = rows[0][-1]
            data_rows = [row[:-1] for row in rows]
            approvals_list = convert_rows_to_list_of_approval_models(cursor, data_rows)

            logger.info(f"Fetched {len(approvals_list)} approval records for user: {email} (total: {total_count})")
            return jsonify({"data": approvals_list, "totalCount": total_count}), 200

    except snowflake.connector.errors.DatabaseError as db_ex:
        logger.info(f"Database error fetching approval data for user {email}: {str(db_ex)}", exc_info=True)
        return jsonify({"error": "Database error", "details": str(db_ex)}), 500
    except Exception as ex:
        logger.info(f"Unexpected error fetching approval data for user {email}: {str(ex)}", exc_info=True)
        return jsonify({"error": "Internal server error", "details": str(ex)}), 500
  
# Approval Tag Retrieval Endpoint  
@app.route('/api/Home/GetApprovalTag', methods=['GET'])
def get_approval_tag():
    """
    Retrieves approval records for a specific tag.
    Only returns records with PENDING status to check if tag is in approval queue.
    """
    try:
        tag_no = request.args.get('tagNo')
        tag_type = request.args.get('tagType')
        
        if not tag_no:
            logger.info("Missing 'tagNo' query parameter.")
            return jsonify({"error": "tagNo parameter is required."}), 400

        if not tag_type or tag_type not in TABLE_MAPPING:
            logger.info(f"Invalid or missing tagType: {tag_type}")
            return jsonify({"error": "Valid tagType parameter is required."}), 400

        table_name = TABLE_MAPPING[tag_type]['history']

        with get_snowflake_connection() as conn, conn.cursor() as cursor:
            # ✅ FIXED: Only return records with PENDING status
            # This prevents REJECTED/APPROVED tags from blocking resubmission
            query = f"""
                SELECT *
                FROM {table_name}
                WHERE TAG_NO = %s AND APPROVAL_STATUS = 'PENDING'
                ORDER BY MODIFIED_DATE DESC
            """
            cursor.execute(query, (tag_no,))
            rows = cursor.fetchall()

            if not rows:
                logger.info(f"No PENDING approval records found for tag: {tag_no}")
                return jsonify({"data": []}), 200

            approvals_list = convert_rows_to_list_of_approval_models(cursor, rows)
            logger.info(f"Found {len(approvals_list)} PENDING approval record(s) for tag: {tag_no}")
            
            return jsonify({"data": approvals_list}), 200

    except snowflake.connector.errors.DatabaseError as db_ex:
        logger.info(f"Database error fetching approval tag {tag_no}: {str(db_ex)}", exc_info=True)
        return jsonify({"error": "Database error", "details": str(db_ex)}), 500
    except Exception as ex:
        logger.info(f"Unexpected error fetching approval tag {tag_no}: {str(ex)}", exc_info=True)
        return jsonify({"error": "Internal server error", "details": str(ex)}), 500

# Contributor Delete Endpoint (Updated to use UPDATE instead of DELETE)
@app.route('/api/Home/ContributorDelete', methods=['DELETE'])
def contributor_delete_tag_endpoint():
    """
    Updates the APPROVAL_STATUS of a contributor record instead of deleting it.
    """
    app.logger.info("Contributor Delete API called.")
    try:
        tag_no = request.args.get('tagNo')
        tag_type = request.args.get('tagType')
        status = request.args.get('Status')  # 'Approved' or 'Rejected'
        AdminEmail = request.args.get('user')
        comment = request.args.get('Comment')
        action = request.args.get('Action')
        
        # Insert approval status data (keep this as is)
        Inserting_Approval_Status_Data(tag_no, status, comment, action, AdminEmail, tag_type)

        if not tag_no:
            app.logger.info("tagNo is mandatory and cannot be empty.")
            return jsonify({'error': 'tagNo is mandatory and cannot be empty'}), 400

        if not tag_type or tag_type not in TABLE_MAPPING:
            app.logger.info(f"Invalid or missing tagType: {tag_type}")
            return jsonify({'error': 'Invalid or missing tagType'}), 400

        table_info = TABLE_MAPPING[tag_type]
        table_name = table_info['history']
        mapping = table_info['mapping']

        tag_column = mapping.get('tagNo')
        if not tag_column:
            app.logger.info("TAG column mapping not found in TABLE_MAPPING.")
            return jsonify({'error': 'TAG column mapping not found in TABLE_MAPPING'}), 500

        conn = get_snowflake_connection()
        with conn.cursor() as cursor:
            # ✅ CHANGE: UPDATE instead of DELETE
            if status == 'Approved':
                # For approved records, DELETE them (they're moved to main table)
                delete_query = f'DELETE FROM {table_name} WHERE "{tag_column}" = %s'
                cursor.execute(delete_query, (tag_no,))
                rows_affected = cursor.rowcount
            else:
                # For rejected records, UPDATE the APPROVAL_STATUS to 'REJECTED'
                update_query = f'''
                    UPDATE {table_name} 
                    SET APPROVAL_STATUS = %s,
                        COMMENTS = %s,
                        MODIFIED_BY = %s,
                        MODIFIED_DATE = CURRENT_TIMESTAMP()
                    WHERE "{tag_column}" = %s
                '''
                cursor.execute(update_query, ('REJECTED', comment, AdminEmail, tag_no))
                rows_affected = cursor.rowcount
            
            conn.commit()

        app.logger.info(f"Updated/Deleted {rows_affected} row(s) in {table_name} for tagNo: {tag_no} with status: {status}")
        return jsonify({'success': rows_affected > 0}), 200

    except Exception as ex:
        app.logger.info(f"Error during contributor delete: {ex}", exc_info=True)
        return jsonify({'success': False, 'error': str(ex)}), 500

# Data Retrieval Helper Functions
def to_camel_case(snake_str):
    components = snake_str.split('_')
    return components[0].lower() + ''.join(x.title() for x in components[1:])

# Data Retrieval Endpoint
@app.route('/api/Home/GetData', methods=['GET'])
def get_data_endpoint():
    """
    Retrieves paginated tag data from the specified tagType table, with optional tag search.
    """
    try:
        # Parse and validate parameters upfront
        try:
            page = int(request.args.get('page', 1))
            page_size = int(request.args.get('pageSize', 10))
        except (ValueError, TypeError):
            logger.info("Non-integer page or pageSize received.")
            return jsonify({"error": "Page and pageSize must be integers."}), 400

        tag_search = request.args.get('tagSearch')
        tag_type = request.args.get('tagType')

        # Early validation
        if not tag_type or tag_type not in TABLE_MAPPING:
            logger.info(f"Invalid or missing tagType: {tag_type}")
            return jsonify({"error": "Invalid or missing tagType."}), 400

        if page < 1 or page_size < 1:
            logger.info("Page and pageSize must be positive integers.")
            return jsonify({"error": "Page and pageSize must be positive integers."}), 400

        # Calculate offset once
        table_name = TABLE_MAPPING[tag_type]['main']
        offset = (page - 1) * page_size

        # Optimize query logic - fix the reversed conditions
        if tag_search:
            # Exact match when search term provided
            where_clause = "WHERE TAG_NO = %s"
            params = [tag_search]
        else:
            # Pattern match when no search term (this was reversed in original)
            where_clause = "WHERE TAG_NO ILIKE %s"
            params = [f"%{tag_search or ''}%"]

        # Use single query with window functions for better performance
        combined_query = f"""
            WITH paginated_data AS (
                SELECT *, COUNT(*) OVER() as total_count
                FROM {table_name}
                {where_clause}
                ORDER BY DATE_UPDATED DESC
                LIMIT %s OFFSET %s
            )
            SELECT * FROM paginated_data
        """
        
        query_params = params + [page_size, offset]

        with get_snowflake_connection() as conn, conn.cursor() as cursor:
            cursor.execute(combined_query, query_params)
            
            # Get column names once
            columns = [desc[0] for desc in cursor.description]
            camel_case_columns = [to_camel_case(col) for col in columns]
            
            rows = cursor.fetchall()
            
            if rows:
                total_count = rows[0][-1]  # Get total_count from window function
                # Remove total_count column from data
                data_rows = [row[:-1] for row in rows]
                data_list = [dict(zip(camel_case_columns[:-1], row)) for row in data_rows]
            else:
                total_count = 0
                data_list = []

        logger.info(f"Fetched {len(data_list)} records from {table_name} (total: {total_count}).")
        return jsonify({
            "data": data_list,
            "totalCount": total_count
        }), 200

    except Exception as ex:
        logger.info(f"Error in get_data_endpoint: {str(ex)}", exc_info=True)
        return jsonify({"error": "Internal server error", "details": str(ex)}), 500

# Approval Data Endpoint
@app.route('/api/Home/GetAllApprovals', methods=['GET'])
def get_all_approvals_endpoint():
    """
    Retrieves paginated approval records from the specified history table.
    """
    try:
        page = request.args.get('page', 1)
        page_size = request.args.get('pageSize', 100)
        tag_type = request.args.get('tagType', 'Equipment')
        approval_status = request.args.get('approvalStatus', 'pending')  # ✅ ADD THIS LINE

        try:
            page = int(page)
            page_size = int(page_size)
        except ValueError:
            logger.info("Non-integer page or pageSize received.")
            return jsonify({"error": "Page and pageSize must be integers."}), 400

        if tag_type not in TABLE_MAPPING:
            logger.info(f"Invalid tagType received: {tag_type}")
            return jsonify({"error": "Invalid tagType."}), 400

        if page < 1 or page_size < 1:
            logger.info("Page and pageSize must be positive integers.")
            return jsonify({"error": "Page and pageSize must be positive integers."}), 400

        table_name = TABLE_MAPPING[tag_type]['history']
        offset = (page - 1) * page_size

        conn = get_snowflake_connection()
        with conn.cursor() as cursor:
            # ✅ MODIFY COUNT QUERY - Add WHERE clause
            count_query = f"SELECT COUNT(*) FROM {table_name} WHERE APPROVAL_STATUS = %s"
            cursor.execute(count_query, (approval_status.upper(),))  # Use 'PENDING' or 'REJECTED'
            total_count = cursor.fetchone()[0]

            # ✅ MODIFY SELECT QUERY - Add WHERE clause
            query = f"""
                SELECT *
                FROM {table_name}
                WHERE APPROVAL_STATUS = %s
                ORDER BY MODIFIED_DATE DESC
                LIMIT %s OFFSET %s
            """
            cursor.execute(query, (approval_status.upper(), page_size, offset))
            rows = cursor.fetchall()
            approvals_list = convert_rows_to_list_of_approval_models(cursor, rows)

        logger.info(f"Fetched {len(approvals_list)} approvals from {table_name} (total: {total_count}).")
        return jsonify({"data": approvals_list, "totalCount": total_count}), 200

    except Exception as ex:
        logger.info(f"Failed to fetch all approval data: {ex}", exc_info=True)
        return jsonify({"error": "Internal server error", "details": str(ex)}), 500

# Helper function to convert Snowflake rows to list of dicts
def convert_rows_to_list_of_approval_models(cursor, rows):
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in rows]

# Search Data Endpoint
@app.route('/api/Home/EditData', methods=['GET'])
def edit_data_endpoint():
    """
    Fetches data for a given tagNo by searching across all supported tag types.
    Returns the first match found along with its tagType.
    """
    tag_no = request.args.get('tagNo')
    logger.info(f"EditData API called with tagNo: {tag_no}")

    if not tag_no:
        logger.info("Missing required parameter: tagNo")
        return jsonify({"error": "tagNo parameter is required."}), 400

    try:
        conn = get_snowflake_connection()
        tag_types = ['Equipment', 'Instrument', 'Line', 'Cable']

        for tag_type in tag_types:
            print("Tag Type---------", tag_types)
            table_name = TABLE_MAPPING[tag_type]['main']
            # query = f"SELECT * FROM {table_name} WHERE TAG_NO = %s LIMIT 1"
            query = f"SELECT * FROM {table_name} WHERE TAG_NO ILIKE %s LIMIT 1"

            with conn.cursor() as cursor:
                cursor.execute(query, (tag_no,))
                rows = cursor.fetchall()

                if rows:
                    data_list = convert_rows_to_list_of_approval_models(cursor, rows)
                    for item in data_list:
                        item['tagType'] = tag_type

                    logger.info(f"Data found for tagNo: {tag_no} in tagType: {tag_type}, table: {table_name}")
                    return jsonify(data_list), 200

        logger.info(f"No data found for tagNo: {tag_no} in any tagType")
        return jsonify({"data": [], "tagType": None}), 200

    except Exception as ex:
        logger.info(f"Error fetching data for tagNo: {tag_no}: {ex}", exc_info=True)
        return jsonify({"error": "Internal server error", "details": str(ex)}), 500

# Delete Data Endpoint
@app.route('/api/Home/DeleteData', methods=['DELETE'])
def delete_data_endpoint():
    """
    Deletes a tag entry from the specified tagType table using the provided tagNo.
    """
    tag_type = request.args.get('tagType')
    tag_no = request.args.get('tagNo')
    logger.info(f"DeleteData API called with tagType: {tag_type}, tagNo: {tag_no}")

    # Validate required parameters
    if not tag_type or not tag_no:
        logger.info("Missing required parameters: tagType or tagNo")
        return jsonify({"success": False, "error": "Both tagType and tagNo parameters are required."}), 400

    # Validate tagType against allowed table mappings
    if tag_type not in TABLE_MAPPING:
        logger.info(f"Invalid tagType received: {tag_type}")
        return jsonify({"success": False, "error": "Invalid tagType."}), 400

    try:
        conn = get_snowflake_connection()
        table_name = TABLE_MAPPING[tag_type]['main']
        delete_query = f"DELETE FROM {table_name} WHERE TAG_NO = %s"

        with conn.cursor() as cursor:
            cursor.execute(delete_query, (tag_no,))
            rows_affected = cursor.rowcount
        conn.commit();

        logger.info(f"{rows_affected} row(s) deleted from {table_name} for tagNo: {tag_no}")
        return jsonify({"success": rows_affected > 0}), 200

    except Exception as ex:
        logger.info(f"Error while deleting tagNo: {tag_no} from {tag_type} - {ex}", exc_info=True)
        return jsonify({"success": False, "error": "Internal server error", "details": str(ex)}), 500

# Edit Data Endpoint
@app.route('/api/Home/SaveData', methods=['POST'])
def save_data_endpoint():
    """
    Updates the edited data in the main tables based on tagType.
    """
    logger.info("SaveData endpoint called.")

    model_data = request.get_json()
    logger.debug(f"Received request data: {model_data}")

    if not model_data:
        logger.info("Request body is empty or invalid.")
        return jsonify({"success": False, "message": "Invalid request body."}), 400

    tag_no = model_data.get('tagNo')
    tag_type = model_data.get('tagType')

    logger.debug(f"Extracted tagNo: '{tag_no}', tagType: '{tag_type}'")

    if not tag_no:
        logger.info("Missing 'tagNo' in the request.")
        return jsonify({"success": False, "message": "Request body must contain a tagNo."}), 400

    if not tag_type or tag_type not in TABLE_MAPPING:
        logger.info(f"Invalid or missing 'tagType': '{tag_type}'")
        return jsonify({"success": False, "message": "Invalid 'tagType'."}), 400

    target_table = TABLE_MAPPING[tag_type]['main']
    column_mapping = TABLE_MAPPING[tag_type]['mapping']
    logger.info(f"Target table: '{target_table}'")

    try:
        with get_snowflake_connection() as conn, conn.cursor() as cursor:
            # Prepare parameters dictionary using the column mapping
            params_dict = {}
            for json_key, snowflake_col in column_mapping.items():
                if json_key in model_data and model_data[json_key] is not None and snowflake_col != 'DATE_UPDATED':
                    params_dict[snowflake_col] = model_data[json_key]

            # params_dict = {
            #     snowflake_col: model_data.get(json_key)
            #     for json_key, snowflake_col in column_mapping.items()
            #     if snowflake_col !='DATE_UPDATED'
            # }

            # Always update both timestamps on modification
            # now = datetime.now()
            # params_dict['DATE_UPDATED'] = now
            now = get_est_now()
            params_dict['MODIFIED_DATE'] = now
            params_dict['MODIFIED_BY'] = model_data.get('userEmail')

            logger.debug(f"Prepared parameters for update: {params_dict}")

            # Construct the SET clause for the UPDATE query
            set_clause = ', '.join([f"{col} = %s" for col in params_dict.keys()])
            values = tuple(params_dict.values()) + (tag_no,)

            update_query = f"""
                UPDATE {target_table}
                SET {set_clause}
                WHERE TAG_NO = %s
            """

            logger.debug(f"Executing SQL: {update_query}")
            logger.debug(f"With values: {values}")
            cursor.execute(update_query, values)
            rows_affected = cursor.rowcount

            conn.commit()
            logger.info(f"Successfully updated {rows_affected} row(s) for tag '{tag_no}' in '{target_table}'.")

            if rows_affected == 0:
                logger.info(f"No rows updated for tag '{tag_no}' in '{target_table}'.")
                return jsonify({"success": False, "message": f"No data found for tagNo '{tag_no}'."}), 404

            return jsonify({"success": True, "message": f"Tag '{tag_no}' updated in {target_table}."}), 200

    except Exception as ex:
        logger.info(f"Failed to save data: {ex}", exc_info=True)
        return jsonify({"success": False, "message": "Internal server error", "details": str(ex)}), 500
 
# Edit Approval Data Endpoint
@app.route('/api/Home/ApprovalEditData', methods=['POST'])
def edit_approval_endpoint():
    """
    Updates only the HISTORY tables.
    If tagType changes, safely moves record from old history table to new one
    (insert → delete, transaction protected).
    """
    logger.info("Approval Edit endpoint called.")
    model_data = request.get_json()

    if not model_data:
        return jsonify({"success": False, "message": "Invalid request body."}), 400

    original_tag_no = model_data.get('tagNo')
    new_tag_type = model_data.get('tagType')
    original_tag_type = model_data.get('originalTagType')

    if not original_tag_no:
        return jsonify({"success": False, "message": "Missing tagNo"}), 400

    if not new_tag_type or new_tag_type not in TABLE_MAPPING:
        return jsonify({"success": False, "message": f"Invalid tagType '{new_tag_type}'"}), 400

    try:
        with get_snowflake_connection() as conn, conn.cursor() as cursor:

            # ================================================================
            # CASE 1: Tag type changed → move record between history tables
            # ================================================================
            if original_tag_type and original_tag_type != new_tag_type:
                old_table = TABLE_MAPPING[original_tag_type]['history']
                new_table = TABLE_MAPPING[new_tag_type]['history']
                old_map = TABLE_MAPPING[original_tag_type]['mapping']
                new_map = TABLE_MAPPING[new_tag_type]['mapping']

                logger.info(f"TagType changed: {original_tag_type} → {new_tag_type}")

                # Fetch old record
                cursor.execute(f"SELECT * FROM {old_table} WHERE TAG_NO = %s", (original_tag_no,))
                record = cursor.fetchone()
                if not record:
                    return jsonify({"success": False, "message": f"Record not found in {old_table}"}), 404

                old_columns = [col[0] for col in cursor.description]
                existing_data = dict(zip(old_columns, record))

                # Common keys between old/new tag type
                common_keys = set(old_map.keys()) & set(new_map.keys())
                logger.debug(f"Common columns: {common_keys}")

                # Build insert
                insert_cols, insert_vals = [], []

                for json_key in common_keys:
                    new_col = new_map[json_key]
                    old_col = old_map[json_key]
                    val = model_data.get(json_key, existing_data.get(old_col))
                    if val == "":
                        val = None
                    insert_cols.append(new_col)
                    insert_vals.append(val)

                # Add required columns if missing
                REQUIRED = {
                    'Equipment': ['ACTION','STATUS', 'TAG_NO', 'TYPE', 'DESCRIPTION', 'ASSET', 'FUNCTION', 'SUFFIX_1', 'SUFFIX_2', 'SEQUENCE','APPROVAL_STATUS','MODIFIED_BY','MODIFIED_DATE'],
                    'Instrument': ['ACTION','STATUS', 'TAG_NO', 'TYPE', 'DESCRIPTION', 'ASSET', 'FUNCTION', 'SUFFIX_1', 'SUFFIX_2', 'SEQUENCE','APPROVAL_STATUS','MODIFIED_BY','MODIFIED_DATE'],
                    'Line': ['ACTION','TAG_NO','DIAMETER', 'CLASS_SPECIFICATION', 'FLUID_CODE_SERVICE','APPROVAL_STATUS','STATUS', 'MODIFIED_BY','MODIFIED_DATE'],
                    'Cable': ['ACTION','STATUS', 'TAG_NO', 'TAG_TYPE', 'DESCRIPTION', 'ASSET', 'SUFFIX_1', 'SUFFIX_2', 'SEQUENCE','APPROVAL_STATUS','MODIFIED_BY','MODIFIED_DATE'],
                }
                for col in REQUIRED.get(new_tag_type, []):
                    if col not in insert_cols:
                        val =model_data.get(col,existing_data.get(col))
                        if val == "":
                            val = None
                        insert_cols.append(col)
                        insert_vals.append(val)
                        

                # Force DATE_UPDATED
                now = get_est_now()
                if 'DATE_UPDATED' in insert_cols:
                    insert_vals[insert_cols.index('DATE_UPDATED')] = now
                else:
                    insert_cols.append('DATE_UPDATED')
                    insert_vals.append(now)

                placeholders = ', '.join(['%s'] * len(insert_vals))
                columns_str = ', '.join(insert_cols)
                insert_query = f"INSERT INTO {new_table} ({columns_str}) VALUES ({placeholders})"

                # Transaction safety
                try:
                    cursor.execute(insert_query, tuple(insert_vals))
                    cursor.execute(f"DELETE FROM {old_table} WHERE TAG_NO = %s", (original_tag_no,))
                    conn.commit()
                    logger.info(f"Moved record from {old_table} → {new_table}")
                except Exception as e:
                    conn.rollback()
                    logger.error(f"Error during record move: {e}", exc_info=True)
                    return jsonify({"success": False, "message": "Failed to move record", "details": str(e)}), 500

                return jsonify({
                    "success": True,
                    "mode": "move",
                    "message": f"Record moved from {original_tag_type} → {new_tag_type} history table.",
                    "movedFrom": old_table,
                    "movedTo": new_table,
                    "commonColumns": list(common_keys)
                }), 200

            # CASE 2: Normal update (same tagType → update in same history)
            tag_type = original_tag_type or new_tag_type
            history_table = TABLE_MAPPING[tag_type]['history']
            mapping = TABLE_MAPPING[tag_type]['mapping']

            params = {}
            for json_key, col in mapping.items():
                if json_key in model_data:
                    val = model_data[json_key]
                    if val == "":
                        val = None
                    params[col] = val

            now = get_est_now()
            params['DATE_UPDATED'] = now

            set_clause = ', '.join([f"{col} = %s" for col in params])
            values = tuple(params.values()) + (original_tag_no,)

            update_query = f"""
                UPDATE {history_table}
                SET {set_clause}
                WHERE TAG_NO = %s
            """
            cursor.execute(update_query, values)
            affected = cursor.rowcount
            conn.commit()

            if affected == 0:
                logger.warning(f"No record updated in {history_table} for tag {original_tag_no}")

            return jsonify({
                "success": True,
                "mode": "update",
                "message": "Record updated successfully in history table.",
                "updatedRows": affected
            }), 200

    except Exception as ex:
        logger.error(f"Unhandled error in ApprovalEditData: {ex}", exc_info=True)
        return jsonify({"success": False, "message": "Internal server error", "details": str(ex)}), 500

# Tag classification logic
TAG_PATTERNS = {
    "Line": re.compile(r'^[A-Z0-9-]+_[0-9]+[A-Z]*-\d+-\d+$'),
    "Instrument": re.compile(r'^[A-Z0-9-]+_([A-Z]{2,}|[A-Z]+\d+)-\d+(-\d+)?$'),
    "Equipment": re.compile(r'^[A-Z0-9-]+_[A-Z]{1,3}-\d+(-[A-Z0-9]+)?$'),
    "Cable": re.compile(r'^[A-Z0-9-]+_(CBL|CABLE|CAB|CB)-?\d+')
}

# Classify tag based on TAG_NO and FileType
def classify_tag(tag_no: str, file_type: str) -> str:
    tag_no = tag_no.strip().upper()
    file_type = file_type.strip().upper()
    print("tagNo and FileType-----------", file_type, tag_no)

    # --- Rule 1: File-type special handling ---
    if file_type.startswith("A7001"):
    # Cause & Effect → includes Equipment + Instrument only
        func_code_match = re.search(r'_([A-Z/]+)-', tag_no)
        if func_code_match:
            func_code = func_code_match.group(1)
            # Multi-letter function codes 
            if len(func_code) >= 2:
                return "Instrument"
            # Single-letter function codes
            else:
                return "Equipment"
        return "Instrument"

    elif file_type.startswith(("A6")):
        return "Cable"
    
    elif file_type.startswith(("A7")):
        return "Instrument"
    
    # --- Rule 2: Pattern-based classification ---
    if TAG_PATTERNS["Cable"].search(tag_no):
        return "Cable"
    elif TAG_PATTERNS["Instrument"].search(tag_no):
        return "Instrument"
    elif TAG_PATTERNS["Equipment"].search(tag_no):
        return "Equipment"
    elif TAG_PATTERNS["Line"].search(tag_no):
        return "Line"

    # --- Rule 3: File-type fallback ---
    if file_type.startswith(("A4", "B5")):
        func_code_match = re.search(r'_([A-Z/]+)-', tag_no)
        # func_code_match = re.search(r'_([A-Z]+)-', tag_no)
        if func_code_match:
            func_code = func_code_match.group(1)
            if len(func_code) >= 2:
                return "Instrument"
            else:
                return "Equipment"
        return "Line"
    elif file_type.startswith("A42"):
        return "Equipment"
    elif file_type.startswith("A3"):
        return "Line"

    # --- Default fallback ---
    return "Unknown"

# Parse TAG_NO into components based on tag type
def parse_tag_details(tag_no: str, tag_type: str):
    """
    Parse TAG_NO into attributes depending on tag_type:
      - Equipment / Instrument: Asset_Function-Seq-Suffix1-Suffix2
      - Line: Asset_Diameter-Class-FluidCode-SeqNumber[-Suffix]
      - Cable: Asset_Prefix-Seq-Suffix1-Suffix2
    """
    t = (tag_no or "").strip()
    details = {}

    if not t:
        return details

    try:
        tt = (tag_type or "").lower()

        # Equipment / Instrument
        if tt in ("equipment", "instrument"):
            if "_" in t:
                asset, rest = t.split("_", 1)
                details["ASSET"] = asset
                parts = rest.split("-")
                if len(parts) >= 1:
                    details["FUNCTION"] = parts[0]
                if len(parts) >= 2:
                    details["SEQUENCE"] = parts[1]
                if len(parts) >= 3:
                    details["SUFFIX_1"] = parts[2]
                if len(parts) >= 4:
                    details["SUFFIX_2"] = parts[3]
            else:
                parts = t.split("-")
                if len(parts) >= 1:
                    details["FUNCTION"] = parts[0]

        # Line: Asset_Diameter-Class-FluidCode-SeqNumber[-Suffix]
        elif tt == "line":
            if "_" in t:
                asset, rest = t.split("_", 1)
                details["ASSET"] = asset

                # Normalize multiple spaces
                rest = " ".join(rest.split())
                parts = rest.split("-")

                if len(parts) > 0:
                    # Split first token into DIAMETER and CLASS_SPECIFICATION
                    first = parts[0].split(" ", 1)
                    details["DIAMETER"] = first[0] + (" " if len(first) > 1 else "")
                    if len(first) > 1:
                        details["CLASS_SPECIFICATION"] = first[1]

                if len(parts) >= 2:
                    details["FLUID_CODE_SERVICE"] = parts[1]
                if len(parts) >= 3:
                    details["SEQUENCE"] = parts[2]
                if len(parts) >= 4:
                    details["SUFFIX_1"] = parts[3]

        # Cable
        elif tt == "cable":
            if "_" in t:
                asset, rest = t.split("_", 1)
                details["ASSET"] = asset
                parts = rest.split("-")
                if len(parts) >= 1:
                    details["PREFIX"] = parts[0]
                if len(parts) >= 2:
                    details["SEQUENCE"] = parts[1]
                if len(parts) >= 3:
                    details["SUFFIX_1"] = parts[2]
                if len(parts) >= 4:
                    details["SUFFIX_2"] = parts[3]
            else:
                parts = t.split("-")
                if len(parts) >= 2:
                    details["PREFIX"] = parts[0]
                    details["SEQUENCE"] = parts[1]

        return {k: v for k, v in details.items() if v}

    except Exception as e:
        print(f"parse_tag_details error for {tag_no} {tag_type}: {e}")
        return {k: v for k, v in details.items() if v}

# Delete tag by S_NO
@app.route('/api/Home/delete-tag', methods=['POST'])
def delete_tag_endpoint():
    
    """
    Deletes a tag from ENI_NL_DEV_TAGREGISTRY.DEV.TBL_TAG_INFO table by tag_ID.
    """
    try:
        data = request.get_json()
        logger.info(f"Incoming delete request: {data}")

        s_no =data.get('S_NO') or data.get('sNo')
        if not s_no:
            return jsonify({"success": False, "message": "S_NO is required"}), 400

        conn = get_snowflake_connection()
        
        cursor = conn.cursor()

        cursor.execute('DELETE FROM ENI_NL_DEV_TAGREGISTRY.DEV.TBL_TAG_INFO WHERE "S_NO" = %s', (s_no,))
        conn.commit()

        return jsonify({"success": True, "message": f"Tag '{s_no}' deleted successfully"}), 200

    except Exception as ex:
        logger.info(f"Error deleting tag: {ex}", exc_info=True)
        return jsonify({"success": False, "error": str(ex)}), 500
 
# Get existing PIDs for a given TAG_NO
@app.route("/api/Home/get-existing-pids", methods=["GET"])
def get_existing_pids():
    tag_no = request.args.get("tagNo")
    file_type = request.args.get("fileType")  # Get fileType from request
    if not tag_no:
        return jsonify({"error": "Missing tagNo"}), 400

    try:
        conn = get_snowflake_connection()
        with conn.cursor() as cursor:
            classification = classify_tag(tag_no, file_type)

            # Choose correct table and column based on classification
            if classification == "Equipment":
                table = 'ENI_NL_DEV_TAGREGISTRY.DEV.TBL_EQUIPMENT_TAGS'
                pid_col = '"PIMS_DOCUMENT"'
            elif classification == "Instrument":
                table = 'ENI_NL_DEV_TAGREGISTRY.DEV.TBL_INSTRUMENT_TAGS'
                pid_col = '"PIMS_DOCUMENT"'
            elif classification == "Line":
                table = 'ENI_NL_DEV_TAGREGISTRY.DEV.TBL_LINE_TAGS'
                pid_col = '"PIMS_DOCUMENT"'  
            elif classification == "Cable":
                table = 'ENI_NL_DEV_TAGREGISTRY.DEV.TBL_CABLE_TAGS'
                pid_col = '"PIMS_DOCUMENT"'
            else:
                return jsonify({"error": f"Unknown classification: {classification}"}), 400

            query = f'SELECT {pid_col} FROM {table} WHERE "TAG_NO" = %s'
            cursor.execute(query, (tag_no,))
            row = cursor.fetchone()
            existing_pids = row[0] if row else ""

            return jsonify({"pids": existing_pids}), 200

    except Exception as e:
        logger.exception(f"Error fetching existing PIDs for {tag_no}")
        return jsonify({"error": str(e)}), 500

# Approve tag equipment endpoint
@app.route("/api/Home/approve-tag-equipment", methods=["POST"])
def approve_tag_equipment():
    data = request.get_json()
    logger.info(f"Approve request received: {data}")
    data = {k: v for k, v in (data or {}).items() if v not in ("", None)}

    tag_no = data.get("tagNo")
    tag_type = data.get("tagType")
    user_email = data.get("userEmail")
    file_type = data.get("file_Type")
    action = data.get("action")
    role_type = data.get("userRole", "")
    pid = data.get("PID")
    new_url = data.get("URL")
    s_no = data.get("S_NO") or data.get("sNo")

    # Input validation
    if not s_no:
        return jsonify({"error": "Missing S_NO"}), 400
    if not tag_no:
        return jsonify({"error": "Missing tagNo"}), 400
    if not tag_type:
        return jsonify({"error": "Missing tagType; please provide tagType"}), 400
    if tag_type not in TABLE_MAPPING:
        return jsonify({"error": f"Unsupported tag type: {tag_type}"}), 400

    table_info = TABLE_MAPPING[tag_type]
    target_table = table_info["main"] if role_type.lower() == "admin" else table_info["history"]
    column_mapping = table_info["mapping"]

    conn = None
    result_status = None  # <-- used for frontend messaging

    try:
        conn = get_snowflake_connection()
        with conn.cursor() as cursor:
            pid_col = '"PIMS_DOCUMENT"'

            # Step 1: Identify which table currently holds this tag
            existing_table = None
            actual_tag_type = tag_type
            row = None

            for tbl_key, tbl_info in TABLE_MAPPING.items():
                check_table = tbl_info["main"] if role_type.lower() == "admin" else tbl_info["history"]
                cursor.execute(f'SELECT "URL", {pid_col} FROM {check_table} WHERE "TAG_NO" = %s', (tag_no,))
                existing_row = cursor.fetchone()
                if existing_row:
                    existing_table = check_table
                    actual_tag_type = tbl_key
                    row = existing_row
                    logger.info(f"Tag {tag_no} found in {check_table} (tag type: {tbl_key})")
                    break

            if not existing_table:
                existing_table = target_table
                actual_tag_type = tag_type

            now = get_est_now()

            # Step 2: Update if tag exists
            if row:
                current_urls = row[0] or ""
                current_pid = row[1] or ""

                urls = set(u.strip() for u in current_urls.split(";") if u.strip())
                pids = set(p.strip() for p in current_pid.split(";") if p.strip())

                new_pids = [p.strip() for p in (pid or "").split(";") if p.strip()]
                new_urls = [u.strip() for u in (new_url or "").split(";") if u.strip()]

                urls.update(u for u in new_urls if u and u not in urls)
                pids.update(p for p in new_pids if p and p not in pids)

                updated_urls = ";".join(sorted(urls)) if urls else ""
                updated_pid = ";".join(sorted(pids)) if pids else ""

                update_query = f'''
                    UPDATE {existing_table}
                    SET "URL" = %s,
                        {pid_col} = %s,
                        "DATE_UPDATED" = %s,
                        "MODIFIED_DATE" = %s,
                        "MODIFIED_BY" = %s
                    WHERE "TAG_NO" = %s
                '''
                cursor.execute(update_query, (updated_urls, updated_pid, now, now, user_email, tag_no))

                result_status = "update"  # <-- existing tag updated
                logger.info(f"Updated {actual_tag_type} tag '{tag_no}' with new documents")

            # Step 3: Insert if tag does NOT exist
            else:
                columns, values, added_cols = [], [], set()

                for json_key, db_col in column_mapping.items():
                    if json_key in data:
                        col_name = f'"{db_col}"'
                        if col_name not in added_cols:
                            columns.append(col_name)
                            values.append(data[json_key])
                            added_cols.add(col_name)

                if new_url and '"URL"' not in added_cols:
                    columns.append('"URL"')
                    values.append(new_url)
                    added_cols.add('"URL"')

                if pid and pid_col not in added_cols:
                    columns.append(pid_col)
                    values.append(pid)
                    added_cols.add(pid_col)

                parsed = parse_tag_details(tag_no, tag_type)
                logger.info(f"Parsed tag details for {tag_no}: {parsed}")
                for k, v in parsed.items():
                    col_name = f'"{k}"'
                    if col_name not in added_cols:
                        columns.append(col_name)
                        values.append(v)
                        added_cols.add(col_name)

                for audit_col in ['"DATE_UPDATED"', '"MODIFIED_DATE"']:
                    if audit_col not in added_cols:
                        columns.append(audit_col)
                        values.append(now)
                        added_cols.add(audit_col)

                if '"MODIFIED_BY"' not in added_cols:
                    columns.append('"MODIFIED_BY"')
                    values.append(user_email)
                    added_cols.add('"MODIFIED_BY"')

                if role_type.lower() == "contributor" and action and '"ACTION"' not in added_cols:
                    columns.append('"ACTION"')
                    values.append(action)
                    added_cols.add('"ACTION"')

                if role_type.lower() == "contributor" and '"APPROVAL_STATUS"' not in added_cols:
                    columns.append('"APPROVAL_STATUS"')
                    values.append("PENDING")
                    added_cols.add('"APPROVAL_STATUS"')

                if not any(col.strip('"').upper() in ('TAG_TYPE', 'TYPE') for col in columns):
                    columns.append('"TAG_TYPE"')
                    values.append(tag_type)

                placeholders = ", ".join(["%s"] * len(values))
                insert_query = f'INSERT INTO {existing_table} ({", ".join(columns)}) VALUES ({placeholders})'
                cursor.execute(insert_query, tuple(values))

                result_status = "insert"  # <-- new tag inserted
                logger.info(f"Inserted new {tag_type} tag '{tag_no}' into {existing_table}")

            # Step 4: Delete from staging
            delete_query = 'DELETE FROM ENI_NL_DEV_TAGREGISTRY.DEV.TBL_TAG_INFO WHERE "S_NO" = %s AND "TAG_NO" = %s'
            try:
                s_no_int = int(s_no)
                cursor.execute(delete_query, (s_no_int, tag_no))
                logger.info(f"Deleted staging record for S_NO={s_no_int}, TAG_NO={tag_no}")
            except ValueError:
                logger.warning(f"Invalid S_NO (non-integer): {s_no}")

        conn.commit()

        # Final output with frontend flag
        return jsonify({
            "success": True,
            "tagNo": tag_no,
            "actualTagType": actual_tag_type,
            "documentType": tag_type,
            "isNewInsert": True if result_status == "insert" else False
        }), 200

    except Exception as e:
        if conn:
            conn.rollback()
        logger.exception(f"Error approving tag {tag_no}")
        return jsonify({"error": str(e)}), 500

    finally:
        if conn:
            conn.close()

# Helper functions for tag processing
MANDATORY_FIELDS_BY_TAG = {
    'Equipment': ['STATUS', 'TYPE', 'TAG_NO', 'ASSET', 'FUNCTION', 'SEQUENCE', 'DESCRIPTION'],
    'Instrument': ['STATUS', 'TYPE', 'TAG_NO', 'ASSET', 'FUNCTION', 'INSTRUMENT_TYPE_FUNCTION_DESCRIPTION', 'SEQUENCE', 'DESCRIPTION'],
    'Line': ['STATUS', 'TAG_NO', 'ASSET', 'DIAMETER', 'CLASS_SPECIFICATION', 'FLUID_CODE_SERVICE', 'SEQUENCE'],
    'Cable': ['STATUS', 'TAG_NO', 'ASSET', 'PREFIX', 'SEQUENCE', 'FROM_LOCATION', 'TO_LOCATION']
}

# Columns that need to be merged with semicolon separation
MERGE_COLUMNS_BY_TAG = {
    'Equipment': ['pimsDocument', 'document', 'relatedTags'],
    'Instrument': ['pimsDocument', 'document', 'relatedTags'],
    'Line': ['pimsDocument', 'document', 'relatedTags'],
    'Cable': ['pimsDocument', 'document', 'relatedTags']
}

# Normalize text for comparison
def normalize(text):
    return re.sub(r'\W+', '', str(text or '')).lower()

# Merge two semicolon-separated values into a unique semicolon-separated string
def merge_semicolon_values(existing, new):
    if not existing and not new:
        return ""
    if not existing:
        return str(new).strip() if new else ""
    if not new:
        return str(existing).strip() if existing else ""
    existing_items = [item.strip() for item in str(existing).split(';') if item.strip()]
    new_items = [item.strip() for item in str(new).split(';') if item.strip()]
    merged = existing_items.copy()
    for item in new_items:
        if item not in merged:
            merged.append(item)
    result = ';'.join(merged)
    logger.debug(f"Merge: existing=[{existing}] + new=[{new}] = [{result}]")
    return result

# Build TAG_NO from components based on tag type
def build_tag_no_by_formula(tag_type_key, candidate_values):
    tt = (tag_type_key or '').lower()

    def clean_value(v):
        if v is None or str(v).strip().lower() in ('nan', 'none', ''):
            return None
        try:
            if isinstance(v, (int, float)):
                if float(v).is_integer():
                    return str(int(v))
                else:
                    return str(int(v)) if str(v).endswith('.0') else str(v).strip()
        except Exception:
            pass
        return str(v).strip()

    def part(*keys):
        for k in keys:
            v = candidate_values.get(k) or candidate_values.get(k.lower()) or candidate_values.get(k.upper())
            cv = clean_value(v)
            if cv:
                return cv
        return None

    seq = part('SEQUENCE', 'SEQ', 'Sequence No', 'SEQ_NO')
    if seq:
        candidate_values['SEQUENCE'] = seq

    if 'equipment' in tt or 'instrument' in tt:
        asset = part('ASSET', 'Asset')
        function = part('FUNCTION', 'Function', 'Function Description', 'FUNC')
        seq = part('SEQUENCE', 'SEQ', 'SEQ_NO')
        s1 = part('SUFFIX_1', 'SUFFIX-1', 'Suffix 1')
        s2 = part('SUFFIX_2', 'SUFFIX-2', 'Suffix 2')

        if not (asset and function and seq):
            return None

        parts = [f"{asset}_{function}-{seq}"]
        if s1: parts.append(s1)
        if s2: parts.append(s2)
        return "-".join(filter(lambda x: x not in (None, '', 'nan'), parts))

    if 'line' in tt:
        asset = part('ASSET')
        diameter = part('DIAMETER', 'DN')
        class_spec = part('CLASS_SPECIFICATION', 'CLASS', 'CLASS_SPEC')
        fluid = part('FLUID_CODE_SERVICE', 'FLUID', 'SERVICE')
        seq = part('SEQUENCE', 'SEQ', 'SEQ_NO')
        if not (asset and diameter and class_spec and fluid and seq):
            return None

        tag_parts = [f"{asset}_{diameter}", class_spec, fluid, seq]
        return "-".join(filter(lambda x: x not in (None, '', 'nan'), tag_parts))

    if 'cable' in tt:
        asset = part('ASSET')
        prefix = part('PREFIX')
        seq = part('SEQUENCE', 'SEQ', 'SEQ_NO')
        s1 = part('SUFFIX_1', 'SUFFIX-1', 'Suffix 1')
        s2 = part('SUFFIX_2', 'SUFFIX-2', 'Suffix 2')
        if not (asset and prefix and seq):
            return None

        tag_parts = [f"{asset}_{prefix}", seq, s1, s2]
        return "-".join(filter(lambda x: x not in (None, '', 'nan'), tag_parts))

    return None

# Map Excel columns to database columns based on normalization and heuristics
def map_excel_to_db_columns(excel_columns, db_mapping):
    mapped, used = {}, set()
    norm_db = {normalize(k): v for k, v in db_mapping.items()}
    tag_number_variations = ['tagno', 'lineno', 'cableno']
    for excel_col in excel_columns:
        norm_excel = normalize(excel_col)
        if norm_excel in tag_number_variations and 'TAG_NO' in db_mapping.values() and 'TAG_NO' not in used:
            mapped[excel_col] = 'TAG_NO'
            used.add('TAG_NO')
            continue
        if norm_excel in norm_db and norm_db[norm_excel] not in used:
            mapped[excel_col] = norm_db[norm_excel]
            used.add(norm_db[norm_excel])
            continue
        for json_key, db_col in db_mapping.items():
            if db_col not in used and re.search(normalize(json_key), norm_excel, re.IGNORECASE):
                mapped[excel_col] = db_col
                used.add(db_col)
                break
    logger.debug(f"Excel to DB column mapping: {mapped}")
    return mapped

# Upload Excel endpoint
@app.route('/api/Home/uploadExcel', methods=['POST'])
def upload_excel_endpoint():
    logger.info("Upload Excel API called.")

    if 'file' not in request.files:
        logger.warning("No file part in the request.")
        return jsonify({"message": "No file part in the request."}), 400

    files = request.files.getlist('file')
    if not files:
        logger.warning("No selected file.")
        return jsonify({"message": "No selected file."}), 400

    overall_summary = []

    for file in files:
        file_summary = {
            "fileName": file.filename,
            "insertedCount": 0,
            "updatedCount": 0,
            "notInsertedCount": 0,
            "totalRowsProcessed": 0,
            "errors": []
        }

        logger.info(f"Processing file: {file.filename}")

        if not file.filename:
            file_summary["errors"].append("Unnamed file skipped.")
            overall_summary.append(file_summary)
            continue

        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            file_summary["errors"].append(
                f"File '{file.filename}' has an invalid format. Only .xlsx or .xls allowed."
            )
            overall_summary.append(file_summary)
            continue

        try:
            content = file.stream.read()
            df = pd.read_excel(io.BytesIO(content))

            # Sanitize column names
            df.columns = (
                df.columns.str.strip()
                .str.replace('.', '', regex=False)
                .str.replace(r'\s+', '', regex=True)
                .str.lower()
            )

            if df.empty:
                file_summary["errors"].append(f"File '{file.filename}' is empty.")
                overall_summary.append(file_summary)
                continue

            normalized_cols = {normalize(col): col for col in df.columns}

            # Detect tagno column
            tagno_col = None
            for cand in ['tagno', 'lineno', 'cableno']:
                if normalized_cols.get(cand):
                    tagno_col = normalized_cols[cand]
                    break

            tagtype_col = normalized_cols.get('tagtype')

            if not tagno_col:
                file_summary["errors"].append(
                    f"File '{file.filename}' missing required column: Tag No./Line No./Cable No."
                )
                overall_summary.append(file_summary)
                continue

            # Rename to standardized keys
            rename_map = {tagno_col: 'tagno'}
            if tagtype_col:
                rename_map[tagtype_col] = 'tagtype'
            df.rename(columns=rename_map, inplace=True)

            records = df.to_dict(orient='records')
            file_summary["totalRowsProcessed"] = len(records)

            # ---------- Validation pass ----------
            validation_errors = []
            validated_actions = []

            conn = get_snowflake_connection()
            cursor = conn.cursor()

            for idx, row in enumerate(records, start=1):
                original_row = row.copy()

                raw_tagno_val = row.get('tagno')
                tag_no = ''
                if raw_tagno_val is not None and not (
                    isinstance(raw_tagno_val, float) and pd.isna(raw_tagno_val)
                ):
                    tag_no_str = str(raw_tagno_val).strip()
                    if tag_no_str.lower() not in ('nan', 'none', 'null', ''):
                        tag_no = tag_no_str

                tag_type_raw = row.get('tagtype')
                matched_tag_type = None
                mapping = None
                table_info = None

                # --- Determine if this tag already exists in any main table ---
                if tag_no:
                    for candidate_tt, tbl_info in TABLE_MAPPING.items():
                        cursor.execute(
                            f"SELECT COUNT(*) FROM {tbl_info['main']} WHERE TAG_NO = %s", (tag_no,)
                        )
                        count = cursor.fetchone()[0]
                        if count > 0:
                            matched_tag_type = candidate_tt
                            table_info = tbl_info
                            mapping = table_info['mapping']
                            logger.info(f"Detected existing TAG '{tag_no}' belongs to type '{matched_tag_type}'.")
                            break

                # --- If not found in DB, fallback to Tag Type column or inference ---
                if not matched_tag_type:
                    if tag_type_raw:
                        norm_tt = normalize(tag_type_raw)
                        matched_tag_type = next(
                            (k for k in TABLE_MAPPING if normalize(k) in norm_tt), None
                        )
                        if matched_tag_type:
                            table_info = TABLE_MAPPING[matched_tag_type]
                            mapping = table_info['mapping']
                            logger.info(f"Tag Type column indicates type '{matched_tag_type}' for row {idx}.")

                excel_to_db = {}
                if mapping:
                    excel_to_db = map_excel_to_db_columns(original_row.keys(), mapping)

                # Build excel_input
                excel_input = {}
                for excel_col, db_col in excel_to_db.items():
                    val = row.get(excel_col)
                    if val is None:
                        val = original_row.get(excel_col)
                    if val is not None and not (isinstance(val, float) and pd.isna(val)) and str(val).strip() != "":
                        excel_input[db_col] = val

                # Gather candidate values
                raw_candidate_values = {
                    k.strip().upper(): v
                    for k, v in original_row.items()
                    if not (v is None or (isinstance(v, float) and pd.isna(v)))
                }

                # Infer tag type if not found
                if not matched_tag_type:
                    candidate_values = {}
                    candidate_values.update(excel_input)
                    candidate_values.update(raw_candidate_values)
                    for candidate_tt in TABLE_MAPPING.keys():
                        built = build_tag_no_by_formula(candidate_tt, candidate_values)
                        if built:
                            tag_no = built
                            matched_tag_type = candidate_tt
                            table_info = TABLE_MAPPING[matched_tag_type]
                            mapping = table_info['mapping']
                            excel_input['TAG_NO'] = built
                            logger.info(
                                f"Validation: inferred TAG_NO '{built}' and tag type '{matched_tag_type}' for row {idx}."
                            )
                            break

                if not mapping:
                    validation_errors.append(
                        f"Row {idx}: Unable to determine tag type or table mapping for TAG '{tag_no or ''}'."
                    )
                    continue

                # Determine Add or Edit
                is_edit = bool(tag_no)
                existing = None

                if is_edit:
                    # --- Verify if tag exists in main table ---
                    cursor.execute(
                        f"SELECT COUNT(*) FROM {table_info['main']} WHERE TAG_NO = %s", (tag_no,)
                    )
                    exists_count = cursor.fetchone()[0]

                    if exists_count == 0:
                        msg = f"Row {idx}: Tag '{tag_no}' is not present in register. Please Add instead of Edit."
                        logger.error(msg)
                        validation_errors.append(msg)
                        continue
                    logger.info(f"Validation: TAG_NO '{tag_no}' marked as Edit (verified exists in main table).")

                    # ✅ ADD THIS SECTION - Fetch existing values for merge columns
                    merge_columns = MERGE_COLUMNS_BY_TAG.get(matched_tag_type, [])
                    if merge_columns:
                        # Get the DB column names for merge columns
                        merge_db_cols = []
                        for merge_field in merge_columns:
                            db_col = mapping.get(merge_field)
                            if db_col:
                                merge_db_cols.append(db_col)
                        
                        if merge_db_cols:
                            # Fetch existing values from main table
                            fetch_cols = ', '.join(merge_db_cols)
                            cursor.execute(
                                f"SELECT {fetch_cols} FROM {table_info['main']} WHERE TAG_NO = %s", (tag_no,)
                            )
                            existing_row = cursor.fetchone()
                            
                            if existing_row:
                            # Merge semicolon values correctly
                                for i, db_col in enumerate(merge_db_cols):
                                    existing_value = existing_row[i]
                                    new_value = excel_input.get(db_col)

                                    # If Excel provided nothing new, keep existing
                                    if new_value is None or str(new_value).strip() == "":
                                        merged_value = existing_value
                                    else:
                                        merged_value = merge_semicolon_values(existing_value, new_value)

                                    excel_input[db_col] = merged_value
                                    logger.debug(f"Row {idx}: Merged {db_col}: existing=[{existing_value}] + new=[{new_value}] = [{merged_value}]")
     
                                    if existing_value or new_value:
                                        merged_value = merge_semicolon_values(existing_value, new_value)
                                        excel_input[db_col] = merged_value
                                        logger.debug(f"Row {idx}: Merged {db_col}: existing=[{existing_value}] + new=[{new_value}] = [{merged_value}]")

                else:
                    candidate_values = {}
                    candidate_values.update(excel_input)
                    candidate_values.update(raw_candidate_values)
                    generated = build_tag_no_by_formula(matched_tag_type, candidate_values)
                    if generated:
                        tag_no = generated
                        excel_input['TAG_NO'] = generated
                        logger.info(f"Validation: auto-generated TAG_NO '{generated}' for Add on row {idx}.")

                        cursor.execute(
                            f"SELECT COUNT(*) FROM {table_info['main']} WHERE TAG_NO = %s", (tag_no,)
                        )
                        exists_count = cursor.fetchone()[0]

                        if exists_count > 0:
                            msg = f"Row {idx}: Tag '{tag_no}' already exists in {matched_tag_type} register. Please use Edit instead of Add."
                            # msg = f"Row {idx}: Tag '{tag_no}' is already present in register. Please Edit instead of Add."
                            logger.error(msg)
                            validation_errors.append(msg)
                            continue
                    else:
                        validation_errors.append(
                            f"Row {idx}: Unable to auto-generate TAG_NO (missing Asset/Function/Sequence or required components)."
                        )
                        continue

                # Auto-populate fields
                parsed_fields = parse_tag_no_to_fields(tag_no, matched_tag_type, mapping)
                for db_col_name, field_value in parsed_fields.items():
                    if db_col_name not in excel_input or not excel_input.get(db_col_name):
                        excel_input[db_col_name] = field_value
                        
                    # ✅ NEW: For Edit operations, verify tag components haven't changed
                    if is_edit and parsed_fields:
                        for db_col_name, parsed_value in parsed_fields.items():
                            user_provided_value = excel_input.get(db_col_name)
                            
                            # If user provided a value and it's different from what's in the tag
                            if user_provided_value and str(user_provided_value).strip() != str(parsed_value).strip():
                                # Find the friendly field name for error message
                                friendly_name = next((k for k, v in mapping.items() if v == db_col_name), db_col_name)
                                validation_errors.append(
                                    f"Cannot modify tag component '{friendly_name}' in Edit mode. "
                                    f"Tag '{tag_no}' contains '{parsed_value}' but you provided '{user_provided_value}'. "
                                    f"Tag no cannot be changed. Please check and reupload."
                                )
                                break
                            
                if is_edit:
                    # ✅ For EDIT: Only TAG_NO is mandatory (tag already exists with other data)
                    if not tag_no or str(tag_no).strip() == "":
                        validation_errors.append(f"Row {idx}: TAG_NO is required for Edit operation")
                        continue
                    # Skip other mandatory field checks for Edit - user can update any field(s)
                    logger.info(f"Row {idx}: Edit operation for tag '{tag_no}' - will update provided fields only")
                else:
                    # ✅ For ADD: All mandatory fields required (creating new tag)
                    mandatory_list = MANDATORY_FIELDS_BY_TAG.get(matched_tag_type, [])
                    missing_mandatory = []
                    for mf in mandatory_list:
                        db_col_name = mapping.get(mf) if mapping else None
                        v = excel_input.get(db_col_name) if db_col_name else excel_input.get(mf)
                        if v is None or (isinstance(v, float) and pd.isna(v)) or str(v).strip() == "":
                            missing_mandatory.append(mf)

                    if missing_mandatory:
                        validation_errors.append(
                            f"Row {idx}: Tag '{tag_no}' is missing mandatory fields: {', '.join(missing_mandatory)}"
                        )
                        continue
                

                validated_actions.append({
                    "action": "Edit" if is_edit else "Add",
                    "tag_no": tag_no,
                    "matched_tag_type": matched_tag_type,
                    "table_info": table_info,
                    "mapping": mapping,
                    "params": excel_input,
                    "existing": existing
                })

            cursor.close()
            conn.close()

            # If any validation errors exist
            if validation_errors:
                logger.error(f"Validation failed for file '{file.filename}': {validation_errors}")
                file_summary["errors"].extend(validation_errors)
                file_summary["notInsertedCount"] = file_summary["totalRowsProcessed"]
                overall_summary.append(file_summary)
                continue

            # ---------- Commit pass ----------
            conn = get_snowflake_connection()
            cursor = conn.cursor()

            try:
                for action in validated_actions:
                    table_name = action["table_info"]["history"]
                    mapping = action["mapping"]
                    params = action["params"]
                    tag_no = action["tag_no"]
                    matched_tag_type = action["matched_tag_type"]
                    logger.info(
                        f"Committing TAG '{tag_no}' as {action['action']} into history table {table_name}."
                    )
                    tag_type_column = 'TAG_TYPE' if 'TAG_TYPE' in mapping.values() else 'TYPE'
                    params.update({
                        tag_type_column: matched_tag_type,
                        'DATE_UPDATED': get_est_now(),
                        'ACTION': action["action"],
                        'MODIFIED_BY': request.form.get('uploadedBy', 'system'),
                        'MODIFIED_DATE': get_est_now(),
                        'APPROVAL_STATUS': 'PENDING'
                    })

                    columns = ', '.join(params.keys())
                    placeholders = ', '.join(['%s'] * len(params))
                    query = f"INSERT INTO {table_name} ({columns}) VALUES ({placeholders})"
                    values = [params[k] for k in params.keys()]

                    cursor.execute(query, values)
                    file_summary["insertedCount"] += 1

                    logger.info(
                        f"Inserted TAG '{tag_no}' ({action['action']}) into history table {table_name}."
                    )

                conn.commit()

            except Exception as commit_ex:
                conn.rollback()
                logger.exception(f"DB commit error for file '{file.filename}': {commit_ex}")
                file_summary["errors"].append(f"DB commit error: {commit_ex}")
                file_summary["notInsertedCount"] = file_summary["totalRowsProcessed"]

            finally:
                cursor.close()
                conn.close()

            overall_summary.append(file_summary)

        except Exception as ex:
            logger.exception(f"Exception while processing file '{file.filename}': {ex}")
            file_summary["errors"].append(str(ex))
            file_summary["notInsertedCount"] = file_summary["totalRowsProcessed"]
            overall_summary.append(file_summary)
            continue

    # ---------- Build final response ----------
    response = {"files": overall_summary}
    any_file_rejected_entirely = False
    error_message_to_show = None

    response = {"files": overall_summary}
    any_file_rejected_entirely = False
    error_message_to_show = None

    for f in overall_summary:
        if f["notInsertedCount"] > 0 and f["notInsertedCount"] == f["totalRowsProcessed"]:
            any_file_rejected_entirely = True
            if any("already exists in" in e and "register" in e for e in f["errors"]):
            # Extract the specific error to show register name
                error_message_to_show = next((e for e in f["errors"] if "already exists in" in e), 
                                            "One or more tags already exist in register. Please use Edit instead of Add.")
            elif any("not present in register" in e for e in f["errors"]):
                error_message_to_show = "One or more tags do not exist in register. Please use Add instead of Edit."
            elif any("missing mandatory fields" in e for e in f["errors"]):
                error_message_to_show = "One or more tags are missing mandatory fields. Please check the errors and update the file."
            elif f["errors"]:
                # Show first error as summary
                error_message_to_show = f["errors"][0]
            else:
                error_message_to_show = "Upload failed. Please check the file and try again."
            break


    if any_file_rejected_entirely:
        response["message"] = error_message_to_show
        logger.error(response["message"])
        return jsonify(response), 400

    response["message"] = "Upload successful."
    logger.info(response["message"])
    return jsonify(response), 200

# Parse TAG_NO into component fields
def parse_tag_no_to_fields(tag_no_str, tag_type_key, db_mapping):
    """
    Extract component fields from TAG_NO and map to DB column names using db_mapping.
    Returns a mapping {db_column_name: value}
    """
    fields = {}
    tt = (tag_type_key or '').lower()

    def get_db_column(search_keys):
        for key in search_keys:
            if key in db_mapping:
                return db_mapping[key]
            for map_key, map_val in db_mapping.items():
                if normalize(map_key) == normalize(key):
                    return map_val
        return None

    try:
        if 'equipment' in tt or 'instrument' in tt:
            parts = tag_no_str.split('_')
            logger.debug(f"[Equipment/Instrument] Parsing TAG '{tag_no_str}': parts after underscore split = {parts}")
            if len(parts) >= 2:
                asset_col = get_db_column(['asset', 'ASSET', 'Asset'])
                if asset_col:
                    fields[asset_col] = parts[0]
                right_parts = parts[1].split('-')
                if len(right_parts) >= 2:
                    func_col = get_db_column(['function', 'FUNCTION', 'Function'])
                    seq_col = get_db_column(['sequence', 'SEQUENCE', 'SEQ'])
                    if func_col:
                        fields[func_col] = right_parts[0]
                    if seq_col:
                        fields[seq_col] = right_parts[1]
                    if len(right_parts) >= 3:
                        suffix1_col = get_db_column(['suffix1', 'SUFFIX_1', 'SUFFIX-1', 'SUFFIX1'])
                        if suffix1_col:
                            fields[suffix1_col] = right_parts[2]
                    if len(right_parts) >= 4:
                        suffix2_col = get_db_column(['suffix2', 'SUFFIX_2', 'SUFFIX-2', 'SUFFIX2'])
                        if suffix2_col:
                            fields[suffix2_col] = right_parts[3]

        elif 'line' in tt:
            logger.debug(f"[Line] Starting to parse TAG_NO: '{tag_no_str}'")
            
            # Split by underscore (limit to 1 split to handle asset codes with dashes like L7-F)
            parts = tag_no_str.split('_', 1)
            logger.debug(f"[Line] After underscore split: parts = {parts}")
            
            if len(parts) < 2:
                logger.warning(f"[Line] TAG '{tag_no_str}' does not contain underscore separator")
                return fields
                
            # Extract ASSET
            asset_col = get_db_column(['asset', 'ASSET', 'Asset'])
            if asset_col:
                fields[asset_col] = parts[0]
                logger.debug(f"[Line] Extracted ASSET: {parts[0]} -> DB column: {asset_col}")

            # Now parse the right side: "4" 250DS-1-004"
            right_side = parts[1]
            logger.debug(f"[Line] Right side to parse: '{right_side}'")
            
            # Split by dash
            right_parts = right_side.split('-')
            logger.debug(f"[Line] After dash split: right_parts = {right_parts}, length = {len(right_parts)}")
            
            # Get DB column names - use exact camelCase keys from line_column_mapping
            diam_col = get_db_column(['diameter', 'DIAMETER', 'DN', 'Diameter'])
            class_col = get_db_column(['classSpecification', 'CLASS_SPECIFICATION', 'CLASS', 'class_specification'])
            fluid_col = get_db_column(['fluidCode', 'FLUID_CODE_SERVICE', 'FLUID', 'fluid_code_service'])
            seq_col = get_db_column(['sequence', 'SEQUENCE', 'SEQ', 'Sequence'])
            
            logger.debug(f"[Line] DB Column mapping: DIAMETER={diam_col}, CLASS={class_col}, FLUID={fluid_col}, SEQ={seq_col}")

            if len(right_parts) == 3:
                # Format: "4" 250DS-1-004"
                # right_parts[0] = '4" 250DS' (diameter + class with space)
                # right_parts[1] = '1' (fluid)
                # right_parts[2] = '004' (sequence)
                
                first_segment = right_parts[0].strip()
                logger.debug(f"[Line] First segment (diameter+class): '{first_segment}'")
                
                # Split by space to separate diameter and class
                if ' ' in first_segment:
                    space_parts = first_segment.split(' ', 1)
                    diameter = space_parts[0].strip()
                    class_spec = space_parts[1].strip() if len(space_parts) > 1 else None
                    logger.debug(f"[Line] Split by space: diameter='{diameter}', class_spec='{class_spec}'")
                else:
                    # No space found - entire segment is diameter
                    diameter = first_segment
                    class_spec = None
                    logger.warning(f"[Line] No space found in first segment. Using entire segment as diameter: '{diameter}'")
                
                fluid = right_parts[1].strip() if len(right_parts) > 1 else None
                seq = right_parts[2].strip() if len(right_parts) > 2 else None
                
                logger.debug(f"[Line] Extracted values: diameter='{diameter}', class='{class_spec}', fluid='{fluid}', seq='{seq}'")
                
                # Assign to fields
                if diam_col and diameter:
                    fields[diam_col] = diameter
                    logger.debug(f"[Line] Assigned DIAMETER: {diameter} -> {diam_col}")
                else:
                    logger.warning(f"[Line] Could not assign DIAMETER. diam_col={diam_col}, diameter={diameter}")
                    
                if class_col and class_spec:
                    fields[class_col] = class_spec
                    logger.debug(f"[Line] Assigned CLASS: {class_spec} -> {class_col}")
                else:
                    logger.warning(f"[Line] Could not assign CLASS. class_col={class_col}, class_spec={class_spec}")
                    
                if fluid_col and fluid:
                    fields[fluid_col] = fluid
                    logger.debug(f"[Line] Assigned FLUID: {fluid} -> {fluid_col}")
                else:
                    logger.warning(f"[Line] Could not assign FLUID. fluid_col={fluid_col}, fluid={fluid}")
                    
                if seq_col and seq:
                    fields[seq_col] = seq
                    logger.debug(f"[Line] Assigned SEQUENCE: {seq} -> {seq_col}")
                else:
                    logger.warning(f"[Line] Could not assign SEQUENCE. seq_col={seq_col}, seq={seq}")
                    
            elif len(right_parts) == 4:
                # Alternative format: DIAMETER-CLASS-FLUID-SEQ (all separated by dashes)
                # Example: L7-F_4"-250DS-1-004
                diameter = right_parts[0].strip()
                class_spec = right_parts[1].strip()
                fluid = right_parts[2].strip()
                seq = right_parts[3].strip()
                
                logger.debug(f"[Line] 4-part format detected: diameter='{diameter}', class='{class_spec}', fluid='{fluid}', seq='{seq}'")
                
                if diam_col and diameter:
                    fields[diam_col] = diameter
                    logger.debug(f"[Line] Assigned DIAMETER: {diameter} -> {diam_col}")
                if class_col and class_spec:
                    fields[class_col] = class_spec
                    logger.debug(f"[Line] Assigned CLASS: {class_spec} -> {class_col}")
                if fluid_col and fluid:
                    fields[fluid_col] = fluid
                    logger.debug(f"[Line] Assigned FLUID: {fluid} -> {fluid_col}")
                if seq_col and seq:
                    fields[seq_col] = seq
                    logger.debug(f"[Line] Assigned SEQUENCE: {seq} -> {seq_col}")
            else:
                logger.warning(f"[Line] Unexpected format for TAG '{tag_no_str}'. Expected 3 or 4 parts after dash split, got {len(right_parts)}: {right_parts}")
            
            logger.debug(f"[Line] Final parsed fields for TAG '{tag_no_str}': {fields}")

        elif 'cable' in tt:
            parts = tag_no_str.split('_')
            logger.debug(f"[Cable] Parsing TAG '{tag_no_str}': parts after underscore split = {parts}")
            if len(parts) >= 2:
                asset_col = get_db_column(['asset', 'ASSET', 'Asset'])
                if asset_col:
                    fields[asset_col] = parts[0]
                right_parts = parts[1].split('-')
                if len(right_parts) >= 2:
                    prefix_col = get_db_column(['prefix', 'PREFIX', 'Prefix'])
                    seq_col = get_db_column(['sequence', 'SEQUENCE', 'SEQ'])
                    if prefix_col:
                        fields[prefix_col] = right_parts[0]
                    if seq_col:
                        fields[seq_col] = right_parts[1]
                    if len(right_parts) >= 3:
                        suffix1_col = get_db_column(['suffix1', 'SUFFIX_1', 'SUFFIX-1', 'SUFFIX1'])
                        if suffix1_col:
                            fields[suffix1_col] = right_parts[2]
                    if len(right_parts) >= 4:
                        suffix2_col = get_db_column(['suffix2', 'SUFFIX_2', 'SUFFIX-2', 'SUFFIX2'])
                        if suffix2_col:
                            fields[suffix2_col] = right_parts[3]
                            
    except Exception as e:
        logger.exception(f"Error parsing TAG_NO '{tag_no_str}': {e}")

    return fields

# ------------------- SCHEDULER -------------------
# Define EST timezone
est = pytz.timezone('US/Eastern')

# 🔹 User-defined function to sync data
def sync_tags():
    print("Running scheduled sync...")

    try:
        conn = get_snowflake_connection()
        with conn.cursor() as cursor:
            # Fetch tags inserted today (you can adjust the filter as needed)
            query = """SELECT "PIMS_DOCUMENT","STATUS","FILE_NAME","TAG_NO","DATE_UPDATED","MODIFIED_DATE"
                    FROM ENI_NL_DEV_TAGREGISTRY.DEV.TBL_EQUIPMENT_TAGS
                    WHERE COALESCE("MODIFIED_DATE", "DATE_UPDATED") 
                        BETWEEN DATEADD('hour', 5, CURRENT_DATE()) 
                            AND DATEADD('hour', 5, DATEADD('day', 1, CURRENT_DATE()))

                    UNION ALL

                    SELECT "PIMS_DOCUMENT","STATUS","FILE_NAME","TAG_NO","DATE_UPDATED","MODIFIED_DATE"
                    FROM ENI_NL_DEV_TAGREGISTRY.DEV.TBL_INSTRUMENT_TAGS
                    WHERE COALESCE("MODIFIED_DATE", "DATE_UPDATED") 
                        BETWEEN DATEADD('hour', 5, CURRENT_DATE()) 
                            AND DATEADD('hour', 5, DATEADD('day', 1, CURRENT_DATE()))

                    UNION ALL

                    SELECT "PIMS_DOCUMENT","STATUS","FILE_NAME","TAG_NO","DATE_UPDATED","MODIFIED_DATE"
                    FROM ENI_NL_DEV_TAGREGISTRY.DEV.TBL_LINE_TAGS
                    WHERE COALESCE("MODIFIED_DATE", "DATE_UPDATED") 
                        BETWEEN DATEADD('hour', 5, CURRENT_DATE()) 
                            AND DATEADD('hour', 5, DATEADD('day', 1, CURRENT_DATE()))

                    UNION ALL

                    SELECT "PIMS_DOCUMENT","STATUS","FILE_NAME","TAG_NO","DATE_UPDATED","MODIFIED_DATE"
                    FROM ENI_NL_DEV_TAGREGISTRY.DEV.TBL_CABLE_TAGS
                    WHERE COALESCE("MODIFIED_DATE", "DATE_UPDATED") 
                        BETWEEN DATEADD('hour', 5, CURRENT_DATE()) 
                            AND DATEADD('hour', 5, DATEADD('day', 1, CURRENT_DATE()));
                            
                 """
            cursor.execute(query)
            rows = cursor.fetchall()

            if not rows:
                print("No new tags found for today.")
                return {"status": "no_data", "message": "No new tags for today"}

            # Build payload for PIMS
            json_with_names = []
            for row in rows:
                pims_document, status, file_name, tag_no, date_updated, modified_date = row
                document_ids = set()  

                # Extract DocumentID from FILE_NAME ---
                if file_name:
                    parts = file_name.split("_")
                    extracted_id = "_".join(parts[:2]) if len(parts) >= 2 else file_name.strip()
                    document_ids.add(extracted_id)

                # Extract one or more DocumentIDs from PIMS_DOCUMENT ---
                if pims_document:  # make sure it's not None
                    pims_docs = [doc.strip() for doc in pims_document.split(";") if doc.strip()]
                    document_ids.update(pims_docs)

                # Skip if no document IDs found ---
                if not document_ids:
                    logger.info(f"Skipping row with missing FILE_NAME and PIMS_DOCUMENT: {row}")
                    continue

                # Add one entry per DocumentID
                for doc_id in document_ids:
                    logger.info( f"PIMS Data ---> DocumentID: {doc_id}, TagNo: {tag_no}, TagStatus: {status}" )
                    json_with_names.append({
                        "DocumentID": doc_id,
                        "TagNo": tag_no,
                        "TagStatus":status
                    })

            request_data = {
                "operation": "create",
                "resourceName": "atbv_DCS_API_DocumentsTags",
                "uniqueName": "dcs_api_documentstags",
                "excludeFieldNames": False,
                "bulk": True,
                "fields": [
                    {"name": "DocumentID"},
                    {"name": "TagNo"},
                    {"name": "TagStatus"}
                ],
                "data": json_with_names
            }

            headers = {
                "Accept": "application/json",
                "Content-Type": "application/json",
                # "Apikey":"0659415DC465E51A2C9F52BCDC9548C5AEE239F7B512FF9EC5E66D6DE4D17D0FA6D2F30AF1BD54F177EC7A0E14580ECD2AE38FCA50FD2CE43B9D13024D2FB4D0DC343458F1096964CC1FC526EE4C5DA88A7875A8E7D6D7DA8A817C27C1B896290FA27FBEA9C166FA33834AD64D14305D"
                "ApiKey": "3CC7D646A5A176204EDC8AF61E39EE64057E040848CABBED02C9AD18A80CEE428A5AF94A53357A71D1F25F5E188FF48AE322772A058614BD9E5C4005531DEEB5916E018372C9D823AC170108464C83D7FA9DA9D38EA26CE5ECCCD04C069DC462422E85CB10DE8F925AE7D031EECABF3B"
            }

            response = requests.post(
                "https://eni-dev.pimshosting.com/api/data",
                # "https://pims.infra01.net/api/data",
                headers=headers,
                data=json.dumps(request_data)
            )

            print("Sync completed:", response.status_code, response.text[:200])
            return {"status": "success", "count": len(json_with_names), "response": response.text[:200]}

    except Exception as e:
        print("Sync error:", str(e))
        return {"status": "error", "message": str(e)}
# 🔹 Scheduler (runs in background)

# Initialize scheduler with EST timezone
scheduler = BackgroundScheduler(timezone=est)

# Schedule initial job at 5:00 AM EST
scheduler.add_job(sync_tags, "cron", hour=5, minute=0)

scheduler.start()

# ------------------- API ENDPOINTS -------------------

# 1️⃣ Manually trigger sync
@app.route("/sync", methods=["POST"])
def trigger_sync():
    result = sync_tags()
    return jsonify(result)

SCHEDULE_FILE = "scheduled_time.json"
booked_times = None
 
def load_schedule():
    global booked_times
    if os.path.exists(SCHEDULE_FILE):
        with open(SCHEDULE_FILE) as f:
            data = json.load(f)
            booked_times = data.get("time")
 
def save_schedule():
    with open(SCHEDULE_FILE, "w") as f:
        json.dump({"time": booked_times}, f)
 
# load once at startup
load_schedule()
 
# Fetch current schedule
@app.route("/schedule", methods=["GET"])
def get_schedule():
    file_path = SCHEDULE_FILE
    if os.path.exists(file_path):
        with open(file_path) as f:
            data = json.load(f)
        app.logger.info(f"Loaded schedule from {file_path}: {data}")
        return jsonify({"time": data.get("time")})
    else:
        app.logger.warning(f"Schedule file not found at {file_path}")
    return jsonify({"time": None})
 
# Update schedule from frontend
@app.route("/schedule", methods=["POST"])
def update_schedule():
    global booked_times
    try:
        data = request.get_json()
        time = data.get("time")  
        if not time:
            return jsonify({"error": "time (HH:mm) is required"}), 400
       
       
        booked_times = time
        save_schedule()  
       
        # hour, minute = map(int, time.split(":"))
        # Clear existing jobs and add new one
        scheduler.remove_all_jobs()
        scheduler.add_job(sync_tags, "cron", hour=int(time.split(":")[0]), minute=int(time.split(":")[1]), timezone=est)
        print(f"✅ Scheduled job at {time} EST")
        # booked_times.append(time)
        return jsonify({"message": f"✅ New schedule set for {time} EST", "time": time})
   
    except Exception as e:
        print("❌ Scheduling error:", str(e))
        return jsonify({"error": f"Invalid time format or scheduling error: {str(e)}"}), 400

# Update TAG_NO by S_NO
@app.route('/api/Home/UpdateTagNoBySno', methods=['PUT'])
def update_tagno_by_sno():
    """Update TAG_NO and TAG_TYPE in the database based on unique column S_NO."""
    logger.info("UpdateTagNoBySno endpoint called.")
    
    model_data = request.get_json()
    if not model_data:
        return jsonify({"success": False, "message": "Invalid request body."}), 400

    s_no = model_data.get("sNo") or model_data.get("s_No")  
    new_tag_no = model_data.get("tagNo")
    new_tag_type = model_data.get("tagType") 

    if not s_no:
        return jsonify({
            "success": False,
            "message": "Request must contain 'sNo' (or 's_No')."
        }), 400

    try:
        with get_snowflake_connection() as conn:
            with conn.cursor() as cursor:
                # Fetch existing values first
                select_query = """
                    SELECT "TAG_NO", "TAGTYPE"
                    FROM ENI_NL_DEV_TAGREGISTRY.DEV.TBL_TAG_INFO
                    WHERE "S_NO" = %s
                """
                cursor.execute(select_query, (s_no,))
                result = cursor.fetchone()

                if not result:
                    return jsonify({
                        "success": False,
                        "message": f"No record found for S_NO '{s_no}'."
                    }), 404

                old_tag_no, old_tag_type = result

                # Prepare update fields dynamically
                fields_to_update = []
                values = []

                if new_tag_no and new_tag_no != old_tag_no:
                    fields_to_update.append('"TAG_NO" = %s')
                    values.append(new_tag_no)

                if new_tag_type and new_tag_type != old_tag_type:
                    fields_to_update.append('"TAGTYPE" = %s')
                    values.append(new_tag_type)

                if not fields_to_update:
                    return jsonify({
                        "success": True,
                        "message": "No changes detected.",
                        "data": {"sNo": s_no}
                    }), 200

                # Build and execute dynamic update query
                update_query = f"""
                    UPDATE ENI_NL_DEV_TAGREGISTRY.DEV.TBL_TAG_INFO
                    SET {', '.join(fields_to_update)}
                    WHERE "S_NO" = %s
                """
                values.append(s_no)
                cursor.execute(update_query, tuple(values))
                rows_affected = cursor.rowcount

        return jsonify({
            "success": True,
            "message": f"Record updated successfully for S_NO '{s_no}'.",
            "data": {
                "sNo": s_no,
                "oldTagNo": old_tag_no,
                "newTagNo": new_tag_no or old_tag_no,
                "oldTagType": old_tag_type,
                "newTagType": new_tag_type or old_tag_type,
                "rowsUpdated": rows_affected
            }
        }), 200

    except Exception as ex:
        logger.info(f"Failed to update tag info: {ex}", exc_info=True)
        return jsonify({
            "success": False,
            "message": "Internal server error",
            "details": str(ex)
        }), 500


# ------------------- BLOB PROCESSING -------------------

PROCESSED_FILE = "processed_files.txt"
MAX_RETRIES = 5

processed_blobs = set()
failed_blobs = set()
retry_counts = {}

# Azure Blob Initialization
blob_service_client = BlobServiceClient.from_connection_string(connection_string)
container_client = blob_service_client.get_container_client(container_name)
success_container_client = blob_service_client.get_container_client(success_container_name)

# Load / Save Processed Files
def load_processed_files():
    if os.path.exists(PROCESSED_FILE):
        with open(PROCESSED_FILE, "r") as f:
            return set(line.strip() for line in f if line.strip())
    return set()

def save_processed_file(blob_name):
    with open(PROCESSED_FILE, "a") as f:
        f.write(blob_name + "\n")

processed_blobs = load_processed_files()

# File Type Detection
pid_regex = re.compile(r'(?:^|_)(A40\d{2}|A41\d{2}|A45\d{2,})', re.IGNORECASE)

def is_pid_file(filename):
    return bool(pid_regex.search(os.path.basename(filename)))

def is_supported_file(filename):
    return filename.lower().endswith(('.jpg', '.jpeg', '.png', '.pdf'))

# Move Blob After Success
def move_blob_to_success_container(blob_name):
    try:
        source_blob = container_client.get_blob_client(blob_name)
        dest_blob = success_container_client.get_blob_client(blob_name)

        # Start copy
        dest_blob.start_copy_from_url(source_blob.url)
        logger.info(f" Copy started → {blob_name} moved to {success_container_name}")

        # Delete original
        source_blob.delete_blob()
        logger.info(f" Deleted from source container: {blob_name}")

    except Exception as e:
        logger.error(f" Failed to move {blob_name} to success container: {e}")

# Failure Handling
def handle_failure(blob_name):
    retry_counts[blob_name] = retry_counts.get(blob_name, 0) + 1
    failed_blobs.add(blob_name)

    if retry_counts[blob_name] >= MAX_RETRIES:
        logger.error(f" Max retries reached for {blob_name}. Skipping permanently.")
        failed_blobs.discard(blob_name)

# Blob Processing Logic
def process_blob(blob_name):
    try:
        blob_client = container_client.get_blob_client(blob_name)
        blob_bytes = blob_client.download_blob().readall()
        filename_only = os.path.basename(blob_name)

        mime_type, _ = mimetypes.guess_type(filename_only)
        mime_type = mime_type or "application/octet-stream"

        # Decide API Routing
        if is_pid_file(filename_only):
            api_url = api_url_pid
            file_type = "P&ID"

            files = {
                "image": (filename_only, BytesIO(blob_bytes), mime_type)
            }
        else:
            api_url = api_url_sd
            file_type = "STANDARD"

            files = [
                ("files", (filename_only, BytesIO(blob_bytes), mime_type))
            ]

        logger.info(f" Processing: {blob_name} | Type: {file_type}")
        logger.info(f" API Hit Succefully")

        # Send Request
        response = requests.post(api_url, files=files)

        if response.status_code == 200:
            logger.info(f" Success: {blob_name}")

            processed_blobs.add(blob_name)
            save_processed_file(blob_name)

            # MOVE FILE AFTER SUCCESS
            move_blob_to_success_container(blob_name)

            failed_blobs.discard(blob_name)
            retry_counts.pop(blob_name, None)

            return True

        else:
            logger.error(f" Failed {blob_name} | Status {response.status_code} | {response.text}")
            handle_failure(blob_name)
            return False

    except Exception as e:
        logger.error(f" Error processing {blob_name}: {e}")
        handle_failure(blob_name)
        return False

# MAIN LOOP
def blob_processing_loop():
    while True:
        logger.info("Checking for new or failed files...")

        # NEW BLOBS
        for blob in container_client.list_blobs():
            if blob.name not in processed_blobs and is_supported_file(blob.name):
                process_blob(blob.name)

        # RETRIES
        if failed_blobs:
            logger.info(f"Retrying {len(failed_blobs)} failed blobs...")
            for blob_name in list(failed_blobs):
                process_blob(blob_name)

        time.sleep(10)

# Start Background Worker
# threading.Thread(target=blob_processing_loop, daemon=True).start()

# ------------------- API ENDPOINTS FOR LINKED OBJECTS -------------------
def extract_doc_id_from_url(value):
    """
    Extracts document ID from URL or returns the value as-is if it's not a URL.
    Handles both single values and semicolon-separated values.
    """
    if not value:
        return ""
    
    import re
    
    # Split by semicolon if multiple values
    parts = value.split(';')
    extracted_ids = []
    
    for part in parts:
        part = part.strip()
        if not part:
            continue
            
        # Check if it's a URL containing DocID parameter
        if 'http' in part and 'DocID=' in part:
            match = re.search(r'DocID=([^&]+)', part)
            if match:
                extracted_ids.append(match.group(1))
            else:
                extracted_ids.append(part)
        else:
            extracted_ids.append(part)
    
    return ';'.join(extracted_ids)

# Get linked objects by tagNo and tagType
@app.route('/api/Home/GetLinkedObjects', methods=['GET'])
def get_linked_objects_endpoint():
    try:
        tag_no = request.args.get('tagNo')
        tag_type = request.args.get('tagType') 
        
        if not tag_no:
            return jsonify({"error": "tagNo parameter is required."}), 400

        conn = get_snowflake_connection()
        
        
        # if tag_type and tag_type.lower() == "cable":
        if tag_type and "cable" in tag_type.lower():
            with conn.cursor() as cursor:
                cursor.execute("""
                    SELECT "PIMS_DOCUMENT", "DOCUMENT",
                           "RELATED_TAGS"
                    FROM "ENI_NL_DEV_TAGREGISTRY"."DEV"."TBL_CABLE_TAGS"
                    WHERE "TAG_NO" = %s
                """, (tag_no,))
                result = cursor.fetchone()
            conn.close()
            if result:
                return jsonify({
                    "omegaPims": [{"pid": result[0] or ""}],"otherDocs": [{"docOfConformityNumberNotPimsDocId": result[1] or ""}],
                    "tags": [{"relatedTags": result[2] or ""}]
                }), 200
            return jsonify({"error": "No data found for the provided tagNo."}), 404
        
     
        if tag_type and tag_type.lower() == "instrument":
            with conn.cursor() as cursor:
                cursor.execute("""
                    SELECT "PIMS_DOCUMENT", "DOCUMENT", "RELATED_TAGS",
                    FROM "ENI_NL_DEV_TAGREGISTRY"."DEV"."TBL_INSTRUMENT_TAGS"
                    WHERE "TAG_NO" = %s
                """, (tag_no,))
                result = cursor.fetchone()
            conn.close()
            if result:
                return jsonify({
                    "omegaPims": [{"pid": result[0] or ""}],
                    "otherDocs": [{"docOfConformityNumberNotPimsDocId": result[1] or ""}],
                    "tags": [{"relatedTags": result[2] or ""}]
                }), 200
            return jsonify({"error": "No data found for the provided tagNo."}), 404
        
        if tag_type and tag_type.lower() == "equipment":
            with conn.cursor() as cursor:
                cursor.execute("""
                    SELECT "PIMS_DOCUMENT", "DOCUMENT","RELATED_TAGS"
                    FROM "ENI_NL_DEV_TAGREGISTRY"."DEV"."TBL_EQUIPMENT_TAGS"
                    WHERE "TAG_NO" = %s
                """, (tag_no,))
                result = cursor.fetchone()
            conn.close()
            if result:
                return jsonify({
                    "omegaPims": [{"pid": result[0] or ""}],"otherDocs": [{"docOfConformityNumberNotPimsDocId": result[1] or ""}],
                    "tags": [{"relatedTags": result[2] or ""}]
                }), 200
            return jsonify({"error": "No data found for the provided tagNo."}), 404
        
        
        if tag_type and tag_type.lower() == "line":
            with conn.cursor() as cursor:
                cursor.execute("""
                    SELECT "PIMS_DOCUMENT", "DOCUMENT",
                           "RELATED_TAGS"
                    FROM "ENI_NL_DEV_TAGREGISTRY"."DEV"."TBL_LINE_TAGS"
                    WHERE "TAG_NO" = %s
                """, (tag_no,))
                result = cursor.fetchone()
            conn.close()
            if result:
                return jsonify({
                    "omegaPims": [{"pid": result[0] or ""}],"otherDocs": [{"docOfConformityNumberNotPimsDocId": result[1] or ""}],
                    "tags": [{"relatedTags": result[2] or ""}]
                }), 200
            return jsonify({"error": "No data found for the provided tagNo."}), 404
               
        conn.close()
        return jsonify({"error": "tagType parameter is required."}), 400
        
    except Exception as ex:
        logger.error(f"Error in GetLinkedObjects: {ex}", exc_info=True)
        return jsonify({"error": "Internal server error", "details": str(ex)}), 500

# Save linked data by tagNo and tagType
@app.route('/api/Home/LinkSaveData', methods=['POST'])
def save_data_endpoint_LinkSaveData():
    """
    REPLACES data in the correct table with what frontend sends.
    Frontend manages ADD/REMOVE, backend just saves the final state.
    """
    try:
        model_data = request.get_json()
        logger.info(f"LinkSaveData called with: {model_data}")
        
        if not model_data or "tagNo" not in model_data:
            return jsonify({"success": False, "message": "tagNo is required"}), 400
            
        # tag_no = model_data["tagNo"]
        # tag_no = str(model_data["tagNo"]).strip()
        tag_no = model_data.get("tagNo", "").strip()
        tag_type = model_data.get("tagType")
        
        if not tag_type:
            return jsonify({"success": False, "message": "tagType is required"}), 400

        conn = get_snowflake_connection()

        # Helper to clean and join values
        def clean_join(value):
            """Clean semicolon-separated string, remove duplicates while preserving order"""
            if not value:
                return ""
            items = [v.strip() for v in str(value).split(';') if v.strip()]
            seen = set()
            result = []
            for item in items:
                if item not in seen:
                    seen.add(item)
                    result.append(item)
            return ';'.join(result)

        # EQUIPMENT / INSTRUMENT TAGS
        equipment_column_mapping = {
            'pid': 'PIMS_DOCUMENT',
            'docOfConformityNumberNotPimsDocId': 'DOCUMENT',
            'relatedTags': 'RELATED_TAGS',
        }
        
        instrument_column_mapping = {
            'pid': 'PIMS_DOCUMENT',
            'docOfConformityNumberNotPimsDocId': 'DOCUMENT',
            'relatedTags': 'RELATED_TAGS',
        }
        
        line_column_mapping = {
            'pid': 'PIMS_DOCUMENT',
            'docOfConformityNumberNotPimsDocId': 'DOCUMENT',
            'relatedTags': 'RELATED_TAGS',
        }
        
        cable_column_mapping = {
            'pid': 'PIMS_DOCUMENT',
            'docOfConformityNumberNotPimsDocId': 'DOCUMENT',
            'relatedTags': 'RELATED_TAGS',
        }
        

        if tag_type.lower() == "equipment":
            table_name = '"ENI_NL_DEV_TAGREGISTRY"."DEV"."TBL_EQUIPMENT_TAGS"'
            column_mapping = equipment_column_mapping
        elif tag_type.lower() == "instrument":
            table_name = '"ENI_NL_DEV_TAGREGISTRY"."DEV"."TBL_INSTRUMENT_TAGS"'
            column_mapping = instrument_column_mapping
        elif tag_type.lower() == "line":
            table_name = '"ENI_NL_DEV_TAGREGISTRY"."DEV"."TBL_LINE_TAGS"'
            column_mapping = line_column_mapping
        elif tag_type.lower() in ["cable", "cable-electrical", "cable-instrument"]:
            table_name = '"ENI_NL_DEV_TAGREGISTRY"."DEV"."TBL_CABLE_TAGS"'
            column_mapping = cable_column_mapping   
        else:
            conn.close()
            return jsonify({"success": False, "message": "Unsupported tagType"}), 400

        # Fetch existing DB row
        select_cols = ', '.join(f'"{col}"' for col in column_mapping.values())
        with conn.cursor() as cursor:
            cursor.execute(f"""
                SELECT {select_cols}
                FROM {table_name}
                WHERE "TAG_NO" = %s
            """, (tag_no,))
            existing = cursor.fetchone()

        if not existing:
            conn.close()
            return jsonify({"success": False, "message": "Tag not found"}), 404

        existing_data = {list(column_mapping.keys())[i]: (existing[i] if existing[i] else "") for i in range(len(column_mapping))}

        # Prepare updates - REPLACE with frontend values
        set_clauses = []
        values = []
        for key, col in column_mapping.items():
            if key in model_data:
                new_value = model_data[key]
                
                # Extract doc ID for PID fields
                if key == 'pid':
                    new_value = extract_doc_id_from_url(new_value)
                
                # Clean and deduplicate
                final_value = clean_join(new_value)
                
                logger.info(f"{tag_type} tag {tag_no} - {key}: Existing=[{existing_data.get(key, '')}], New=[{final_value}]")
                
                set_clauses.append(f'"{col}" = %s')
                values.append(final_value)
        
        now = get_est_now()
        set_clauses.append('"MODIFIED_DATE" = %s')
        values.append(now)


        if not set_clauses:
            conn.close()
            return jsonify({"success": False, "message": "No valid fields to update"}), 400

        values.append(tag_no)
        update_query = f"""
            UPDATE {table_name}
            SET {', '.join(set_clauses)}
            WHERE "TAG_NO" = %s
        """
        
        with conn.cursor() as cursor:
            cursor.execute(update_query, tuple(values))
            rows_updated = cursor.rowcount
            conn.commit()
        
        logger.info(f"Updated {rows_updated} row(s) for {tag_type} tag: {tag_no}")
        conn.close()

        return jsonify({"success": True, "message": f"{tag_type.capitalize()} tag updated successfully"}), 200

    except Exception as ex:
        logger.error(f"Error in LinkSaveData: {ex}", exc_info=True)
        return jsonify({"success": False, "message": str(ex)}), 500

# Generate Direct Line Token for Copilot/Chatbot
@app.route("/api/directline/token", methods=["POST"])
def get_directline_token():
    try:
        auth_header = request.headers.get("Authorization")
        if not auth_header:
            return jsonify({"error": "Missing Authorization header"}), 401
 
        url = f"https://{COPILOT_REGION}.directline.botframework.com/v3/directline/tokens/generate"
        headers = {
            "Authorization": f"Bearer {DIRECT_LINE_SECRET}",
            "Content-Type": "application/json"
        }
        payload = {"user": {"id": "user-001"}}
 
        response = requests.post(url, headers=headers, json=payload)
        response.raise_for_status()
 
        print("✅ Direct Line token generated")
        return jsonify(response.json())
 
    except Exception as e:
        print("❌ Token error:", e)
        return jsonify({"error": str(e)}), 500

EQUIPMENT_EXCEL_HEADERS = {
    "STATUS": "Status","TYPE": "Tag Type","TAG_NO": "Tag No.","ASSET": "Asset","FUNCTION": "Function",
    "FUNCTION_DESCRIPTION": "Function Description","SEQUENCE": "Sequence","SUFFIX_1": "Suffix-1","SUFFIX_2": "Suffix-2",
    "DESCRIPTION": "Description","SPECIFICATION": "Specification","PED_GRP_CAT_MOD": "P.E.D. Grp / CAT / Mod","SUPPLIER": "Supplier","FLOW_RATE_NM3_HR": "Flow Rate (Nm3 / hr)",
    "CAPACITY_M3": "Capacity (m3)","DUTY_KW": "Duty (kW)","DESIGN_PRESSURE_BARG": "Design Pressure (barg)","TEST_PRESS_BARG": "Test Pressure.(barg)",
    "DESIGN_TEMPERATURE_DEGC": "Design Temperature (Deg.C)","MEDIUM": "Medium","POWER_KW": "Power (kW)","PIMS_DOCUMENT": "PIMS Document",
    "DOCUMENT": "Document","RELATED_TAGS": "Related Tags","LENGTH_T_T_MM": "Length T / T (mm)","WIDTH_MM": "Width (mm)",
    "HEIGHT_MM": "Height (mm)","DIA_ID_MM": "Dia ID (mm)","DRY_WT_TE": "Dry Wt. (Te)","OPERATIONAL_WEIGHT_TE": "Operational Weight (Te)", "REMARKS": "Remarks"
}

INSTRUMENT_EXCEL_HEADERS = {
    "STATUS": "Status","TYPE": "Tag Type","TAG_NO": "Tag No.","ASSET": "Asset","FUNCTION": "Function",
    "INSTRUMENT_TYPE_FUNCTION_DESCRIPTION": "Instrument Type (Function Description)","SEQUENCE": "Sequence","SUFFIX_1": "Suffix-1","SUFFIX_2": "Suffix-2","DESCRIPTION": "Description",
    "LOOP_ID": "Loop ID","DEVICE_TYPE": "Device Type","ATEX": "ATEX","DECK_MODULE_ROOM": "Deck / Module / Room","PACKAGE_PANEL": "Package / Panel","ICSS_SYSTEM": "ICSS System","ICSS_CABINET": "ICSS Cabinet","ICSS_IO_TYPE": "ICSS I/O Type","ICSS_SIGNAL_TYPE": "ICSS Signal Type","PIMS_DOCUMENT": "PIMS Document","CE_DIAGRAM": "CE Diagram","LOCATION_DRAWING": "Location Drawing",
    "LAYOUT": "Layout","LOOP_DRAWING": "Loop Drawing","DATA_SHEET": "Data Sheet","HOOKUP_DRAWING_1": "Hookup Drawing #1","HOOKUP_DRAWING_2": "Hookup Drawing #2","SPECIFICATION": "Specification","MAXIMO_WO": "Maximo WO",
    "MANUFACTURER": "Manufacturer","MODEL": "Model","CALIBRATED_RANGE": "Calibrated Range","UNIT": "Unit","TRIP_LL": "Trip LL","DOCUMENT": "Document","RELATED_TAGS": "Related Tags","ALARM_L": "Alarm L","CTR_L": "Control L",
    "CTR_N": "Control N","CTR_H": "Control H","ALARM_H": "Alarm H","TRIP_HH": "Trip HH","OUTPUT_RANGE_MA": "Output Range (mA)","OUTPUT_SET_POINT_L_MA": "Output Set Point L (mA)","OUTPUT_SET_POINT_H_MA": "Output Set Point H (mA)","REMARKS": "Remarks"
}

LINE_EXCEL_HEADERS = {
    "STATUS": "Status","TAG_NO": "Line No.","TAG_TYPE": "Tag Type","ASSET": "Asset","DIAMETER": "Diameter","CLASS_SPECIFICATION": "Class (Specification)",
    "FLUID_CODE_SERVICE": "Fluid Code (Service)","SEQUENCE": "Sequence","INSULATION": "Insulation","INSULATION_THICKNESS_MM": "Insulation Thickness (mm)","FROM_LOCATION": "From Location","TO_LOCATION": "To Location",
    "OPERATION_PHASE": "Operation - Phase","OPERATION_PRESSURE_BARG": "Operation - Pressure (barg)","OPERATION_TEMPERATURE_DEGC": "Operation - Temperature (Deg.C)","ASSESSMENT_VAP_PRESSURE_BARG": "Assessment - Vapour Pressure (barg)",
    "ASSESSMENT_MIN_DES_TEMP_DEGC": "Assessment - Min Design Temp (Deg.C)","ASSESSMENT_MAX_DES_TEMP_DEGC": "Assessment - Max Design Temp (Deg.C)","ASSESSMENT_DES_PRESS_BARG": "Assessment - Design Pressure (barg)","PIPE_WT_OR_SCH": "Pipe WT Or SCH","PED_CRITICAL": "PED Critical","HAZARDOUS_CATEGORY": "Hazardous Category",
    "PED_TABLE": "PED Table","PED_CAT": "PED Category","PED_MOD": "PED Module","PED_GROUP": "PED Group","KVI_TABLE": "KVI Table","KVL_Y_N": "KVL (Y/N)",
    "TEST_PRESSURE": "Test Pressure","INSPECTION_NUMBER": "Inspection Number","PIMS_DOCUMENT": "PIMS Document","PID_2": "PID 2","PID_3": "PID 3","DOCUMENT": "Document","RELATED_TAGS": "Related Tags","REMARKS": "Remarks"
}

CABLE_EXCEL_HEADERS = {
    "STATUS": "Status","TAG_NO": "Cable No.","TAG_TYPE": "Tag Type","ASSET": "Asset","PREFIX": "Prefix","SEQUENCE": "Sequence","SUFFIX_1": "Suffix-1","SUFFIX_2": "Suffix-2","DESCRIPTION": "Description",
    "FROM_LOCATION": "From Location","TO_LOCATION": "To Location","DOCUMENT": "Document","RELATED_TAGS": "Related Tags","SIZE": "Size","COLOUR": "Colour","LENGTH": "Length","PIMS_DOCUMENT": "PIMS Document","INSTRUMENT_DRAWING_1": "Instrument Drawing #1","INSTRUMENT_DRAWING_2": "Instrument Drawing #2","NRPACA": "nrpaca","NRCOCA": "nrcoca","REMARKS": "Remarks"
}

EXCEL_HEADER_MAP = {
    "Equipment": EQUIPMENT_EXCEL_HEADERS,
    "Instrument": INSTRUMENT_EXCEL_HEADERS,
    "Line": LINE_EXCEL_HEADERS,
    "Cable": CABLE_EXCEL_HEADERS,
    "Cable - Electrical": CABLE_EXCEL_HEADERS,
    "Cable - Instrument": CABLE_EXCEL_HEADERS
}

@app.route('/api/Home/ExportToExcel', methods=['POST'])
def export_to_excel_endpoint():
    logger.info("ExportToExcel endpoint called.")
    
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "Invalid request body"}), 400

        # Handle tagType for all tag categories
        tag_type = data.get("tagType") or data.get("type")
        print(  f" ExportToExcel → tag_type received: {tag_type} "  )
        tag_type = (tag_type or "").strip().title()

        if tag_type not in TABLE_MAPPING:
            return jsonify({"error": f"Invalid or missing tagType. Received: '{tag_type}'"}), 400

        filtered_data = data.get("data", [])
        if not isinstance(filtered_data, list) or not filtered_data:
            return jsonify({"error": "No records available for export"}), 400

        # Backend mapping
        column_mapping = TABLE_MAPPING[tag_type]["mapping"]

        # Excluded columns
        EXCLUDED_COLUMNS_BY_TYPE = {
                "Equipment": {"file_Type", "url", "file_Name", "file_Description",
                            "relatedTags1","relatedTags2","relatedTags3",
                            "dateUpdated", "modifiedBy", "modifiedDate", "maximoWo"},

                "Instrument": {"file_Type", "url", "file_Name", "file_Description",
                            "relatedTags1","relatedTags2","relatedTags3",
                            "dateUpdated", "modifiedBy", "modifiedDate"},

                "Line": {"file_Type", "url", "file_Name", "file_Description",
                        "relatedTags1","relatedTags2","relatedTags3","pid1","pid2","pid3",
                        "dateUpdated", "modifiedBy", "modifiedDate", "maximoWo"},

                "Cable": {"file_Type", "url", "file_Name", "file_Description",
                        "relatedTags1","relatedTags2","relatedTags3",
                        "dateUpdated", "modifiedBy", "modifiedDate", "maximoWo"}
            }
    
        EXCLUDED = EXCLUDED_COLUMNS_BY_TYPE.get(tag_type, set())
        column_mapping = {k: v for k, v in column_mapping.items() if k not in EXCLUDED}

        # Normalize data
        normalized_data = [
            {db_col: record.get(src_key, "") for src_key, db_col in column_mapping.items()}
            for record in filtered_data
        ]

        # DataFrame
        df = pd.DataFrame(normalized_data)
        df = df.reindex(columns=list(column_mapping.values()))

        # Apply EXCEL UI header mapping
        header_map = EXCEL_HEADER_MAP[tag_type]
        df.columns = [header_map.get(col, col) for col in df.columns]

        # Build filename
        filename = f"{tag_type} Tag Register {datetime.utcnow().strftime('%d-%m-%Y')}.xlsx"

        # Excel file creation
        output = BytesIO()
        from openpyxl.styles import Font, Border, Side, PatternFill

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="TagRegister")
            ws = writer.sheets["TagRegister"]
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill

            for col_cells in ws.columns:
                col_letter = col_cells[0].column_letter
                max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

            ws.freeze_panes = "A2"
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border

        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )

    except Exception as ex:
        logger.error(f"Excel export failed: {ex}", exc_info=True)
        return jsonify({"error": "Internal server error", "details": str(ex)}), 500

# ------------------- SECURITY HEADERS -------------------
@app.after_request
def add_security_headers(response):
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Content-Security-Policy"] = "frame-ancestors 'none';"
    response.headers["Referrer-Policy"] = "no-referrer"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response

# ------------------- APP RUN -------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, threaded=True)
    logger.info("This is a debug message for app run start")
    


